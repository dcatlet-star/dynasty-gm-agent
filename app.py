from flask import Flask, request, jsonify, send_from_directory
import anthropic
from collections import defaultdict
import os
import re
import json
import sqlite3
from datetime import datetime, timedelta
import requests
import random

DB_PATH = os.environ.get('DB_PATH', '/data/dynasty.db')
if not os.path.exists('/data'):
    DB_PATH = 'dynasty.db'

app = Flask(__name__, static_folder='static')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024
client = anthropic.Anthropic(api_key=os.environ.get("ANTHROPIC_API_KEY"))

SLEEPER_LEAGUE_IDS = {
    "gentlemans_dynasty": "1314472610167279616",
    "velvet_spade":       "1315445968161734656"
}

def init_db():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS chat_history
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, week_key TEXT, role TEXT, content TEXT, timestamp TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS player_rankings
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, player_name TEXT UNIQUE, position TEXT, team TEXT,
                  elo_score REAL DEFAULT 1500, comparisons INTEGER DEFAULT 0, last_updated TEXT,
                  ktc_tier INTEGER DEFAULT 10)''')
    c.execute('''CREATE TABLE IF NOT EXISTS player_profiles
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, player_name TEXT UNIQUE, profile_data TEXT, last_updated TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS player_values
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, player_name TEXT UNIQUE,
                  position TEXT, my_value INTEGER, ktc_value INTEGER, delta INTEGER,
                  tier TEXT, last_updated TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS league_rosters
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, league TEXT, manager TEXT,
                  player_name TEXT, position TEXT, team TEXT, last_updated TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS gm_tendencies
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, manager TEXT, league TEXT,
                  window TEXT, style TEXT, loves_buy TEXT, loves_sell TEXT,
                  never_sells TEXT, overpays_for TEXT, trade_notes TEXT,
                  last_updated TEXT, UNIQUE(manager, league))''')
    c.execute('''CREATE TABLE IF NOT EXISTS trade_log
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, league TEXT, date TEXT,
                  their_manager TEXT, what_they_want TEXT, what_i_want TEXT,
                  status TEXT, my_next_move TEXT, notes TEXT, last_updated TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS pick_capital
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, league TEXT, year TEXT,
                  round TEXT, original_owner TEXT, current_owner TEXT,
                  label TEXT, last_updated TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS kb_meta
                 (id INTEGER PRIMARY KEY, last_upload TEXT, filename TEXT,
                  tabs_parsed TEXT, row_counts TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS dashboard_cache
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, cache_key TEXT UNIQUE, data TEXT, last_updated TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS vs_rankings
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, player_name TEXT UNIQUE, position TEXT, team TEXT,
                  elo_score REAL DEFAULT 1500, comparisons INTEGER DEFAULT 0, last_updated TEXT,
                  ktc_tier INTEGER DEFAULT 10)''')
    c.execute('''CREATE TABLE IF NOT EXISTS market_data
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, source TEXT NOT NULL, player_name TEXT NOT NULL,
                  rank INTEGER, value INTEGER, position TEXT, team TEXT, updated_at TEXT NOT NULL,
                  UNIQUE(source, player_name))''')
    # Draft activity: picks made, trades, opponent tendencies
    c.execute('''CREATE TABLE IF NOT EXISTS draft_activity
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, draft_id TEXT DEFAULT 'vs_2026',
                  pick_slot TEXT, player_name TEXT, manager TEXT, position TEXT,
                  logged_at TEXT, source TEXT DEFAULT 'manual')''')
    # Manager profiles for trade targeting
    c.execute('''CREATE TABLE IF NOT EXISTS manager_profiles
                 (id INTEGER PRIMARY KEY AUTOINCREMENT, manager TEXT UNIQUE,
                  roster_data TEXT, trade_activity TEXT, tendencies TEXT, updated_at TEXT)''')
    # Alter existing tables to add ktc_tier if missing
    try:
        c.execute("ALTER TABLE player_rankings ADD COLUMN ktc_tier INTEGER DEFAULT 10")
    except: pass
    try:
        c.execute("ALTER TABLE vs_rankings ADD COLUMN ktc_tier INTEGER DEFAULT 10")
    except: pass
    conn.commit()
    conn.close()

init_db()

def get_week_key():
    now = datetime.now()
    week_start = now - timedelta(days=now.weekday())
    return week_start.strftime("%Y-W%U")

LEAGUE_CONTEXT = """
LEAGUE 1: CAPITAL GAINS — FFPC #430
$250 | 12-team SuperFlex TE Premium | 28 rounds | 29+taxi | FAAB $1000
Strategy: ACTIVE REBUILD → 2027 | Priority: #3
UNTOUCHABLES: Drake Maye, Drake London | CORE: LaPorta, Burden, MHJ, Tate
MOVEABLE: Milroe, all RBs, depth WRs/TEs
{CG_ROSTER_BLOCK}
{CG_PICKS_BLOCK}
CONSOLATION: Tank Wks 1-13 → WIN consolation → 1.01 pick 2027
SEPTEMBER CUTS: Neal, Vidal, Wright, Sampson, Noel, Douglas, Coker, Okonkwo, Milroe
INTEL: BostonBlackMambas(Allen/CMC/Saquon-LAST DANCE) | SeizeTheGrey(Burrow/Herbert-CONTENDER) | GNAwin0(Lawrence/Jeanty-MID) | Blunderbuss(Mahomes/Swift-CONTENDER) | TeddySalad(Daniels-MID) | LegendsNeverDie(REBUILDING R6/R7 only) | MayanFactors(Lamar/Goff-CONTENDER) | RiskItBrisket(Hurts/Bijan-CONTENDER) | ShootTheGlass(Allen/Mahomes/Barkley-ELITE)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

LEAGUE 2: TWENTY RUN SAVAGES — FFPC #210
$100 | 12-team SuperFlex TE Premium + K+DST | 20+3IR | FAAB $1000
Strategy: COMPETING NOW | Priority: #2 | K+DST ALWAYS REQUIRED
UNTOUCHABLES: Drake Maye, Bijan Robinson | CORE: Loveland, JSN, Lemon
{TRS_ROSTER_BLOCK}
2027 PICKS: R1(Stinky), R1(own), R2(own), R3-R7
INTEL: ShootTheGlass(ELITE) | BoulderFreeZone(Lamar/CMC-STRONG) | EvilEmpire(C.Williams/Gibbs-CONTENDER) | Settler22$(Judkins/Tate-STRONG) | Stinky(owes dcatlet 2027R1) | NuclearOptions(FULL REBUILD) | H2OSONDC(Daniels/Prescott-STRONG)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

LEAGUE 3: GENTLEMAN'S DYNASTY — SLEEPER FREE
14-team | SuperFlex | TE Premium 1.5 | 4pt TDs | K+DST | 23+3IR+4taxi
Strategy: REBUILD 2027-28 | Priority: #4 | Sleeper ID: 1314472610167279616
UNTOUCHABLES: Mahomes, Bowers | CORE: Judkins, MHJ, Tate
MAX PF WARNING: Taxi squad points COUNT toward Max PF seeding.
{GL_ROSTER_BLOCK}
2027: 1st(own), 2nd(Stiller29), 2nd(own), 4th | 2028: 1st, 2nd, 3rd, 4th
GM INTEL: c1smith11(LaPorta on block-TARGET) | McGido(Desperate TE-sell Njoku) | Goooz(Desperate TE) | SenorHyde(Desperate QB-sell McCarthy/Richardson) | mstan16(Desperate RB-sell Kamara/Etienne)

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

LEAGUE 4: VELVET SPADE — SLEEPER $250 startup (DRAFT COMPLETE)
12-team | SuperFlex | TE Premium 1.5 | 6PT PASSING TDs | 23+5taxi+2IR
FAAB $1000 | Waivers Wed 3AM ET | Trade deadline Wk13
Draft completed May 2026 | Sleeper ID: 1315445968161734656 | Priority: #1
SCORING: All TDs 6pt | Pass 0.04/yd | Rec RB/WR 1.0 TE 1.5
STATUS: TRADE SEASON — draft complete, focus on trades only
{VS_ROSTER_BLOCK}
{VS_PICKS_BLOCK}
NOTE: Capital Gains is a DIFFERENT FFPC league — never confuse with Velvet Spade.
MANAGERS: pdwyer13 | yerkdog | jefisk24 | Smohr609 | ColeTrain8300 | DrTrollPhD | coinball | EazyDakar | jakemills69 | NateSneller | sneller
WINDOW: 2027-2028 | Drake Maye untouchable | See PICKS OWNED below for current pick capital
PRIORITIES: Add proven RB2/RB3 | Add veteran WR | Use pick capital as trade leverage

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

KTC BASELINE May 2026: Maye 9500|Bowers 8100|Bijan 8200|JSN 7800|London 6500|Loveland 6200|LaPorta 5800|Mahomes 5600|MHJ 5200|Tate 4567|Dart 4500|Judkins 3500|Downs 3800
"""

SYSTEM_PROMPT = f"""You are an elite dynasty fantasy football Assistant GM for MJBrutus. You manage 4 leagues with different strategies, scoring systems, and timelines. Deep expertise in dynasty formats, startup drafts, trade theory, and roster construction.

{LEAGUE_CONTEXT}

CORE BEHAVIORS:
1. MY ROSTER DATA IS IN THIS SYSTEM PROMPT — when asked about my rosters, read from the ROSTER blocks above. Do NOT web search for my roster. Do NOT ask me to provide it. The data is already here.
2. ALWAYS web search for current player NFL situation before making any claim about stats, injuries, depth chart, role, or team. Cite the source and date for every NFL fact stated. If unverifiable, flag ⚠ UNVERIFIED — never present estimated or inferred stats as fact.
3. KTC SuperFlex+TE primary. Cross-reference RosterAudit, FantasyPros, Rotoballer, ESPN, Underdog
4. One clear decisive recommendation — not a menu
5. Always identify which league and strategy phase applies — never cross-reference leagues
6. Factor scoring differences — especially 6pt TDs in Velvet Spade
7. Trade messages: under 20 words, 2 sentences max, direct and confident
8. Gentleman's: monitor Max PF taxi implications weekly
9. TRS: confirm K+DST always rostered | Competing window — prioritize players who help WIN NOW
10. Velvet Spade: DRAFT COMPLETE — trade season only | Priority #1 league
11. Capital Gains: always factor two-phase consolation strategy

TRADE PROPOSAL RULES — EVERY TRADE MUST INCLUDE:
a) Asset values for BOTH sides (MY value + KTC) with package discount applied
b) WHY THE OTHER MANAGER WOULD ACCEPT — their roster need, competitive window, what they gain
c) WHY THIS FITS MY STRATEGY — which of my league-specific goals does this advance
d) OTHER SIDE TEST — confirm the deal is realistic from their perspective before presenting it
e) ONE Sleeper message: under 20 words, addresses what THEY want
If a proposed trade fails any of these five checks, revise it or don't present it.

WEEKLY BRIEFING FORMAT — use this exact structure when asked for a briefing:
Lead with VS (Priority #1), then TRS, CG, GL.
For each league:
• STRENGTHS (2 lines max)
• WEAKNESSES (2 lines max)  
• TOP 3 TRADE CHIPS: player | MY value | why sell now
• BEST TRADE PARTNER: one manager | their need | my leverage
• THIS WEEK'S TRADE: assets + values both sides + package discount + why they accept + Sleeper message
Keep each league to one mobile screen. Total briefing = 4 screens max.

RESPONSE FORMAT: Lead with recommendation | Cite source and date for every NFL stat/fact | Apply trade rules | One specific action item | Mobile-friendly

MY PERSONAL PLAYER VALUES (composite: 45% personal rank + 30% RA + 25% KTC, 6pt QB boost applied):
Scale: 0-10000 matching KTC. MY = my composite value. KTC = market. Δ = difference (+ means I value higher).
★ = significant divergence >500pts from KTC market.

Josh Allen (QB) | MY:9999 KTC:9999 Δ:+0 | T1 — Untouchable
Bijan Robinson (RB) | MY:9999 KTC:9991 Δ:+8 | T1 — Untouchable
Ja'Marr Chase (WR) | MY:9991 KTC:9999 Δ:-8 | T1 — Untouchable
Jaxon Smith-Njigba (WR) | MY:9889 KTC:9889 Δ:+0 | T1 — Untouchable
Jahmyr Gibbs (RB) | MY:9553 KTC:9553 Δ:+0 | T1 — Untouchable
Drake Maye (QB) | MY:9416 KTC:9416 Δ:+0 | T1 — Untouchable
Puka Nacua (WR) | MY:8794 KTC:8631 Δ:+163 | T2 — Elite
Brock Bowers (TE) | MY:8631 KTC:8794 Δ:-163 | T2 — Elite
Amon-Ra St. Brown (WR) | MY:8330 KTC:7723 Δ:+607 | T2 — Elite ★
Jayden Daniels (QB) | MY:7943 KTC:7814 Δ:+129 | T2 — Elite
Caleb Williams (QB) | MY:7858 KTC:7943 Δ:-85 | T2 — Elite
Lamar Jackson (QB) | MY:7814 KTC:7676 Δ:+138 | T2 — Elite
Malik Nabers (WR) | MY:7806 KTC:7858 Δ:-52 | T3 — Franchise
Justin Jefferson (WR) | MY:7723 KTC:7806 Δ:-83 | T3 — Franchise
Joe Burrow (QB) | MY:7676 KTC:7494 Δ:+182 | T3 — Franchise
Ashton Jeanty (RB) | MY:7650 KTC:7561 Δ:+89 | T3 — Franchise
Jeremiyah Love (RB) | MY:7561 KTC:7650 Δ:-89 | T3 — Franchise
CeeDee Lamb (WR) | MY:7494 KTC:7310 Δ:+184 | T3 — Franchise
Trey McBride (TE) | MY:7310 KTC:8330 Δ:-1020 | T3 — Franchise ★
Justin Herbert (QB) | MY:6879 KTC:6815 Δ:+64 | T3 — Franchise
DeVon Achane (RB) | MY:6867 KTC:6867 Δ:+0 | T3 — Franchise
Omarion Hampton (RB) | MY:6815 KTC:6711 Δ:+104 | T3 — Franchise
Colston Loveland (TE) | MY:6809 KTC:6689 Δ:+120 | T3 — Franchise
Drake London (WR) | MY:6711 KTC:6879 Δ:-168 | T3 — Franchise
Patrick Mahomes (QB) | MY:6689 KTC:6809 Δ:-120 | T4 — Premium
Jaxson Dart (QB) | MY:6660 KTC:6660 Δ:+0 | T4 — Premium
Jonathan Taylor (RB) | MY:6557 KTC:6063 Δ:+494 | T4 — Premium
Jalen Hurts (QB) | MY:6332 KTC:6191 Δ:+141 | T4 — Premium
Tetairoa McMillan (WR) | MY:6222 KTC:6557 Δ:-335 | T4 — Premium
Bo Nix (QB) | MY:6204 KTC:6142 Δ:+62 | T4 — Premium
Tyler Warren (TE) | MY:6191 KTC:6332 Δ:-141 | T4 — Premium
Brock Purdy (QB) | MY:6142 KTC:5815 Δ:+327 | T4 — Premium
Emeka Egbuka (WR) | MY:6063 KTC:6204 Δ:-141 | T4 — Premium
James Cook (RB) | MY:6046 KTC:6046 Δ:+0 | T4 — Premium
Carnell Tate (WR) | MY:5985 KTC:5833 Δ:+152 | T4 — Premium
Breece Hall (RB) | MY:5833 KTC:5384 Δ:+449 | T4 — Premium
George Pickens (WR) | MY:5815 KTC:5985 Δ:-170 | T4 — Premium
Fernando Mendoza (QB) | MY:5712 KTC:5663 Δ:+49 | T4 — Premium
Garrett Wilson (WR) | MY:5678 KTC:5712 Δ:-34 | T4 — Premium
Nico Collins (WR) | MY:5663 KTC:5678 Δ:-15 | T4 — Premium
Chris Olave (WR) | MY:5630 KTC:5587 Δ:+43 | T4 — Premium
TreVeyon Henderson (RB) | MY:5587 KTC:5351 Δ:+236 | T4 — Premium
Trevor Lawrence (QB) | MY:5460 KTC:6222 Δ:-762 | T5 — Strong ★
Christian McCaffrey (RB) | MY:5447 KTC:5051 Δ:+396 | T5 — Strong
Kenneth Walker III (RB) | MY:5440 KTC:5447 Δ:-7 | T5 — Strong
Quinshon Judkins (RB) | MY:5392 KTC:5460 Δ:-68 | T5 — Strong
Jordan Love (QB) | MY:5384 KTC:5630 Δ:-246 | T5 — Strong
Harold Fannin (TE) | MY:5351 KTC:5440 Δ:-89 | T5 — Strong
Ladd McConkey (WR) | MY:5348 KTC:5318 Δ:+30 | T5 — Strong
Rashee Rice (WR) | MY:5318 KTC:5301 Δ:+17 | T5 — Strong
Luther Burden (WR) | MY:5304 KTC:5216 Δ:+88 | T5 — Strong
Cam Ward (QB) | MY:5301 KTC:5245 Δ:+56 | T5 — Strong
Jordyn Tyson (WR) | MY:5245 KTC:5304 Δ:-59 | T5 — Strong
A.J. Brown (WR) | MY:5216 KTC:4808 Δ:+408 | T5 — Strong
Makai Lemon (WR) | MY:5211 KTC:5055 Δ:+156 | T5 — Strong
Rome Odunze (WR) | MY:5089 KTC:5392 Δ:-303 | T5 — Strong
Tucker Kraft (TE) | MY:5055 KTC:5348 Δ:-293 | T5 — Strong
Saquon Barkley (RB) | MY:5054 KTC:4876 Δ:+178 | T5 — Strong
Marvin Harrison Jr. (WR) | MY:5051 KTC:5054 Δ:-3 | T5 — Strong
Baker Mayfield (QB) | MY:4993 KTC:4758 Δ:+235 | T5 — Strong
Dak Prescott (QB) | MY:4934 KTC:4993 Δ:-59 | T5 — Strong
DeVonta Smith (WR) | MY:4927 KTC:5211 Δ:-284 | T5 — Strong
Chase Brown (RB) | MY:4911 KTC:4882 Δ:+29 | T5 — Strong
Sam LaPorta (TE) | MY:4908 KTC:5089 Δ:-181 | T5 — Strong
Tyler Shough (QB) | MY:4884 KTC:4884 Δ:+0 | T5 — Strong
Jadarian Price (RB) | MY:4882 KTC:4742 Δ:+140 | T5 — Strong
Tee Higgins (WR) | MY:4876 KTC:4911 Δ:-35 | T5 — Strong
Bucky Irving (RB) | MY:4870 KTC:4908 Δ:-38 | T5 — Strong
Brian Thomas Jr. (WR) | MY:4868 KTC:4934 Δ:-66 | T5 — Strong
Jaylen Waddle (WR) | MY:4858 KTC:4870 Δ:-12 | T5 — Strong
Zay Flowers (WR) | MY:4808 KTC:4670 Δ:+138 | T6 — Solid
Jared Goff (QB) | MY:4781 KTC:4591 Δ:+190 | T6 — Solid
Kyren Williams (RB) | MY:4758 KTC:4858 Δ:-100 | T6 — Solid
Cam Skattebo (RB) | MY:4742 KTC:4586 Δ:+156 | T6 — Solid
C.J. Stroud (QB) | MY:4670 KTC:4868 Δ:-198 | T6 — Solid
Kyle Pitts (TE) | MY:4662 KTC:4927 Δ:-265 | T6 — Solid
Jameson Williams (WR) | MY:4652 KTC:4662 Δ:-10 | T6 — Solid
Josh Jacobs (RB) | MY:4591 KTC:4457 Δ:+134 | T6 — Solid
KC Concepcion (WR) | MY:4586 KTC:4516 Δ:+70 | T6 — Solid
Travis Etienne (RB) | MY:4516 KTC:4403 Δ:+113 | T6 — Solid
Kenyon Sadiq (TE) | MY:4505 KTC:4652 Δ:-147 | T6 — Solid
Sam Darnold (QB) | MY:4457 KTC:4781 Δ:-324 | T6 — Solid
Javonte Williams (RB) | MY:4434 KTC:4434 Δ:+0 | T6 — Solid
Kyler Murray (QB) | MY:4403 KTC:4101 Δ:+302 | T6 — Solid
Alec Pierce (WR) | MY:4111 KTC:4111 Δ:+0 | T6 — Solid
Bhayshul Tuten (RB) | MY:4101 KTC:3951 Δ:+150 | T6 — Solid
Bryce Young (QB) | MY:4099 KTC:4505 Δ:-406 | T6 — Solid
Matthew Stafford (QB) | MY:4007 KTC:3661 Δ:+346 | T6 — Solid
Derrick Henry (RB) | MY:3998 KTC:3834 Δ:+164 | T6 — Solid
Omar Cooper (WR) | MY:3973 KTC:3896 Δ:+77 | T6 — Solid
Malik Willis (QB) | MY:3959 KTC:4007 Δ:-48 | T6 — Solid
Daniel Jones (QB) | MY:3951 KTC:3959 Δ:-8 | T6 — Solid
Eli Stowers (TE) | MY:3896 KTC:3998 Δ:-102 | T6 — Solid
George Kittle (TE) | MY:3834 KTC:3708 Δ:+126 | T6 — Solid
Parker Washington (WR) | MY:3745 KTC:3521 Δ:+224 | T6 — Solid
RJ Harvey (RB) | MY:3740 KTC:3498 Δ:+242 | T6 — Solid
David Montgomery (RB) | MY:3724 KTC:3408 Δ:+316 | T6 — Solid
Jordan Addison (WR) | MY:3716 KTC:3973 Δ:-257 | T6 — Solid
D.J. Moore (WR) | MY:3708 KTC:3678 Δ:+30 | T6 — Solid
D'Andre Swift (RB) | MY:3678 KTC:3505 Δ:+173 | T6 — Solid
Michael Wilson (WR) | MY:3661 KTC:3656 Δ:+5 | T6 — Solid
Matthew Golden (WR) | MY:3656 KTC:3745 Δ:-89 | T6 — Solid
Christian Watson (WR) | MY:3654 KTC:3633 Δ:+21 | T6 — Solid
Oronde Gadsden (TE) | MY:3647 KTC:4099 Δ:-452 | T6 — Solid
Denzel Boston (WR) | MY:3633 KTC:3442 Δ:+191 | T6 — Solid
Josh Downs (WR) | MY:3610 KTC:3610 Δ:+0 | T7 — Value
Kyle Monangai (RB) | MY:3590 KTC:3590 Δ:+0 | T7 — Value
Davante Adams (WR) | MY:3543 KTC:3398 Δ:+145 | T7 — Value
D.K. Metcalf (WR) | MY:3522 KTC:3647 Δ:-125 | T7 — Value
Mike Evans (WR) | MY:3521 KTC:3289 Δ:+232 | T7 — Value
Terry McLaurin (WR) | MY:3505 KTC:3450 Δ:+55 | T7 — Value
Jayden Higgins (WR) | MY:3498 KTC:3429 Δ:+69 | T7 — Value
Dalton Kincaid (TE) | MY:3487 KTC:3654 Δ:-167 | T7 — Value
Wan'Dale Robinson (WR) | MY:3450 KTC:3543 Δ:-93 | T7 — Value
Chuba Hubbard (RB) | MY:3442 KTC:3405 Δ:+37 | T7 — Value
Travis Hunter (WR) | MY:3440 KTC:3322 Δ:+118 | T7 — Value
Ricky Pearsall (WR) | MY:3429 KTC:3522 Δ:-93 | T7 — Value
Jaylen Warren (RB) | MY:3408 KTC:3215 Δ:+193 | T7 — Value
Ty Simpson (QB) | MY:3405 KTC:3724 Δ:-319 | T7 — Value
Brenton Strange (TE) | MY:3398 KTC:3440 Δ:-42 | T7 — Value
Jake Ferguson (TE) | MY:3370 KTC:3716 Δ:-346 | T7 — Value
Jonah Coleman (RB) | MY:3360 KTC:3084 Δ:+276 | T7 — Value
Chris Bell (WR) | MY:3344 KTC:2978 Δ:+366 | T7 — Value
Michael Pittman Jr. (WR) | MY:3337 KTC:3370 Δ:-33 | T7 — Value
Jakobi Meyers (WR) | MY:3322 KTC:3211 Δ:+111 | T7 — Value
Zach Charbonnet (RB) | MY:3289 KTC:3360 Δ:-71 | T7 — Value
Michael Penix Jr. (QB) | MY:3288 KTC:3151 Δ:+137 | T7 — Value
Jayden Reed (WR) | MY:3231 KTC:3487 Δ:-256 | T7 — Value
Isaiah Likely (TE) | MY:3228 KTC:3740 Δ:-512 | T7 — Value ★
Blake Corum (RB) | MY:3215 KTC:3212 Δ:+3 | T7 — Value
Germie Bernard (WR) | MY:3214 KTC:2951 Δ:+263 | T7 — Value
Courtland Sutton (WR) | MY:3212 KTC:3231 Δ:-19 | T7 — Value
Rico Dowdle (RB) | MY:3211 KTC:3174 Δ:+37 | T7 — Value
Chigoziem Okonkwo (TE) | MY:3174 KTC:3118 Δ:+56 | T7 — Value
Mark Andrews (TE) | MY:3161 KTC:3161 Δ:+0 | T7 — Value
Romeo Doubs (WR) | MY:3151 KTC:3288 Δ:-137 | T7 — Value
Xavier Worthy (WR) | MY:3118 KTC:3344 Δ:-226 | T7 — Value
Antonio Williams (WR) | MY:3108 KTC:2724 Δ:+384 | T7 — Value
Quentin Johnston (WR) | MY:3084 KTC:3337 Δ:-253 | T7 — Value
Jonathon Brooks (RB) | MY:3079 KTC:3079 Δ:+0 | T7 — Value
Nicholas Singleton (RB) | MY:3056 KTC:2919 Δ:+137 | T7 — Value
Elijah Sarratt (WR) | MY:3042 KTC:2761 Δ:+281 | T7 — Value
Jalen Coker (WR) | MY:3023 KTC:3056 Δ:-33 | T7 — Value
Kenneth Gainwell (RB) | MY:2998 KTC:2755 Δ:+243 | T7 — Value
Zachariah Branch (WR) | MY:2982 KTC:2609 Δ:+373 | T7 — Value
J.K. Dobbins (RB) | MY:2981 KTC:2973 Δ:+8 | T7 — Value
AJ Barner (TE) | MY:2978 KTC:3214 Δ:-236 | T7 — Value
Khalil Shakir (WR) | MY:2973 KTC:3023 Δ:-50 | T7 — Value
Chris Godwin (WR) | MY:2966 KTC:2982 Δ:-16 | T7 — Value
Dallas Goedert (TE) | MY:2964 KTC:2981 Δ:-17 | T7 — Value
T.J. Hockenson (TE) | MY:2951 KTC:3228 Δ:-277 | T8 — Upside
Jalen McMillan (WR) | MY:2949 KTC:2998 Δ:-49 | T8 — Upside
Emmett Johnson (RB) | MY:2923 KTC:2495 Δ:+428 | T8 — Upside
De'Zhaun Stribling (WR) | MY:2919 KTC:1926 Δ:+993 | T8 — Upside ★
Tony Pollard (RB) | MY:2910 KTC:2860 Δ:+50 | T8 — Upside
Chris Brazzell (WR) | MY:2908 KTC:2603 Δ:+305 | T8 — Upside
Rhamondre Stevenson (RB) | MY:2881 KTC:2923 Δ:-42 | T8 — Upside
Kaytron Allen (RB) | MY:2860 KTC:2454 Δ:+406 | T8 — Upside
J.J. McCarthy (QB) | MY:2821 KTC:2881 Δ:-60 | T8 — Upside
Jacory Croskey-Merritt (RB) | MY:2814 KTC:2964 Δ:-150 | T8 — Upside
Mike Washington (RB) | MY:2808 KTC:2444 Δ:+364 | T8 — Upside
Jordan Mason (RB) | MY:2799 KTC:2638 Δ:+161 | T8 — Upside
Jacoby Brissett (QB) | MY:2787 KTC:2702 Δ:+85 | T8 — Upside
Tua Tagovailoa (QB) | MY:2762 KTC:2713 Δ:+49 | T8 — Upside
Travis Kelce (TE) | MY:2762 KTC:2808 Δ:-46 | T8 — Upside
Juwan Johnson (TE) | MY:2761 KTC:2762 Δ:-1 | T8 — Upside
Woody Marks (RB) | MY:2755 KTC:2910 Δ:-155 | T8 — Upside
Malachi Fields (WR) | MY:2753 KTC:2620 Δ:+133 | T8 — Upside
Shedeur Sanders (QB) | MY:2751 KTC:2829 Δ:-78 | T8 — Upside
Rachaad White (RB) | MY:2751 KTC:2689 Δ:+62 | T8 — Upside
Max Klare (TE) | MY:2727 KTC:2448 Δ:+279 | T8 — Upside
Terrance Ferguson (TE) | MY:2724 KTC:3108 Δ:-384 | T8 — Upside
Ja'Kobi Lane (WR) | MY:2717 KTC:2348 Δ:+369 | T8 — Upside
Skyler Bell (WR) | MY:2713 KTC:2457 Δ:+256 | T8 — Upside
Tre Harris (WR) | MY:2712 KTC:2762 Δ:-50 | T8 — Upside
Gunnar Helm (TE) | MY:2702 KTC:3042 Δ:-340 | T8 — Upside
Tyler Allgeier (RB) | MY:2699 KTC:2751 Δ:-52 | T8 — Upside
Rashid Shaheed (WR) | MY:2689 KTC:2949 Δ:-260 | T8 — Upside
Tyrone Tracy (RB) | MY:2684 KTC:2612 Δ:+72 | T8 — Upside
Ted Hurst (WR) | MY:2683 KTC:2015 Δ:+668 | T8 — Upside ★
Chris Rodriguez Jr. (RB) | MY:2672 KTC:2531 Δ:+141 | T8 — Upside
Tyjae Spears (RB) | MY:2669 KTC:2684 Δ:-15 | T8 — Upside
Brandon Aiyuk (WR) | MY:2669 KTC:2394 Δ:+275 | T8 — Upside
Mason Taylor (TE) | MY:2641 KTC:2814 Δ:-173 | T8 — Upside
Jerry Jeudy (WR) | MY:2638 KTC:2644 Δ:-6 | T8 — Upside
Dylan Sampson (RB) | MY:2620 KTC:2672 Δ:-52 | T8 — Upside
Dylan Sampson (RB) | MY:2617 KTC:2669 Δ:-52 | T8 — Upside
Isiah Pacheco (RB) | MY:2612 KTC:2263 Δ:+349 | T8 — Upside
Dalton Schultz (TE) | MY:2600 KTC:2669 Δ:-69 | T8 — Upside
Deshaun Watson (QB) | MY:2539 KTC:2028 Δ:+511 | T8 — Upside ★
Geno Smith (QB) | MY:2531 KTC:2178 Δ:+353 | T8 — Upside
Jauan Jennings (WR) | MY:2495 KTC:2799 Δ:-304 | T8 — Upside
Isaac TeSlaa (WR) | MY:2495 KTC:2821 Δ:-326 | T8 — Upside
David Njoku (TE) | MY:2493 KTC:2966 Δ:-473 | T8 — Upside
DELTA KEY: ★ = my value diverges significantly from KTC market (>800 pts). Positive Δ = I value higher than KTC. Negative Δ = KTC values higher than me.
When evaluating trades: use MY values (not KTC) as the primary valuation. Flag if I'm being offered players where KTC > my value (market overvalues them) or giving away players where my value > KTC (I value more than market)."""

DEFAULT_PLAYERS = [
    # QBs
    ("Josh Allen", "QB", "BUF"), ("Drake Maye", "QB", "NE"), ("Jayden Daniels", "QB", "WAS"),
    ("Caleb Williams", "QB", "CHI"), ("Lamar Jackson", "QB", "BAL"), ("Joe Burrow", "QB", "CIN"),
    ("Patrick Mahomes", "QB", "KC"), ("Jaxson Dart", "QB", "NYG"), ("Justin Herbert", "QB", "LAC"),
    ("Trevor Lawrence", "QB", "JAC"), ("Jalen Hurts", "QB", "PHI"), ("Bo Nix", "QB", "DEN"),
    ("Brock Purdy", "QB", "SF"), ("Jordan Love", "QB", "GB"), ("Cam Ward", "QB", "TEN"),
    ("Fernando Mendoza", "QB", "LV"), ("CJ Stroud", "QB", "HOU"), ("Bryce Young", "QB", "CAR"),
    ("Kyler Murray", "QB", "MIN"), ("Malik Willis", "QB", "MIA"), ("JJ McCarthy", "QB", "MIN"),
    ("Anthony Richardson", "QB", "IND"), ("Shedeur Sanders", "QB", "CLE"), ("Dak Prescott", "QB", "DAL"),
    ("Sam Darnold", "QB", "SEA"), ("Baker Mayfield", "QB", "TB"), ("Tyler Shough", "QB", "NO"),
    ("Deshaun Watson", "QB", "CLE"), ("Will Howard", "QB", "PIT"), ("Jalen Milroe", "QB", "SEA"),
    # RBs
    ("Bijan Robinson", "RB", "ATL"), ("Jahmyr Gibbs", "RB", "DET"), ("Jeremiyah Love", "RB", "ARI"),
    ("Ashton Jeanty", "RB", "LV"), ("De'Von Achane", "RB", "MIA"), ("Christian McCaffrey", "RB", "SF"),
    ("Omarion Hampton", "RB", "LAC"), ("Jonathan Taylor", "RB", "IND"), ("James Cook", "RB", "BUF"),
    ("Breece Hall", "RB", "NYJ"), ("TreVeyon Henderson", "RB", "NE"), ("Kenneth Walker", "RB", "KC"),
    ("Quinshon Judkins", "RB", "CLE"), ("Bucky Irving", "RB", "TB"), ("Saquon Barkley", "RB", "PHI"),
    ("Chase Brown", "RB", "CIN"), ("Jadarian Price", "RB", "SEA"), ("Kyren Williams", "RB", "LAR"),
    ("Cam Skattebo", "RB", "NYG"), ("Jaylen Waddle", "RB", "DEN"), ("Josh Jacobs", "RB", "GB"),
    ("Travis Etienne", "RB", "NO"), ("Javonte Williams", "RB", "DAL"), ("RJ Harvey", "RB", "DEN"),
    ("Rico Dowdle", "RB", "PIT"), ("Kyle Monangai", "RB", "CHI"), ("D Andre Swift", "RB", "CHI"),
    ("Jonah Coleman", "RB", "DEN"), ("Zach Charbonnet", "RB", "SEA"), ("Bhayshul Tuten", "RB", "JAC"),
    ("Jacory Croskey-Merritt", "RB", "WAS"), ("Rhamondre Stevenson", "RB", "NE"), ("Woody Marks", "RB", "HOU"),
    ("Nicholas Singleton", "RB", "TEN"), ("Tony Pollard", "RB", "TEN"), ("Blake Corum", "RB", "LAR"),
    ("JK Dobbins", "RB", "DEN"), ("Jordan Mason", "RB", "MIN"), ("Jaylen Warren", "RB", "PIT"),
    ("Trey Benson", "RB", "ARI"), ("Braelon Allen", "RB", "NYJ"), ("Dylan Sampson", "RB", "CLE"),
    ("Kaytron Allen", "RB", "WAS"), ("Emmett Johnson", "RB", "KC"), ("Tank Bigsby", "RB", "PHI"),
    ("Mike Washington Jr", "RB", "LV"), ("Ollie Gordon", "RB", "MIA"), ("LeQuint Allen", "RB", "JAC"),
    ("Jordan James", "RB", "SF"), ("Kimani Vidal", "RB", "LAC"), ("Jaylen Wright", "RB", "MIA"),
    # WRs
    ("Ja Marr Chase", "WR", "CIN"), ("Jaxon Smith-Njigba", "WR", "SEA"), ("Puka Nacua", "WR", "LAR"),
    ("Malik Nabers", "WR", "NYG"), ("Justin Jefferson", "WR", "MIN"), ("Amon-Ra St. Brown", "WR", "DET"),
    ("CeeDee Lamb", "WR", "DAL"), ("Drake London", "WR", "ATL"), ("Tetairoa McMillan", "WR", "CAR"),
    ("Emeka Egbuka", "WR", "TB"), ("George Pickens", "WR", "DAL"), ("Carnell Tate", "WR", "TEN"),
    ("Garrett Wilson", "WR", "NYJ"), ("Nico Collins", "WR", "HOU"), ("Jordyn Tyson", "WR", "NO"),
    ("Ladd McConkey", "WR", "LAC"), ("Rashee Rice", "WR", "KC"), ("DeVonta Smith", "WR", "PHI"),
    ("Luther Burden", "WR", "CHI"), ("Marvin Harrison Jr", "WR", "ARI"), ("Rome Odunze", "WR", "CHI"),
    ("Makai Lemon", "WR", "PHI"), ("Brian Thomas Jr", "WR", "JAC"), ("Zay Flowers", "WR", "BAL"),
    ("Tee Higgins", "WR", "CIN"), ("AJ Brown", "WR", "PHI"), ("Jameson Williams", "WR", "DET"),
    ("Jaylen Waddle", "WR", "DEN"), ("Chris Olave", "WR", "NO"), ("KC Concepcion", "WR", "CLE"),
    ("Jayden Higgins", "WR", "HOU"), ("Denzel Boston", "WR", "CLE"), ("Travis Hunter", "WR", "JAC"),
    ("Chris Bell", "WR", "MIA"), ("Germie Bernard", "WR", "PIT"), ("Zachariah Branch", "WR", "ATL"),
    ("Antonio Williams", "WR", "WAS"), ("Malachi Fields", "WR", "NYG"), ("Dontayvion Wicks", "WR", "PHI"),
    ("Chimere Dike", "WR", "TEN"), ("Tre Harris", "WR", "LAC"), ("Kyle Williams", "WR", "NE"),
    ("Troy Franklin", "WR", "DEN"), ("Kayshon Boutte", "WR", "NE"), ("Elijah Sarratt", "WR", "BAL"),
    ("Chris Brazzell", "WR", "CAR"), ("Skyler Bell", "WR", "BUF"), ("Keon Coleman", "WR", "BUF"),
    ("Adonai Mitchell", "WR", "NYJ"), ("Xavier Worthy", "WR", "KC"), ("Wan Dale Robinson", "WR", "TEN"),
    ("Josh Downs", "WR", "IND"), ("Parker Washington", "WR", "JAC"), ("Ricky Pearsall", "WR", "SF"),
    ("Jayden Reed", "WR", "GB"), ("De Zhaun Stribling", "WR", "SF"), ("Romeo Doubs", "WR", "NE"),
    ("Terry McLaurin", "WR", "WAS"), ("Matthew Golden", "WR", "GB"), ("Ja Kobi Lane", "WR", "BAL"),
    ("Mike Evans", "WR", "SF"), ("DK Metcalf", "WR", "PIT"), ("Courtland Sutton", "WR", "DEN"),
    ("Calvin Ridley", "WR", "TEN"), ("Brandon Aiyuk", "WR", "SF"), ("Tory Horton", "WR", "SEA"),
    ("Tank Dell", "WR", "HOU"), ("Pat Bryant", "WR", "DEN"), ("Isaac TeSlaa", "WR", "DET"),
    ("Dont'e Thornton", "WR", "LV"), ("Jalen Coker", "WR", "CAR"), ("Jaylin Noel", "WR", "HOU"),
    ("Khalil Shakir", "WR", "BUF"), ("Jalen McMillan", "WR", "TB"), ("Quentin Johnston", "WR", "LAC"),
    ("Xavier Legette", "WR", "CAR"), ("Savion Williams", "WR", "GB"), ("Elic Ayomanor", "WR", "TEN"),
    # TEs
    ("Brock Bowers", "TE", "LV"), ("Trey McBride", "TE", "ARI"), ("Colston Loveland", "TE", "CHI"),
    ("Tyler Warren", "TE", "IND"), ("Tucker Kraft", "TE", "GB"), ("Harold Fannin", "TE", "CLE"),
    ("Sam LaPorta", "TE", "DET"), ("Kyle Pitts", "TE", "ATL"), ("Oronde Gadsden", "TE", "LAC"),
    ("Kenyon Sadiq", "TE", "NYJ"), ("Eli Stowers", "TE", "PHI"), ("Isaiah Likely", "TE", "NYG"),
    ("Gunnar Helm", "TE", "TEN"), ("AJ Barner", "TE", "SEA"), ("TJ Hockenson", "TE", "MIN"),
    ("Terrance Ferguson", "TE", "LAR"), ("Mark Andrews", "TE", "BAL"), ("Brenton Strange", "TE", "JAC"),
    ("Max Klare", "TE", "LAR"), ("Elijah Arroyo", "TE", "SEA"), ("Dallas Goedert", "TE", "PHI"),
    ("Ja Tavion Sanders", "TE", "CAR"), ("Mason Taylor", "TE", "NYJ"), ("Dalton Kincaid", "TE", "BUF"),
    ("Jake Ferguson", "TE", "DAL"), ("George Kittle", "TE", "SF"), ("Oscar Delp", "TE", "NO"),
    ("Justin Joly", "TE", "DEN"), ("Eli Raridon", "TE", "NE"), ("Michael Trigg", "TE", "DAL"),
    ("Darnell Washington", "TE", "PIT"), ("Cole Kmet", "TE", "CHI"), ("Cade Otton", "TE", "TB"),
    ("Pat Freiermuth", "TE", "PIT"), ("David Njoku", "TE", "FA"), ("Greg Dulcich", "TE", "MIA"),
    # PICK ASSETS — Dynasty picks (KTC values May 2026)
    ("2027 Early 1st", "PICK", ""), ("2027 Mid 1st", "PICK", ""), ("2027 Late 1st", "PICK", ""),
    ("2026 Early 1st", "PICK", ""), ("2028 Early 1st", "PICK", ""), ("2026 Mid 1st", "PICK", ""),
    ("2027 Early 2nd", "PICK", ""), ("2028 Mid 1st", "PICK", ""), ("2026 Late 1st", "PICK", ""),
    ("2028 Late 1st", "PICK", ""), ("2027 Mid 2nd", "PICK", ""), ("2027 Late 2nd", "PICK", ""),
    ("2026 Early 2nd", "PICK", ""), ("2027 Early 3rd", "PICK", ""), ("2028 Early 2nd", "PICK", ""),
    ("2026 Mid 2nd", "PICK", ""), ("2027 Mid 3rd", "PICK", ""), ("2028 Mid 2nd", "PICK", ""),
    ("2026 Late 2nd", "PICK", ""), ("2027 Late 3rd", "PICK", ""), ("2028 Late 2nd", "PICK", ""),
    ("2027 Early 4th", "PICK", ""), ("2026 Early 3rd", "PICK", ""), ("2027 Mid 4th", "PICK", ""),
    ("2028 Early 3rd", "PICK", ""), ("2026 Mid 3rd", "PICK", ""), ("2027 Late 4th", "PICK", ""),
    ("2028 Mid 3rd", "PICK", ""), ("2026 Late 3rd", "PICK", ""), ("2028 Late 3rd", "PICK", ""),
    ("2026 Early 4th", "PICK", ""), ("2026 Mid 4th", "PICK", ""), ("2028 Early 4th", "PICK", ""),
]

def seed_players():
    """
    ELO scale maps to KTC tiers with real separation (400-3000):
    Tier 1 (9000+ KTC): ~2800-3000 ELO
    Tier 2 (7500-8999): ~2400-2799 ELO
    Tier 4 (4500-5999): ~1600-1999 ELO
    Tier 7 (under 3000): ~400-1199 ELO
    KTC values updated May 12, 2026
    """
    KTC_SEED_ELOS = {
        'Ja\'Marr Chase': 3000,
        'Josh Allen': 3000,
        'Bijan Robinson': 2998,
        'Jaxon Smith-Njigba': 2971,
        'Jahmyr Gibbs': 2884,
        'Drake Maye': 2848,
        'Brock Bowers': 2687,
        'Puka Nacua': 2644,
        'Trey McBride': 2566,
        'Caleb Williams': 2465,
        'Malik Nabers': 2443,
        'Jayden Daniels': 2432,
        'Justin Jefferson': 2430,
        'Amon-Ra St. Brown': 2408,
        'Lamar Jackson': 2396,
        '2027 Early 1st': 2251,
        '2026 Early 1st': 1883,
        '2027 Mid 1st': 1858,
        'Breece Hall': 1800,
        'TreVeyon Henderson': 1791,
        'Tucker Kraft': 1791,
        'Ladd McConkey': 1783,
        'Jordyn Tyson': 1779,
        'Rashee Rice': 1778,
        'Cam Ward': 1764,
        'Luther Burden': 1756,
        'DeVonta Smith': 1755,
        'Sam LaPorta': 1723,
        'Makai Lemon': 1714,
        'Marvin Harrison Jr.': 1714,
        'Christian McCaffrey': 1713,
        'Dak Prescott': 1698,
        '2028 Early 1st': 1683,
        'Brian Thomas Jr.': 1683,
        'Kyle Pitts': 1681,
        'Tee Higgins': 1677,
        'Bucky Irving': 1676,
        '2027 Late 1st': 1674,
        'Tyler Shough': 1670,
        'Chase Brown': 1669,
        'Saquon Barkley': 1668,
        'Jaylen Waddle': 1666,
        'C.J. Stroud': 1666,
        'Kyren Williams': 1663,
        'A.J. Brown': 1650,
        'Sam Darnold': 1643,
        'Baker Mayfield': 1637,
        'Jadarian Price': 1633,
        'Zay Flowers': 1614,
        'Jameson Williams': 1612,
        'Kenyon Sadiq': 1610,
        '2026 Mid 1st': 1602,
        'Jared Goff': 1594,
        'Cam Skattebo': 1592,
        'KC Concepcion': 1574,
        'Bryce Young': 1571,
        'Josh Jacobs': 1559,
        'Javonte Williams': 1553,
        '2028 Mid 1st': 1551,
        'Travis Etienne': 1545,
        'Alec Pierce': 1469,
        'Kyler Murray': 1466,
        'Oronde Gadsden': 1466,
        'Malik Willis': 1442,
        'Jordan Addison': 1433,
        '2028 Late 1st': 1430,
        'Daniel Jones': 1429,
        'Bhayshul Tuten': 1427,
        '2027 Early 2nd': 1427,
        '2026 Late 1st': 1421,
        'Omar Cooper Jr.': 1413,
        'Derrick Henry': 1397,
        'Matthew Golden': 1374,
        'Isaiah Likely': 1372,
        'Ty Simpson': 1368,
        'Jake Ferguson': 1366,
        'George Kittle': 1364,
        'D.J. Moore': 1356,
        'Matthew Stafford': 1352,
        'Michael Wilson': 1351,
        'Dalton Kincaid': 1350,
        'DK Metcalf': 1348,
        'Christian Watson': 1345,
        'Josh Downs': 1339,
        'Kyle Monangai': 1333,
        'Wan\'Dale Robinson': 1321,
        'Ricky Pearsall': 1316,
        'Parker Washington': 1316,
        'D\'Andre Swift': 1311,
        'RJ Harvey': 1310,
        'Jayden Reed': 1307,
        'Terry McLaurin': 1297,
        'Brenton Strange': 1294,
        'Jayden Higgins': 1292,
        '2027 Mid 2nd': 1292,
        'David Montgomery': 1286,
        'Chuba Hubbard': 1285,
        'Davante Adams': 1284,
        'Michael Pittman': 1276,
        'Zach Charbonnet': 1274,
        'Xavier Worthy': 1270,
        'Quentin Johnston': 1268,
        'Travis Hunter': 1264,
        'Mike Evans': 1255,
        'Romeo Doubs': 1255,
        '2026 Early 2nd': 1252,
        'Courtland Sutton': 1240,
        'T.J. Hockenson': 1239,
        'Jaylen Warren': 1236,
        'AJ Barner': 1236,
        'Blake Corum': 1235,
        'Jakobi Meyers': 1235,
        '2027 Late 2nd': 1230,
        'Rico Dowdle': 1225,
        'Mark Andrews': 1222,
        'Michael Penix Jr.': 1219,
        'Chigoziem Okonkwo': 1211,
    }
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    for name, pos, team in DEFAULT_PLAYERS:
        starting_elo = KTC_SEED_ELOS.get(name, 1400)
        c.execute("INSERT OR IGNORE INTO player_rankings (player_name, position, team, elo_score, comparisons, last_updated) VALUES (?, ?, ?, ?, 0, ?)",
                 (name, pos, team, starting_elo, datetime.now().isoformat()))
    conn.commit()
    conn.close()

def seed_ktc_market_data():
    """Store fresh KTC values in market_data table for tier/trade analysis"""
    KTC_PLAYERS = [
        ('Ja\'Marr Chase', 'WR', 'CIN', 1, 9999, 1),
        ('Josh Allen', 'QB', 'BUF', 2, 9999, 1),
        ('Bijan Robinson', 'RB', 'ATL', 3, 9991, 1),
        ('Jaxon Smith-Njigba', 'WR', 'SEA', 4, 9889, 1),
        ('Jahmyr Gibbs', 'RB', 'DET', 5, 9553, 2),
        ('Drake Maye', 'QB', 'NEP', 6, 9416, 2),
        ('Brock Bowers', 'TE', 'LVR', 7, 8794, 3),
        ('Puka Nacua', 'WR', 'LAR', 8, 8631, 3),
        ('Trey McBride', 'TE', 'ARI', 9, 8330, 3),
        ('Caleb Williams', 'QB', 'CHI', 10, 7943, 4),
        ('Malik Nabers', 'WR', 'NYG', 11, 7858, 4),
        ('Jayden Daniels', 'QB', 'WAS', 12, 7814, 4),
        ('Justin Jefferson', 'WR', 'MIN', 13, 7806, 4),
        ('Amon-Ra St. Brown', 'WR', 'DET', 14, 7723, 4),
        ('Lamar Jackson', 'QB', 'BAL', 15, 7676, 4),
        ('2027 Early 1st', 'PICK', 'FA', 20, 7117, 6),
        ('2026 Early 1st', 'PICK', 'FA', 40, 5702, 12),
        ('2027 Mid 1st', 'PICK', 'FA', 44, 5609, 12),
        ('Breece Hall', 'RB', 'NYJ', 50, 5384, 13),
        ('TreVeyon Henderson', 'RB', 'NEP', 51, 5351, 10),
        ('Tucker Kraft', 'TE', 'GBP', 52, 5348, 10),
        ('Ladd McConkey', 'WR', 'LAC', 53, 5318, 10),
        ('Jordyn Tyson', 'WR', 'NOS', 54, 5304, 10),
        ('Rashee Rice', 'WR', 'KCC', 55, 5301, 10),
        ('Cam Ward', 'QB', 'TEN', 56, 5245, 10),
        ('Luther Burden', 'WR', 'CHI', 57, 5216, 10),
        ('DeVonta Smith', 'WR', 'PHI', 58, 5211, 10),
        ('Sam LaPorta', 'TE', 'DET', 59, 5089, 11),
        ('Makai Lemon', 'WR', 'PHI', 60, 5055, 11),
        ('Marvin Harrison Jr.', 'WR', 'ARI', 61, 5054, 11),
        ('Christian McCaffrey', 'RB', 'SFO', 62, 5051, 11),
        ('Dak Prescott', 'QB', 'DAL', 63, 4993, 11),
        ('2028 Early 1st', 'PICK', 'FA', 64, 4935, 11),
        ('Brian Thomas Jr.', 'WR', 'JAC', 65, 4934, 11),
        ('Kyle Pitts', 'TE', 'ATL', 66, 4927, 11),
        ('Tee Higgins', 'WR', 'CIN', 67, 4911, 11),
        ('Bucky Irving', 'RB', 'TBB', 68, 4908, 11),
        ('2027 Late 1st', 'PICK', 'FA', 69, 4898, 11),
        ('Tyler Shough', 'QB', 'NOS', 70, 4884, 11),
        ('Chase Brown', 'RB', 'CIN', 71, 4882, 11),
        ('Saquon Barkley', 'RB', 'PHI', 72, 4876, 11),
        ('Jaylen Waddle', 'WR', 'DEN', 73, 4870, 11),
        ('C.J. Stroud', 'QB', 'HOU', 74, 4868, 11),
        ('Kyren Williams', 'RB', 'LAR', 75, 4858, 11),
        ('A.J. Brown', 'WR', 'PHI', 76, 4808, 11),
        ('Sam Darnold', 'QB', 'SEA', 77, 4781, 11),
        ('Baker Mayfield', 'QB', 'TBB', 78, 4758, 11),
        ('Jadarian Price', 'RB', 'SEA', 79, 4742, 11),
        ('Zay Flowers', 'WR', 'BAL', 80, 4670, 12),
        ('Jameson Williams', 'WR', 'DET', 81, 4662, 12),
        ('Kenyon Sadiq', 'TE', 'NYJ', 82, 4652, 12),
        ('2026 Mid 1st', 'PICK', 'FA', 83, 4622, 12),
        ('Jared Goff', 'QB', 'DET', 84, 4591, 12),
        ('Cam Skattebo', 'RB', 'NYG', 85, 4586, 12),
        ('KC Concepcion', 'WR', 'CLE', 86, 4516, 13),
        ('Bryce Young', 'QB', 'CAR', 87, 4505, 13),
        ('Josh Jacobs', 'RB', 'GBP', 88, 4457, 13),
        ('Javonte Williams', 'RB', 'DAL', 89, 4434, 13),
        ('2028 Mid 1st', 'PICK', 'FA', 90, 4425, 13),
        ('Travis Etienne', 'RB', 'NOS', 91, 4403, 13),
        ('Alec Pierce', 'WR', 'IND', 92, 4111, 14),
        ('Kyler Murray', 'QB', 'MIN', 93, 4101, 14),
        ('Oronde Gadsden', 'TE', 'LAC', 94, 4099, 14),
        ('Malik Willis', 'QB', 'MIA', 95, 4007, 15),
        ('Jordan Addison', 'WR', 'MIN', 97, 3973, 15),
        ('2028 Late 1st', 'PICK', 'FA', 98, 3963, 15),
        ('Daniel Jones', 'QB', 'IND', 99, 3959, 15),
        ('Bhayshul Tuten', 'RB', 'JAC', 100, 3951, 15),
        ('2027 Early 2nd', 'PICK', 'FA', 101, 3950, 15),
        ('2026 Late 1st', 'PICK', 'FA', 102, 3926, 15),
        ('Omar Cooper Jr.', 'WR', 'NYJ', 103, 3896, 15),
        ('Derrick Henry', 'RB', 'BAL', 104, 3834, 16),
        ('Matthew Golden', 'WR', 'GBP', 105, 3745, 17),
        ('Isaiah Likely', 'TE', 'NYG', 106, 3740, 17),
        ('Ty Simpson', 'QB', 'LAR', 107, 3724, 17),
        ('Jake Ferguson', 'TE', 'DAL', 108, 3716, 17),
        ('George Kittle', 'TE', 'SFO', 109, 3708, 17),
        ('D.J. Moore', 'WR', 'BUF', 110, 3678, 17),
        ('Matthew Stafford', 'QB', 'LAR', 111, 3661, 17),
        ('Michael Wilson', 'WR', 'ARI', 112, 3656, 17),
        ('Dalton Kincaid', 'TE', 'BUF', 113, 3654, 17),
        ('DK Metcalf', 'WR', 'PIT', 114, 3647, 17),
        ('Christian Watson', 'WR', 'GBP', 115, 3633, 17),
        ('Josh Downs', 'WR', 'IND', 116, 3610, 17),
        ('Kyle Monangai', 'RB', 'CHI', 117, 3590, 17),
        ('Wan\'Dale Robinson', 'WR', 'TEN', 118, 3543, 17),
        ('Ricky Pearsall', 'WR', 'SFO', 119, 3522, 17),
        ('Parker Washington', 'WR', 'JAC', 120, 3521, 17),
        ('D\'Andre Swift', 'RB', 'CHI', 121, 3505, 17),
        ('RJ Harvey', 'RB', 'DEN', 122, 3498, 17),
        ('Jayden Reed', 'WR', 'GBP', 123, 3487, 17),
        ('Terry McLaurin', 'WR', 'WAS', 124, 3450, 17),
        ('Brenton Strange', 'TE', 'JAC', 126, 3440, 17),
        ('Jayden Higgins', 'WR', 'HOU', 127, 3429, 17),
        ('2027 Mid 2nd', 'PICK', 'FA', 128, 3429, 17),
        ('David Montgomery', 'RB', 'HOU', 129, 3408, 17),
        ('Chuba Hubbard', 'RB', 'CAR', 130, 3405, 17),
        ('Davante Adams', 'WR', 'LAR', 131, 3398, 17),
        ('Michael Pittman', 'WR', 'PIT', 132, 3370, 17),
        ('Zach Charbonnet', 'RB', 'SEA', 133, 3360, 17),
        ('Xavier Worthy', 'WR', 'KCC', 134, 3344, 17),
        ('Quentin Johnston', 'WR', 'LAC', 135, 3337, 17),
        ('Travis Hunter', 'WR', 'JAC', 136, 3322, 17),
        ('Mike Evans', 'WR', 'SFO', 137, 3289, 17),
        ('Romeo Doubs', 'WR', 'NEP', 138, 3288, 17),
        ('2026 Early 2nd', 'PICK', 'FA', 139, 3277, 17),
        ('Courtland Sutton', 'WR', 'DEN', 140, 3231, 17),
        ('T.J. Hockenson', 'TE', 'MIN', 141, 3228, 17),
        ('Jaylen Warren', 'RB', 'PIT', 142, 3215, 17),
        ('AJ Barner', 'TE', 'SEA', 143, 3214, 17),
        ('Blake Corum', 'RB', 'LAR', 144, 3212, 17),
        ('Jakobi Meyers', 'WR', 'JAC', 145, 3211, 17),
        ('2027 Late 2nd', 'PICK', 'FA', 146, 3192, 17),
        ('Rico Dowdle', 'RB', 'PIT', 147, 3174, 17),
        ('Mark Andrews', 'TE', 'BAL', 148, 3161, 17),
        ('Michael Penix Jr.', 'QB', 'ATL', 149, 3151, 17),
        ('Chigoziem Okonkwo', 'TE', 'WAS', 150, 3118, 17),
    ]
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    # Only seed if empty or stale
    c.execute("SELECT COUNT(*) FROM market_data WHERE source='ktc'")
    existing = c.fetchone()[0]
    if existing > 100:
        conn.close()
        return  # Already have data
    for name, pos, team, rank, value, tier in KTC_PLAYERS:
        c.execute('''INSERT OR REPLACE INTO market_data
                    (source, player_name, rank, value, position, team, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)''',
                 ('ktc', name, rank, value, pos, team, datetime.now().isoformat()))
    conn.commit()
    conn.close()

seed_players()
seed_ktc_market_data()

@app.route('/')
def index():
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/api/chat', methods=['POST'])
def chat():
    data = request.json
    user_message = data.get('message', '')
    image_data   = data.get('image', None)
    file_data    = data.get('file_data', None)
    week_key = get_week_key()

    content = []

    if image_data:
        image_content = image_data.split(',')[1] if ',' in image_data else image_data
        media_type = 'image/png' if 'png' in image_data[:30].lower() else 'image/jpeg'
        content.append({"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_content}})

    elif file_data:
        fname    = file_data.get('name', 'file')
        raw_data = file_data.get('data', '')
        is_excel = file_data.get('isExcel', False)
        is_pdf   = file_data.get('isPDF', False)

        if is_pdf and raw_data:
            pdf_b64 = raw_data.split(',')[1] if ',' in raw_data else raw_data
            content.append({"type": "document",
                "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}})

        elif is_excel and raw_data:
            try:
                import io, base64 as b64lib
                from openpyxl import load_workbook as _lwb
                excel_b64 = raw_data.split(',')[1] if ',' in raw_data else raw_data
                excel_bytes = b64lib.b64decode(excel_b64)
                wb_xl = _lwb(io.BytesIO(excel_bytes), data_only=True)

                text_parts = [f"[Excel file: {fname}]"]

                # Priority order: trade/roster context first, draft tools last
                # Tabs with these keywords get read first
                PRIORITY_KEYWORDS = [
                    'my roster', 'my asset', 'tendencies', 'gm tend',
                    'trade log', 'velvet spade', 'capital gains', 'trs',
                    'gentlemans', "gentleman's", 'dashboard', 'rosters',
                ]
                SKIP_KEYWORDS = [
                    'draft board', 'pick map', 'manager roster',
                    'qb board', 'rb board', 'wr board', 'te board',
                    'pick value', 'methodology', 'tier break', '_mgr_helper',
                ]

                all_visible = [s for s in wb_xl.sheetnames
                               if not s.startswith('_')
                               and wb_xl[s].sheet_state == 'visible']

                def priority_score(name):
                    n = name.lower()
                    if any(k in n for k in SKIP_KEYWORDS): return 99
                    for i, k in enumerate(PRIORITY_KEYWORDS):
                        if k in n: return i
                    return 50

                sheets_ordered = sorted(all_visible, key=priority_score)
                # Drop skip-category sheets entirely
                sheets_to_read = [s for s in sheets_ordered
                                  if priority_score(s) < 99][:8]

                # Per-sheet row caps: smaller for large roster sheets, bigger for summaries
                def row_cap(name):
                    n = name.lower()
                    if any(k in n for k in ['dashboard','log','tendencies']): return 100
                    if any(k in n for k in ['my roster','my asset']): return 200
                    return 150  # roster tabs: enough to cover ~12 teams × 25 players

                total_chars = 0
                CHAR_LIMIT = 45000

                for sname in sheets_to_read:
                    if total_chars >= CHAR_LIMIT:
                        text_parts.append(f"\n[Remaining sheets skipped — context limit reached]")
                        break
                    ws_xl = wb_xl[sname]
                    cap = row_cap(sname)
                    sheet_lines = [f"\n=== {sname} ==="]
                    rows_added = 0
                    for row in ws_xl.iter_rows(min_row=1, max_row=600, values_only=True):
                        if all(v is None for v in row): continue
                        # Skip rows that are purely formula artifacts (all numbers, no names)
                        row_str = '\t'.join([str(v).strip() if v is not None else '' for v in row])
                        if row_str.strip():
                            sheet_lines.append(row_str)
                            rows_added += 1
                        if rows_added >= cap:
                            sheet_lines.append(f"[...{ws_xl.max_row - rows_added} more rows in this sheet]")
                            break
                    sheet_text = '\n'.join(sheet_lines)
                    text_parts.append(sheet_text)
                    total_chars += len(sheet_text)

                text_parts.append(f"\n[Sheets read: {', '.join(sheets_to_read)}]")
                if len(all_visible) > len(sheets_to_read):
                    skipped = [s for s in all_visible if s not in sheets_to_read]
                    text_parts.append(f"[Sheets skipped (draft tools): {', '.join(skipped[:10])}]")

                extracted = '\n'.join(text_parts)
                content.append({"type": "text", "text": extracted})

            except ImportError:
                content.append({"type": "text",
                    "text": f"[SERVER ERROR: openpyxl not installed — cannot parse {fname}. "
                            f"Redeploy with openpyxl in requirements.txt.]"})
            except Exception as ex:
                content.append({"type": "text",
                    "text": f"[ERROR parsing {fname}: {type(ex).__name__}: {ex}]"})
        else:
            content.append({"type": "text", "text": f"[File: {fname}]\n{raw_data[:40000]}"})

    if user_message:
        content.append({"type": "text", "text": user_message})

    if not content:
        content = user_message or ''

    if isinstance(content, list) and len(content) == 1 and content[0].get('type') == 'text':
        content = content[0]['text']

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT role, content FROM chat_history WHERE week_key=? ORDER BY id", (week_key,))
    history_rows = c.fetchall()
    conn.close()

    messages = []
    for role, msg_content in history_rows:
        try:
            parsed = json.loads(msg_content)
            messages.append({"role": role, "content": parsed})
        except:
            messages.append({"role": role, "content": msg_content})

    messages.append({"role": "user", "content": content})
    if len(messages) > 20:
        messages = messages[-20:]

    # Build fully dynamic system prompt — all roster data from DB, nothing hardcoded
    try:
        live_values = get_player_values_block()
    except Exception: live_values = ''
    try:
        kb_context = get_kb_context()
    except Exception: kb_context = ''
    try:
        # Single combined Sleeper fetch per league — gets my roster AND all managers in one pass
        vs_context = get_full_league_context('velvet_spade')
        gl_context = get_full_league_context('gentlemans_dynasty')

        system_with_values = (SYSTEM_PROMPT
            .replace('{VS_ROSTER_BLOCK}',  vs_context['my_roster'])
            .replace('{VS_PICKS_BLOCK}',   vs_context.get('my_picks', '[picks not loaded]'))
            .replace('{GL_ROSTER_BLOCK}',  gl_context['my_roster'])
            .replace('{CG_ROSTER_BLOCK}',  get_roster_block('Capital Gains'))
            .replace('{CG_PICKS_BLOCK}',   'R1 own | R1 Dudesss | R1 GNAwin0 | R1 TeddySalad | R2 own | R2 LegendsDie | R2 GNAwin0 | R3-R7')
            .replace('{TRS_ROSTER_BLOCK}', get_roster_block('TRS')))

        if vs_context['all_rosters']: system_with_values += vs_context['all_rosters']
        if gl_context['all_rosters']: system_with_values += gl_context['all_rosters']
    except Exception: system_with_values = SYSTEM_PROMPT
    if live_values: system_with_values += "\n\n" + live_values
    if kb_context:  system_with_values += "\n\n" + kb_context

    try:
        response = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=2000,
            system=system_with_values,
            tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 5}],
            messages=messages
        )

        assistant_message = ""
        for block in response.content:
            if hasattr(block, 'text'):
                assistant_message += block.text

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        user_content_str = json.dumps(content) if isinstance(content, list) else content
        c.execute("INSERT INTO chat_history (week_key, role, content, timestamp) VALUES (?, ?, ?, ?)",
                 (week_key, 'user', user_content_str, datetime.now().isoformat()))
        c.execute("INSERT INTO chat_history (week_key, role, content, timestamp) VALUES (?, ?, ?, ?)",
                 (week_key, 'assistant', assistant_message, datetime.now().isoformat()))
        conn.commit()
        conn.close()

        return jsonify({"response": assistant_message, "success": True})
    except Exception as e:
        return jsonify({"response": f"Error: {str(e)}", "success": False})

@app.route('/api/chat/history', methods=['GET'])
def get_history():
    week_key = get_week_key()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT role, content, timestamp FROM chat_history WHERE week_key=? ORDER BY id", (week_key,))
    rows = c.fetchall()
    conn.close()
    history = []
    for role, content, timestamp in rows:
        try:
            parsed = json.loads(content)
            if isinstance(parsed, list):
                text = next((b.get('text','') for b in parsed if isinstance(b,dict) and b.get('type')=='text'), '')
                has_image = any(isinstance(b,dict) and b.get('type')=='image' for b in parsed)
                history.append({"role": role, "content": text, "has_image": has_image, "timestamp": timestamp})
            else:
                history.append({"role": role, "content": content, "has_image": False, "timestamp": timestamp})
        except:
            history.append({"role": role, "content": content, "has_image": False, "timestamp": timestamp})
    return jsonify({"history": history, "success": True})

@app.route('/api/chat/clear', methods=['POST'])
def clear_chat():
    week_key = get_week_key()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM chat_history WHERE week_key=?", (week_key,))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route('/api/chat/search', methods=['POST'])
def search_chat():
    query = request.json.get('query', '').lower()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT role, content, timestamp, week_key FROM chat_history ORDER BY id DESC LIMIT 500")
    rows = c.fetchall()
    conn.close()
    results = []
    for role, content, timestamp, week_key in rows:
        if query in content.lower():
            results.append({"role": role, "content": content[:300], "timestamp": timestamp, "week": week_key})
        if len(results) >= 20:
            break
    return jsonify({"results": results, "success": True})

@app.route('/api/dashboard/static', methods=['GET'])
def get_static_dashboard():
    """Instant-loading static dashboard from known context — no AI calls"""
    data = {
        "roster_summary": [
            {"league": "Capital Gains", "strategy": "REBUILD → 2027", "priority": "#3",
             "top_assets": ["Drake Maye 9,412", "Drake London 6,887", "Sam LaPorta 5,042"],
             "picks": "4x 2027 R1 · R1(own) + R1(Dudesss) + R1(GNAwin0) + R1(TeddySalad)",
             "alert": "29 players — cut to 20 by September"},
            {"league": "Twenty Run Savages", "strategy": "COMPETING NOW", "priority": "#2",
             "top_assets": ["Drake Maye 9,412", "Bijan Robinson 9,987", "JSN 9,918"],
             "picks": "2027 R1(Stinky) + R1(own)",
             "alert": "Draft still ongoing — picks 3.07, 4.03, 4.07 remaining"},
            {"league": "Gentleman's Dynasty", "strategy": "REBUILD → 2027-28", "priority": "#4",
             "top_assets": ["Brock Bowers 8,765", "Patrick Mahomes 6,814", "Carnell Tate 5,830"],
             "picks": "2027 R1(own) · 2028 R1+R2+R3+R4",
             "alert": "Acquire LaPorta from c1smith11 — #1 priority"},
            {"league": "Velvet Spade", "strategy": "STARTUP COMPLETE — Trade season active", "priority": "#1",
             "top_assets": ["Pick 1.02 overall", "2.11", "3.11"],
             "picks": "28-round startup — pick #2 overall",
             "alert": "Velvet Spade roster — update via Sleeper sync"}
        ],
        "quick_actions": [
            {"league": "Gentleman's", "action": "Send LaPorta offer to c1smith11", "message": "Njoku and 2027 2nd for LaPorta. Works for both of us."},
            {"league": "Gentleman's", "action": "Sell McCarthy to SenorHyde", "message": "McCarthy for your 2027 2nd. You need the QB."},
            {"league": "TRS", "action": "Add K and backup DST via FAAB", "message": "Streaming K and DST needed before Week 1"},
            {"league": "Velvet Spade", "action": "Trade season — add RB2 using pick capital leverage", "message": "4x 2027 firsts = massive trade chip. Target proven RBs."}
        ],
        "ktc_baseline_date": "May 8, 2026"
    }
    return jsonify({"data": data, "success": True})

@app.route('/api/dashboard', methods=['GET'])
def get_dashboard():
    """AI-powered dashboard with live web searches — triggered on demand"""
    # Check cache first (6 hour TTL)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute("SELECT data, last_updated FROM dashboard_cache WHERE cache_key='main'")
        row = c.fetchone()
        if row:
            last_updated = datetime.fromisoformat(row[1])
            if datetime.now() - last_updated < timedelta(hours=6):
                conn.close()
                return jsonify({"data": json.loads(row[0]), "success": True, "cached": True,
                               "cached_at": row[1]})
    except Exception:
        pass
    conn.close()

    try:
        response = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=2000,
            system=SYSTEM_PROMPT,
            tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 3}],
            messages=[{"role": "user", "content": """Generate a dynasty GM dashboard for MJBrutus. Search for 1-2 breaking news items relevant to his rosters. Return ONLY valid JSON:
{"alerts":[{"type":"warning|info|opportunity","league":"league name","message":"short alert","action":"what to do"}],
"news":[{"player":"name","team":"team","position":"QB","leagues":["leagues"],"headline":"news","impact":"dynasty impact","recommendation":"BUY|SELL|HOLD|MONITOR"}],
"movers":[{"player":"name","direction":"up|down","change":"+/-amount","reason":"why","action":"what to do"}],
"trade_targets":[{"player":"name","owner":"their team","league":"which league","offer":"under 20 words","rationale":"why"}],
"weekly_priorities":["priority 1","priority 2","priority 3"]}"""}]
        )
        text = "".join(b.text for b in response.content if hasattr(b, 'text'))
        start = text.find('{')
        end = text.rfind('}') + 1
        if start >= 0 and end > start:
            data = json.loads(text[start:end])
            # Cache it
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute("INSERT OR REPLACE INTO dashboard_cache (cache_key, data, last_updated) VALUES (?, ?, ?)",
                     ('main', json.dumps(data), datetime.now().isoformat()))
            conn.commit()
            conn.close()
            return jsonify({"data": data, "success": True, "cached": False})
    except Exception as e:
        pass

    # Fallback
    return jsonify({"data": {
        "alerts": [{"type": "opportunity", "league": "Gentleman's", "message": "LaPorta on block at c1smith11", "action": "Send Njoku + 2027 2nd offer"}],
        "news": [{"player": "Drake Maye", "team": "NE", "position": "QB", "leagues": ["Capital Gains", "TRS"], "headline": "AJ Brown trade expected post June 1", "impact": "Maye value increases with elite WR1", "recommendation": "HOLD"}],
        "movers": [{"player": "Carnell Tate", "direction": "up", "change": "+150", "reason": "Strong pre-draft buzz as TEN WR1", "action": "Hold — core asset"}],
        "trade_targets": [{"player": "Sam LaPorta", "owner": "c1smith11", "league": "Gentleman's Dynasty", "offer": "Njoku and 2027 2nd for LaPorta", "rationale": "TE premium upgrade"}],
        "weekly_priorities": ["Identify VS trade targets — add proven RB2", "Send LaPorta offer in Gentleman's", "Evaluate CG pick capital trade opportunities"]
    }, "success": True, "cached": False})


@app.route('/api/player/profile', methods=['POST'])
def get_player_profile():
    player_name = request.json.get('player', '')
    league = request.json.get('league', 'all')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT profile_data, last_updated FROM player_profiles WHERE player_name=?", (player_name,))
    row = c.fetchone()
    conn.close()
    if row:
        last_updated = datetime.fromisoformat(row[1])
        if datetime.now() - last_updated < timedelta(hours=24):
            return jsonify({"profile": json.loads(row[0]), "success": True, "cached": True})
    try:
        response = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=2000,
            system=SYSTEM_PROMPT,
            tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 5}],
            messages=[{"role": "user", "content": f"""Create dynasty player profile for {player_name}, league context: {league}. Return ONLY valid JSON:
{{"name":"","position":"","team":"","age":0,"ktc_value":0,"ktc_trend":"","position_rank":"","trade_liquidity":"High|Medium|Low","recommendation":"BUY|SELL|HOLD","recommendation_reason":"","league_specific":{{"Capital Gains":"","TRS":"","Gentlemans":"","Velvet Spade":""}},"latest_news":[],"dynasty_outlook":"","injury_history":"","contract_status":"","depth_chart":"","startup_adp":"","ppg_2025":0,"ppg_2024":0}}"""}]
        )
        assistant_message = "".join(b.text for b in response.content if hasattr(b, 'text'))
        cleaned = assistant_message.strip().lstrip('```json').lstrip('```').rstrip('```').strip()
        profile = json.loads(cleaned)
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("INSERT OR REPLACE INTO player_profiles (player_name, profile_data, last_updated) VALUES (?, ?, ?)",
                 (player_name, json.dumps(profile), datetime.now().isoformat()))
        conn.commit()
        conn.close()
        return jsonify({"profile": profile, "success": True})
    except Exception as e:
        return jsonify({"error": str(e), "success": False})

@app.route('/api/rankings/pair', methods=['GET'])
def get_ranking_pair():
    """
    Proximity-based matchup using hybrid tier system:
    - Uses KTC tier until player has 10+ comparisons, then ELO-based tier
    - 60% same tier, 30% 1 tier apart, 10% 2-3 tiers apart
    """
    position_filter = request.args.get('position', 'ALL')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    if position_filter != 'ALL':
        c.execute("""SELECT player_name, position, team, elo_score, comparisons, COALESCE(ktc_tier, 10)
                    FROM player_rankings WHERE position=?
                    ORDER BY comparisons ASC, RANDOM() LIMIT 80""", (position_filter,))
    else:
        c.execute("""SELECT player_name, position, team, elo_score, comparisons, COALESCE(ktc_tier, 10)
                    FROM player_rankings ORDER BY comparisons ASC, RANDOM() LIMIT 120""")
    rows = c.fetchall()
    conn.close()
    if len(rows) < 2:
        return jsonify({"players": [], "success": False, "error": "Not enough players"})
    players = [{"name": r[0], "position": r[1], "team": r[2],
                "elo": r[3], "comparisons": r[4], "ktc_tier": r[5]} for r in rows]

    def get_comp_tier(p):
        if p['comparisons'] < 10:
            kt = p['ktc_tier']
            if kt <= 1: return 1
            if kt <= 3: return 2
            if kt <= 6: return 3
            if kt <= 10: return 4
            if kt <= 14: return 5
            if kt <= 18: return 6
            return 7
        else:
            elo = p['elo']
            if elo >= 2800: return 1
            if elo >= 2400: return 2
            if elo >= 2000: return 3
            if elo >= 1600: return 4
            if elo >= 1200: return 5
            if elo >= 800:  return 6
            return 7

    is_filtered = position_filter != 'ALL'
    pool = sorted(players, key=lambda x: x['comparisons'])
    p1 = pool[random.randint(0, min(9, len(pool)-1))]
    p1_tier = get_comp_tier(p1)
    rand = random.random()
    if rand < 0.60:
        target_tiers = [p1_tier]
    elif rand < 0.90:
        target_tiers = [p1_tier - 1, p1_tier + 1]
    else:
        target_tiers = [p1_tier + i for i in range(-3, 4) if i != 0]
    target_tiers = [t for t in target_tiers if 1 <= t <= 7]
    p2_candidates = [p for p in players if p['name'] != p1['name'] and get_comp_tier(p) in target_tiers]
    if not p2_candidates and is_filtered:
        for spread in range(1, 4):
            expanded = [p1_tier + i for i in range(-spread, spread+1) if i != 0]
            p2_candidates = [p for p in players if p['name'] != p1['name'] and get_comp_tier(p) in [t for t in expanded if 1<=t<=7]]
            if p2_candidates: break
    if not p2_candidates:
        p2_candidates = [p for p in players if p['name'] != p1['name']]
    p2_candidates.sort(key=lambda x: x['comparisons'])
    p2 = p2_candidates[random.randint(0, min(9, len(p2_candidates)-1))]
    return jsonify({"players": [p1, p2], "success": True,
                   "tiers": [p1_tier, get_comp_tier(p2)]})

@app.route('/api/rankings/boundary_pair', methods=['GET'])
def get_boundary_pair():
    """
    Boundary Training mode: serves comparisons specifically designed
    to calibrate tier break points. Always pairs one player from the
    bottom of a tier with one from the top of the adjacent tier.
    
    Strategy:
    - Find all KTC tier boundaries (e.g. T3/T4, T6/T7, T7/T8)
    - Pick a boundary that has the least comparison data
    - Return one player from each side of that boundary
    - After ~3-5 votes per boundary, that break point is calibrated
    
    With 8 meaningful tier boundaries and 4 votes each = 32 comparisons
    to get accurate personal tiers.
    """
    position_filter = request.args.get('position', 'ALL')
    pool_type = request.args.get('pool', 'standard')  # 'standard' or 'vs'
    table = 'vs_rankings' if pool_type == 'vs' else 'player_rankings'

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    if position_filter != 'ALL':
        c.execute(f"""SELECT player_name, position, team, elo_score, comparisons, COALESCE(ktc_tier, 10)
                    FROM {table} WHERE position=? AND ktc_tier IS NOT NULL
                    ORDER BY ktc_tier ASC, elo_score DESC""", (position_filter,))
    else:
        c.execute(f"""SELECT player_name, position, team, elo_score, comparisons, COALESCE(ktc_tier, 10)
                    FROM {table} WHERE ktc_tier IS NOT NULL
                    ORDER BY ktc_tier ASC, elo_score DESC""")

    rows = c.fetchall()
    conn.close()

    if len(rows) < 4:
        return jsonify({"players": [], "success": False,
                       "error": "Not enough players with KTC tier data. Run Refresh KTC Tiers first."})

    # Group players by KTC tier
    tier_groups = defaultdict(list)
    for name, pos, team, elo, comps, kt in rows:
        tier_groups[kt].append({
            "name": name, "position": pos, "team": team,
            "elo": elo, "comparisons": comps, "ktc_tier": kt
        })

    # Find all adjacent tier boundaries that have players on both sides
    sorted_tiers = sorted(tier_groups.keys())
    boundaries = []
    for i in range(len(sorted_tiers) - 1):
        t_low = sorted_tiers[i]
        t_high = sorted_tiers[i + 1]
        low_players = tier_groups[t_low]
        high_players = tier_groups[t_high]
        if not low_players or not high_players:
            continue
        # Boundary "hardness": avg comparisons of players at this boundary
        # Pick players at the boundary edge (worst of lower tier, best of upper tier)
        # Sort: lower tier players by ELO ascending (weakest last in tier)
        # Upper tier players by ELO descending (strongest first in tier)
        low_edge = sorted(low_players, key=lambda x: x['elo'])[:3]   # bottom of lower tier
        high_edge = sorted(high_players, key=lambda x: -x['elo'])[:3]  # top of upper tier
        avg_comps = sum(p['comparisons'] for p in low_edge + high_edge) / max(len(low_edge + high_edge), 1)
        boundaries.append({
            'tier_low': t_low,
            'tier_high': t_high,
            'low_edge': low_edge,
            'high_edge': high_edge,
            'avg_comps': avg_comps,
            'boundary_label': f"KTC T{t_low} / T{t_high}",
        })

    if not boundaries:
        return jsonify({"players": [], "success": False, "error": "No tier boundaries found"})

    # Pick the boundary with fewest comparisons (most needs calibration)
    # Add slight randomness so we don't always pick same boundary
    boundaries.sort(key=lambda b: b['avg_comps'] + random.uniform(0, 2))
    chosen = boundaries[0]

    # Pick one player from each side
    p_low = random.choice(chosen['low_edge'])
    p_high = random.choice(chosen['high_edge'])

    # Count remaining boundaries that need work (avg < 5 comps)
    uncalibrated = sum(1 for b in boundaries if b['avg_comps'] < 5)
    total_boundaries = len(boundaries)
    calibrated = total_boundaries - uncalibrated

    return jsonify({
        "players": [p_low, p_high],
        "success": True,
        "boundary": chosen['boundary_label'],
        "boundary_context": f"Is the bottom of KTC T{chosen['tier_low']} better or worse than the top of KTC T{chosen['tier_high']}?",
        "calibrated_boundaries": calibrated,
        "total_boundaries": total_boundaries,
        "pct_done": round(calibrated / total_boundaries * 100) if total_boundaries else 0,
    })


@app.route('/api/rankings/vote', methods=['POST'])
def vote_ranking():
    """Record comparison vote and update ELO"""
    data = request.json
    winner = data.get('winner', '')
    loser = data.get('loser', '')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT elo_score, comparisons FROM player_rankings WHERE player_name=?", (winner,))
    wr = c.fetchone()
    c.execute("SELECT elo_score, comparisons FROM player_rankings WHERE player_name=?", (loser,))
    lr = c.fetchone()
    if wr and lr:
        K = 32
        we, le = wr[0], lr[0]
        exp_w = 1 / (1 + 10 ** ((le - we) / 400))
        new_we = we + K * (1 - exp_w)
        new_le = le + K * (0 - (1 - exp_w))
        c.execute("UPDATE player_rankings SET elo_score=?, comparisons=comparisons+1, last_updated=? WHERE player_name=?",
                 (new_we, datetime.now().isoformat(), winner))
        c.execute("UPDATE player_rankings SET elo_score=?, comparisons=comparisons+1, last_updated=? WHERE player_name=?",
                 (new_le, datetime.now().isoformat(), loser))
    conn.commit()
    conn.close()
    return jsonify({"success": True})


@app.route('/api/rankings/reset', methods=['POST'])
def reset_rankings():
    """Reset all rankings and re-seed with full player list + KTC tiers"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM player_rankings")
    conn.commit()
    conn.close()
    seed_players()
    _do_refresh_ktc_tiers()
    return jsonify({"success": True, "message": f"Rankings reset with {len(DEFAULT_PLAYERS)} players and KTC tiers applied"})

def _do_refresh_ktc_tiers():
    """Populate ktc_tier from market_data KTC source using fuzzy name matching"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT player_name, rank FROM market_data WHERE source='ktc'")
    ktc_rows = c.fetchall()

    def rank_to_ktc_tier(rank):
        if rank <= 3: return 1
        if rank <= 9: return 2
        if rank <= 19: return 3
        if rank <= 30: return 4
        if rank <= 40: return 5
        if rank <= 55: return 6
        if rank <= 75: return 7
        if rank <= 92: return 8
        if rank <= 100: return 9
        if rank <= 115: return 10
        if rank <= 125: return 11
        if rank <= 140: return 12
        if rank <= 152: return 13
        if rank <= 165: return 14
        if rank <= 175: return 15
        if rank <= 185: return 16
        return 17

    def normalize(name):
        import re as _re
        name = _re.sub(r"[^a-z0-9\s]", "", name.lower()).strip()
        name = _re.sub(r'\b(jr|sr|ii|iii|iv)\b', '', name).strip()
        return _re.sub(r'\s+', '', name)

    ktc_normalized = {normalize(name): rank_to_ktc_tier(rank) for name, rank in ktc_rows}

    for table in ['player_rankings', 'vs_rankings']:
        c.execute(f"SELECT player_name FROM {table}")
        for (player_name,) in c.fetchall():
            norm = normalize(player_name)
            tier = ktc_normalized.get(norm)
            if tier:
                c.execute(f"UPDATE {table} SET ktc_tier=? WHERE player_name=?", (tier, player_name))

    conn.commit()
    conn.close()


@app.route('/api/rankings/refresh_tiers', methods=['POST'])
def refresh_ktc_tiers():
    """Re-run KTC tier seeding on existing rankings without wiping comparison history"""
    _do_refresh_ktc_tiers()
    return jsonify({"success": True, "message": "KTC tiers refreshed on all players"})


def get_ranking_pair():
    """
    Proximity-based matchup using hybrid tier system:
    - Uses KTC tier until player has 10+ comparisons, then ELO-based tier
    - 60% same tier, 30% 1 tier apart, 10% 2-3 tiers apart
    - Relaxes constraints for position-filtered pools
    """
    position_filter = request.args.get('position', 'ALL')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    if position_filter != 'ALL':
        c.execute("""SELECT player_name, position, team, elo_score, comparisons, COALESCE(ktc_tier, 10)
                    FROM player_rankings WHERE position=?
                    ORDER BY comparisons ASC, RANDOM() LIMIT 80""", (position_filter,))
    else:
        c.execute("""SELECT player_name, position, team, elo_score, comparisons, COALESCE(ktc_tier, 10)
                    FROM player_rankings ORDER BY comparisons ASC, RANDOM() LIMIT 120""")

    rows = c.fetchall()
    conn.close()

    if len(rows) < 2:
        return jsonify({"players": [], "success": False, "error": "Not enough players"})

    players = [{"name": r[0], "position": r[1], "team": r[2],
                "elo": r[3], "comparisons": r[4], "ktc_tier": r[5]} for r in rows]

    def get_comparison_tier(p):
        """Hybrid: use KTC tier until 10+ comps, then derive from ELO rank"""
        if p['comparisons'] < 10:
            # Map KTC tiers (1-21) to comparison buckets (1-7)
            kt = p['ktc_tier']
            if kt <= 1: return 1
            if kt <= 3: return 2
            if kt <= 6: return 3
            if kt <= 10: return 4
            if kt <= 14: return 5
            if kt <= 18: return 6
            return 7
        else:
            # ELO-based after enough comparisons
            elo = p['elo']
            if elo >= 2800: return 1
            if elo >= 2400: return 2
            if elo >= 2000: return 3
            if elo >= 1600: return 4
            if elo >= 1200: return 5
            if elo >= 800: return 6
            return 7


    # Position-filtered pools are smaller — relax tier constraints
    is_filtered = position_filter != 'ALL'
    pool = sorted(players, key=lambda x: x['comparisons'])
    p1 = pool[random.randint(0, min(9, len(pool)-1))]
    p1_tier = get_comparison_tier(p1)

    rand = random.random()
    if rand < 0.60:
        target_tiers = [p1_tier]
        max_spread = 1 if is_filtered else 0
    elif rand < 0.90:
        target_tiers = [p1_tier - 1, p1_tier + 1]
        max_spread = 2 if is_filtered else 1
    else:
        target_tiers = [p1_tier + i for i in range(-3, 4) if i != 0]
        max_spread = 3

    target_tiers = [t for t in target_tiers if 1 <= t <= 7]

    p2_candidates = [p for p in players
                    if p['name'] != p1['name'] and get_comparison_tier(p) in target_tiers]

    # Fallback: expand by 1 tier at a time until we find candidates
    if not p2_candidates and is_filtered:
        for spread in range(1, 4):
            expanded = [p1_tier + i for i in range(-spread, spread+1) if i != 0]
            expanded = [t for t in expanded if 1 <= t <= 7]
            p2_candidates = [p for p in players
                            if p['name'] != p1['name'] and get_comparison_tier(p) in expanded]
            if p2_candidates:
                break

    if not p2_candidates:
        p2_candidates = [p for p in players if p['name'] != p1['name']]

    p2_candidates.sort(key=lambda x: x['comparisons'])
    p2 = p2_candidates[random.randint(0, min(9, len(p2_candidates)-1))]

    return jsonify({
        "players": [p1, p2],
        "success": True,
        "tiers": [p1_tier, get_comparison_tier(p2)]
    })

@app.route('/api/rankings/adjust', methods=['POST'])
def adjust_ranking():
    """Manually adjust a player's ELO up or down"""
    name = request.json.get('name', '')
    direction = request.json.get('direction', 'up')  # 'up' or 'down'
    positions = request.json.get('positions', 5)  # how many spots to move

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Get current rank
    c.execute("SELECT elo_score FROM player_rankings WHERE player_name=?", (name,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({"success": False, "error": "Player not found"})

    current_elo = row[0]

    # Get the ELO of the player N positions above/below
    c.execute("""SELECT elo_score FROM player_rankings
                ORDER BY elo_score DESC""")
    all_elos = [r[0] for r in c.fetchall()]

    # Find current position
    try:
        curr_pos = next(i for i, e in enumerate(all_elos) if abs(e - current_elo) < 0.01)
    except StopIteration:
        curr_pos = len(all_elos) // 2

    if direction == 'up':
        target_pos = max(0, curr_pos - positions)
    else:
        target_pos = min(len(all_elos) - 1, curr_pos + positions)

    if target_pos == curr_pos:
        conn.close()
        return jsonify({"success": True, "new_elo": current_elo})

    # Set new ELO midpoint between surrounding players
    if direction == 'up' and target_pos > 0:
        above_elo = all_elos[target_pos - 1]
        at_elo = all_elos[target_pos]
        new_elo = (above_elo + at_elo) / 2
    elif direction == 'down' and target_pos < len(all_elos) - 1:
        at_elo = all_elos[target_pos]
        below_elo = all_elos[target_pos + 1]
        new_elo = (at_elo + below_elo) / 2
    else:
        new_elo = all_elos[target_pos]

    c.execute("UPDATE player_rankings SET elo_score=?, last_updated=? WHERE player_name=?",
             (new_elo, datetime.now().isoformat(), name))
    conn.commit()
    conn.close()

    return jsonify({"success": True, "new_elo": round(new_elo), "moved_to": target_pos + 1})


@app.route('/api/rankings/reorder', methods=['POST'])
def reorder_rankings():
    """
    Accept a new ordered list of player names and assign ELO scores
    that preserve the requested order with equal spacing.
    Handles both standard ('standard') and VS ('vs') pools.
    """
    data = request.json
    ordered_names = data.get('ordered_names', [])
    pool = data.get('pool', 'standard')  # 'standard' or 'vs'
    table = 'vs_rankings' if pool == 'vs' else 'player_rankings'

    if not ordered_names:
        return jsonify({"success": False, "error": "No names provided"})

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Get current ELO range to maintain scale
    c.execute(f"SELECT MIN(elo_score), MAX(elo_score) FROM {table}")
    row = c.fetchone()
    elo_min = row[0] or 400
    elo_max = row[1] or 3000

    n = len(ordered_names)
    step = (elo_max - elo_min) / max(n, 1)

    # Assign ELOs: rank 1 gets elo_max, rank n gets elo_min
    for i, name in enumerate(ordered_names):
        new_elo = elo_max - (i * step)
        c.execute(f"UPDATE {table} SET elo_score=?, last_updated=? WHERE player_name=?",
                 (new_elo, datetime.now().isoformat(), name))

    conn.commit()
    conn.close()
    return jsonify({"success": True, "reordered": n, "pool": pool})


@app.route('/api/rankings/vote_elo', methods=['POST'])
def submit_vote():
    w_row = c.fetchone()
    c.execute("SELECT elo_score FROM player_rankings WHERE player_name=?", (loser,))
    l_row = c.fetchone()
    if w_row and l_row:
        k = 32
        w_elo, l_elo = w_row[0], l_row[0]
        exp_w = 1 / (1 + 10 ** ((l_elo - w_elo) / 400))
        exp_l = 1 / (1 + 10 ** ((w_elo - l_elo) / 400))
        c.execute("UPDATE player_rankings SET elo_score=?, comparisons=comparisons+1, last_updated=? WHERE player_name=?",
                 (w_elo + k * (1 - exp_w), datetime.now().isoformat(), winner))
        c.execute("UPDATE player_rankings SET elo_score=?, comparisons=comparisons+1, last_updated=? WHERE player_name=?",
                 (l_elo + k * (0 - exp_l), datetime.now().isoformat(), loser))
        conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route('/api/rankings/list', methods=['GET'])
def get_rankings():
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT player_name, position, team, elo_score, comparisons FROM player_rankings ORDER BY elo_score DESC LIMIT 200")
    rows = c.fetchall()
    conn.close()
    rankings = [{"rank": i+1, "name": r[0], "position": r[1], "team": r[2], "elo": round(r[3]), "comparisons": r[4]} for i, r in enumerate(rows)]
    return jsonify({"rankings": rankings, "success": True})

@app.route('/api/rankings/add', methods=['POST'])
def add_player():
    name = request.json.get('name', '')
    position = request.json.get('position', '')
    team = request.json.get('team', '')
    if not name:
        return jsonify({"success": False, "error": "Name required"})
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("INSERT OR IGNORE INTO player_rankings (player_name, position, team, elo_score, comparisons, last_updated) VALUES (?, ?, ?, 1500, 0, ?)",
             (name, position, team, datetime.now().isoformat()))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route('/api/draft/chart', methods=['GET'])
def get_draft_chart():
    chart = {
        "pick_values": {
            "1.01":10000,"1.02":9500,"1.03":8800,"1.04":7800,"1.05":7000,"1.06":6300,
            "1.07":5600,"1.08":5000,"1.09":4400,"1.10":3900,"1.11":3400,"1.12":2900,
            "2.01":2600,"2.02":2400,"2.03":2200,"2.04":2050,"2.05":1900,"2.06":1750,
            "2.07":1600,"2.08":1450,"2.09":1300,"2.10":1150,"2.11":1000,"2.12":850,
            "3.01":800,"3.02":750,"3.03":700,"3.04":650,"3.05":600,"3.06":560,
            "3.07":520,"3.08":480,"3.09":440,"3.10":410,"3.11":380,"3.12":350,
            "4.01":320,"4.02":300,"4.03":280,"4.04":260,"4.05":245,"4.06":230,
            "4.07":215,"4.08":200,"4.09":185,"4.10":172,"4.11":160,"4.12":148
        },
        "my_picks": ["1.02","2.11","3.11","4.02","5.11","6.02","7.11","8.02","9.11","10.02"],
        "scoring_notes": "6pt TD format. Elite QBs (Maye, Allen, Lamar) worth ~20% premium.",
        "execute_threshold": 0.15
    }
    return jsonify({"chart": chart, "success": True})

@app.route('/api/kb/import', methods=['POST'])
def import_knowledge_base():
    """
    Parse uploaded Excel knowledge base and store all key data in DB.
    Called once on upload — data persists and is injected into every subsequent request.
    Handles: My Rosters, GM Tendencies, Trade Log, Traded Picks, league roster tabs.
    """
    data     = request.json
    file_obj = data.get('file_data', {})
    raw      = file_obj.get('data', '')
    fname    = file_obj.get('name', 'knowledge_base.xlsx')

    if not raw:
        return jsonify({"success": False, "error": "No file data received"})

    try:
        import io, base64 as b64lib
        from openpyxl import load_workbook as _lwb
        b64 = raw.split(',')[1] if ',' in raw else raw
        wb_xl = _lwb(io.BytesIO(b64lib.b64decode(b64)), data_only=True)
    except Exception as e:
        return jsonify({"success": False, "error": f"Could not open Excel file: {e}"})

    conn = sqlite3.connect(DB_PATH)
    c    = conn.cursor()
    now  = datetime.now().isoformat()
    results = {}

    def cell(row, col):
        v = row[col] if col < len(row) else None
        return str(v).strip() if v is not None and str(v).strip() not in ('None','') else ''

    # ── PARSE GM TENDENCIES TAB ───────────────────────────────────────────────
    gm_sheets = [s for s in wb_xl.sheetnames
                 if any(k in s.lower() for k in ['tend', 'gm', 'manager'])]
    gm_count = 0
    for sname in gm_sheets:
        ws = wb_xl[sname]
        headers = [str(ws.cell(1, i).value or '').lower() for i in range(1, 15)]
        def hcol(keywords):
            for kw in keywords:
                for i, h in enumerate(headers):
                    if kw in h: return i
            return None
        mgr_c    = hcol(['manager','name'])
        league_c = hcol(['league'])
        window_c = hcol(['window'])
        style_c  = hcol(['style'])
        buy_c    = hcol(['buy','loves buy'])
        sell_c   = hcol(['sell','loves sell'])
        never_c  = hcol(['never'])
        over_c   = hcol(['overpay','over'])
        notes_c  = hcol(['notes','intel'])

        if mgr_c is None: continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            mgr = cell(row, mgr_c)
            if not mgr or mgr.lower() in ('manager','name'): continue
            league = cell(row, league_c) if league_c is not None else ''
            c.execute("""INSERT OR REPLACE INTO gm_tendencies
                        (manager,league,window,style,loves_buy,loves_sell,
                         never_sells,overpays_for,trade_notes,last_updated)
                        VALUES (?,?,?,?,?,?,?,?,?,?)""", (
                mgr, league,
                cell(row, window_c) if window_c is not None else '',
                cell(row, style_c)  if style_c  is not None else '',
                cell(row, buy_c)    if buy_c    is not None else '',
                cell(row, sell_c)   if sell_c   is not None else '',
                cell(row, never_c)  if never_c  is not None else '',
                cell(row, over_c)   if over_c   is not None else '',
                cell(row, notes_c)  if notes_c  is not None else '',
                now))
            gm_count += 1
    results['gm_tendencies'] = gm_count

    # ── PARSE TRADE LOG TAB ───────────────────────────────────────────────────
    trade_sheets = [s for s in wb_xl.sheetnames
                    if any(k in s.lower() for k in ['trade log','trade activity','log'])]
    trade_count = 0
    for sname in trade_sheets:
        ws = wb_xl[sname]
        headers = [str(ws.cell(1, i).value or '').lower() for i in range(1, 12)]
        def hcol2(keywords):
            for kw in keywords:
                for i, h in enumerate(headers):
                    if kw in h: return i
            return None
        c.execute("DELETE FROM trade_log")
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row): continue
            league  = cell(row, hcol2(['league'])) if hcol2(['league']) is not None else ''
            date    = cell(row, hcol2(['date']))    if hcol2(['date'])   is not None else ''
            them    = cell(row, hcol2(['their','manager','team'])) if hcol2(['their','manager','team']) is not None else ''
            want_me = cell(row, hcol2(['they want','their ask','want'])) if hcol2(['they want','their ask','want']) is not None else ''
            want_i  = cell(row, hcol2(['i want','my target'])) if hcol2(['i want','my target']) is not None else ''
            status  = cell(row, hcol2(['status'])) if hcol2(['status']) is not None else ''
            move    = cell(row, hcol2(['next move','my next'])) if hcol2(['next move','my next']) is not None else ''
            notes   = cell(row, hcol2(['notes'])) if hcol2(['notes']) is not None else ''
            if not any([league, them, want_me, want_i]): continue
            c.execute("""INSERT INTO trade_log
                        (league,date,their_manager,what_they_want,what_i_want,
                         status,my_next_move,notes,last_updated)
                        VALUES (?,?,?,?,?,?,?,?,?)""",
                     (league,date,them,want_me,want_i,status,move,notes,now))
            trade_count += 1
    results['trade_log'] = trade_count

    # ── PARSE MY ROSTERS TAB ──────────────────────────────────────────────────
    roster_sheets = [s for s in wb_xl.sheetnames
                     if any(k in s.lower() for k in ['my roster','my asset','⭐'])]
    roster_count = 0
    for sname in roster_sheets:
        ws = wb_xl[sname]
        headers = [str(ws.cell(1, i).value or '').lower() for i in range(1, 12)]
        def hcol3(keywords):
            for kw in keywords:
                for i, h in enumerate(headers):
                    if kw in h: return i
            return None
        league_c = hcol3(['league'])
        name_c   = hcol3(['player','name'])
        pos_c    = hcol3(['pos','position'])
        team_c   = hcol3(['team','nfl'])
        val_c    = hcol3(['my value','ktc','value'])
        if name_c is None: continue
        current_league = ''
        for row in ws.iter_rows(min_row=2, values_only=True):
            if all(v is None for v in row): continue
            name = cell(row, name_c)
            # Detect league header rows
            if not name and league_c is not None:
                lg = cell(row, league_c)
                if lg: current_league = lg
                continue
            if not name: continue
            pos   = cell(row, pos_c)   if pos_c   is not None else ''
            team  = cell(row, team_c)  if team_c  is not None else ''
            league = cell(row, league_c) if league_c is not None else current_league
            my_val = 0
            if val_c is not None:
                try: my_val = int(float(str(row[val_c] or 0)))
                except: pass
            # Update player_values table with any new values
            if my_val > 0:
                c.execute("""INSERT OR REPLACE INTO player_values
                            (player_name,position,my_value,ktc_value,delta,tier,last_updated)
                            VALUES (?,?,?,COALESCE((SELECT ktc_value FROM player_values WHERE player_name=?),0),
                            ?-COALESCE((SELECT ktc_value FROM player_values WHERE player_name=?),0),'',?)""",
                         (name,pos,my_val,name,my_val,name,now))
            roster_count += 1
    results['my_rosters'] = roster_count

    # ── PARSE GM TENDENCIES FROM FFPC TABS ────────────────────────────────────
    # Also parse roster tabs for Capital Gains and TRS
    ffpc_sheets = [s for s in wb_xl.sheetnames
                   if any(k in s.lower() for k in ['capital gains','trs','ffpc'])]
    ffpc_count = 0
    for sname in ffpc_sheets:
        ws = wb_xl[sname]
        league_name = 'Capital Gains' if 'capital' in sname.lower() else 'TRS'
        current_mgr = ''
        c.execute("DELETE FROM league_rosters WHERE league=?", (league_name,))
        for row in ws.iter_rows(min_row=3, values_only=True):
            if all(v is None for v in row): continue
            first = str(row[0] or '').strip()
            if not first: continue
            # Manager header rows have their name and "— X players"
            if '—' in first or 'players' in first.lower():
                current_mgr = first.split('—')[0].strip().replace('YOUR TEAM','').strip()
                continue
            # Player rows
            if current_mgr and len(row) >= 3:
                name = str(row[1] or '').strip() if len(row) > 1 else ''
                pos  = str(row[2] or '').strip() if len(row) > 2 else ''
                team = str(row[3] or '').strip() if len(row) > 3 else ''
                if name and pos in ('QB','RB','WR','TE','K','DEF'):
                    c.execute("""INSERT INTO league_rosters
                                (league,manager,player_name,position,team,last_updated)
                                VALUES (?,?,?,?,?,?)""",
                             (league_name, current_mgr, name, pos, team, now))
                    ffpc_count += 1
    results['ffpc_rosters'] = ffpc_count

    # ── UPDATE KB META ────────────────────────────────────────────────────────
    import json as _json
    c.execute("""INSERT OR REPLACE INTO kb_meta
                (id, last_upload, filename, tabs_parsed, row_counts)
                VALUES (1,?,?,?,?)""",
             (now, fname,
              _json.dumps(list(wb_xl.sheetnames)),
              _json.dumps(results)))
    conn.commit()
    conn.close()

    total = sum(results.values())
    return jsonify({
        "success": True,
        "message": f"Knowledge base imported: {total} records stored across {len(results)} categories",
        "details": results,
        "tabs_found": list(wb_xl.sheetnames),
        "last_updated": now
    })


@app.route('/api/kb/status', methods=['GET'])
def kb_status():
    """Return metadata about the last imported knowledge base."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT last_upload, filename, tabs_parsed, row_counts FROM kb_meta WHERE id=1")
    row = c.fetchone()
    c.execute("SELECT COUNT(*) FROM gm_tendencies")
    gm_count = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM trade_log")
    trade_count = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM player_values")
    val_count = c.fetchone()[0]
    conn.close()
    if not row:
        return jsonify({"success": True, "imported": False,
                        "message": "No knowledge base imported yet"})
    import json as _json
    return jsonify({
        "success": True, "imported": True,
        "last_upload": row[0], "filename": row[1],
        "tabs_parsed": _json.loads(row[2] or '[]'),
        "row_counts":  _json.loads(row[3] or '{}'),
        "gm_tendencies": gm_count,
        "trade_log": trade_count,
        "player_values": val_count
    })


def get_kb_context():
    """
    Build full knowledge base context string injected into EVERY request.
    Pulls GM tendencies, active trade log, and pick capital from DB.
    """
    conn = sqlite3.connect(DB_PATH)
    c    = conn.cursor()

    # GM Tendencies
    c.execute("""SELECT manager, league, window, style, loves_buy, loves_sell,
                        never_sells, overpays_for, trade_notes
                 FROM gm_tendencies ORDER BY league, manager""")
    gm_rows = c.fetchall()

    # Active trades
    c.execute("""SELECT league, their_manager, what_they_want, what_i_want,
                        status, my_next_move, notes
                 FROM trade_log
                 WHERE status NOT IN ('Closed','Done','Completed','')
                 ORDER BY league""")
    trade_rows = c.fetchall()

    # KB meta
    c.execute("SELECT last_upload, filename FROM kb_meta WHERE id=1")
    meta = c.fetchone()

    conn.close()

    if not gm_rows and not trade_rows:
        return ""

    lines = []
    if meta:
        lines.append(f"\n━━ KNOWLEDGE BASE (imported: {meta[0][:10] if meta[0] else 'unknown'} | {meta[1]}) ━━")

    # GM Tendencies grouped by league
    if gm_rows:
        lines.append("\nGM TENDENCIES:")
        current_league = None
        for mgr, league, window, style, buy, sell, never, over, notes in gm_rows:
            if league != current_league:
                lines.append(f"\n  [{league}]")
                current_league = league
            parts = []
            if window: parts.append(f"Window:{window}")
            if style:  parts.append(f"Style:{style}")
            if buy:    parts.append(f"Buys:{buy}")
            if sell:   parts.append(f"Sells:{sell}")
            if never:  parts.append(f"Never:{never}")
            if over:   parts.append(f"Overpays:{over}")
            if notes:  parts.append(f"Intel:{notes}")
            lines.append(f"  {mgr} — {' | '.join(parts)}" if parts else f"  {mgr}")

    # Active trade negotiations
    if trade_rows:
        lines.append("\nACTIVE TRADE NEGOTIATIONS:")
        for league, them, want_me, want_i, status, move, notes in trade_rows:
            lines.append(f"  [{league}] {them}: They want {want_me} | I want {want_i} | "
                        f"Status:{status} | Next:{move}" + (f" | {notes}" if notes else ""))

    return '\n'.join(lines)



@app.route('/api/debug/rosters', methods=['GET'])
def debug_rosters():
    """Debug: show what's in league_rosters table."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='league_rosters'")
    has_table = c.fetchone() is not None
    if not has_table:
        conn.close()
        return jsonify({"has_table": False, "message": "league_rosters table does not exist yet — sync Sleeper first"})
    c.execute("SELECT league, manager, COUNT(*) as cnt FROM league_rosters GROUP BY league, manager ORDER BY league, cnt DESC")
    rows = c.fetchall()
    conn.close()
    return jsonify({"has_table": True,
                    "summary": [{"league": r[0], "manager": r[1], "players": r[2]} for r in rows]})



@app.route('/api/sleeper/sync_rosters', methods=['POST'])
def sync_all_rosters():
    total_synced = 0
    errors = []

    try:
        players_r = requests.get("https://api.sleeper.app/v1/players/nfl", timeout=30)
        players_db = players_r.json() if players_r.status_code == 200 else {}
    except Exception as e:
        return jsonify({"success": False, "error": f"Could not fetch player DB: {e}"})

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    for league_key, league_id in SLEEPER_LEAGUE_IDS.items():
        try:
            users_r   = requests.get(f"https://api.sleeper.app/v1/league/{league_id}/users", timeout=10)
            rosters_r = requests.get(f"https://api.sleeper.app/v1/league/{league_id}/rosters", timeout=10)
            if users_r.status_code != 200 or rosters_r.status_code != 200:
                errors.append(f"{league_key}: API error")
                continue

            users   = users_r.json()
            rosters = rosters_r.json()
            user_map = {u['user_id']: u.get('display_name', u.get('username', '?'))
                        for u in users}

            # Clear old roster data for this league
            c.execute("DELETE FROM league_rosters WHERE league=?", (league_key,))

            for roster in rosters:
                mgr = user_map.get(roster.get('owner_id', ''), f"Team{roster['roster_id']}")
                for pid in (roster.get('players', []) or []):
                    p = players_db.get(pid, {})
                    name = f"{p.get('first_name','')} {p.get('last_name','')}".strip() or pid
                    pos  = p.get('position', '?')
                    team = p.get('team', 'FA') or 'FA'
                    c.execute("""INSERT INTO league_rosters
                                (league, manager, player_name, position, team, last_updated)
                                VALUES (?,?,?,?,?,?)""",
                             (league_key, mgr, name, pos, team, now))
                    total_synced += 1

        except Exception as e:
            errors.append(f"{league_key}: {str(e)}")

    conn.commit()
    conn.close()
    return jsonify({"success": True, "synced": total_synced, "errors": errors,
                    "message": f"Synced {total_synced} player-roster records across Sleeper leagues"})


def get_league_roster_context(league_key):
    """
    Build a compact roster summary for all managers in a league.
    Used to inject into trade analysis so the model knows who owns what.
    Returns empty string if no data synced yet.
    """
    # Map display league names to DB keys
    league_map = {
        'Velvet Spade':        'velvet_spade',
        'velvet_spade':        'velvet_spade',
        "Gentleman's Dynasty": 'gentlemans_dynasty',
        'gentlemans_dynasty':  'gentlemans_dynasty',
    }
    db_key = league_map.get(league_key, league_key.lower().replace(' ', '_').replace("'", ''))

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""SELECT manager, position, player_name, team
                 FROM league_rosters WHERE league=?
                 ORDER BY manager, position, player_name""", (db_key,))
    rows = c.fetchall()
    c.execute("SELECT MAX(last_updated) FROM league_rosters WHERE league=?", (db_key,))
    last_sync = c.fetchone()[0]
    conn.close()

    if not rows:
        return ""

    # Group by manager
    mgr_rosters = defaultdict(lambda: defaultdict(list))
    for mgr, pos, name, team in rows:
        mgr_rosters[mgr][pos].append(f"{name}({team})")

    lines = [f"\n{league_key.upper()} FULL LEAGUE ROSTERS (synced: {last_sync[:10] if last_sync else 'unknown'}):"]
    lines.append("Use this to identify who owns each player when analyzing trade blocks.")
    for mgr, positions in sorted(mgr_rosters.items()):
        pos_parts = []
        for pos in ['QB','RB','WR','TE']:
            if pos in positions:
                pos_parts.append(f"{pos}: {', '.join(positions[pos])}")
        lines.append(f"\n{mgr}: {' | '.join(pos_parts)}")

    return '\n'.join(lines)



@app.route('/api/sleeper/my_roster', methods=['POST'])
def my_sleeper_roster():
    """Fetch dcatlet's roster from a specific Sleeper league, with player names and KTC values."""
    data        = request.json
    league_key  = data.get('league', 'velvet_spade')
    my_username = 'dcatlet'

    league_id = SLEEPER_LEAGUE_IDS.get(league_key)
    if not league_id:
        return jsonify({"success": False, "error": f"Unknown league: {league_key}"})

    try:
        # Get player database
        players_r = requests.get("https://api.sleeper.app/v1/players/nfl", timeout=30)
        players_db = players_r.json() if players_r.status_code == 200 else {}

        # Get users to find dcatlet's user_id
        users_r = requests.get(f"https://api.sleeper.app/v1/league/{league_id}/users", timeout=10)
        if users_r.status_code != 200:
            return jsonify({"success": False, "error": "Could not fetch league users"})
        users = users_r.json()
        my_user_id = next(
            (u['user_id'] for u in users
             if u.get('username','').lower() == my_username.lower()
             or u.get('display_name','').lower() == my_username.lower()),
            None
        )
        if not my_user_id:
            return jsonify({"success": False,
                "error": f"Could not find {my_username} in league. Users: {[u.get('username') for u in users]}"})

        # Get rosters
        rosters_r = requests.get(f"https://api.sleeper.app/v1/league/{league_id}/rosters", timeout=10)
        if rosters_r.status_code != 200:
            return jsonify({"success": False, "error": "Could not fetch rosters"})
        rosters = rosters_r.json()

        my_roster = next((r for r in rosters if r.get('owner_id') == my_user_id), None)
        if not my_roster:
            return jsonify({"success": False, "error": "Could not find your roster"})

        # Pull values from DB — three tiers:
        # 1. player_values (your composite — most accurate)
        # 2. market_data KTC source (from KTC paste — broader coverage)
        # 3. None — show as unranked
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT player_name, my_value, ktc_value FROM player_values")
        pv_rows = c.fetchall()
        c.execute("SELECT player_name, value FROM market_data WHERE source='ktc'")
        md_rows = c.fetchall()
        conn.close()

        def normalize(s):
            import re as _re
            return _re.sub(r"[^a-z0-9]", "", s.lower()) if s else ""

        # Build lookups
        pv_map   = {row[0].lower(): row[2] for row in pv_rows}   # name → ktc_value
        pv_norm  = {normalize(row[0]): row[2] for row in pv_rows}
        md_map   = {row[0].lower(): row[1] for row in md_rows}    # name → ktc value from paste
        md_norm  = {normalize(row[0]): row[1] for row in md_rows}

        def get_value(name):
            if not name: return None
            nl = name.lower()
            nn = normalize(name)
            # Tier 1: player_values exact
            if nl in pv_map:   return pv_map[nl]
            if nn in pv_norm:  return pv_norm[nn]
            # Tier 2: market_data KTC paste
            if nl in md_map:   return md_map[nl]
            if nn in md_norm:  return md_norm[nn]
            return None

        taxi_ids = set(my_roster.get('taxi', []) or [])
        ir_ids   = set(my_roster.get('reserve', []) or [])

        players_out = []
        for pid in (my_roster.get('players', []) or []):
            p = players_db.get(pid, {})
            name = f"{p.get('first_name','')} {p.get('last_name','')}".strip()
            if not name: name = pid
            players_out.append({
                'name':      name,
                'position':  p.get('position', '?'),
                'team':      p.get('team', 'FA') or 'FA',
                'age':       p.get('age'),
                'ktc_value': get_value(name),
                'taxi':      pid in taxi_ids,
                'ir':        pid in ir_ids,
            })

        # Sort by position priority
        pos_order = {'QB':0,'RB':1,'WR':2,'TE':3,'K':4,'DEF':5}
        players_out.sort(key=lambda x: (pos_order.get(x['position'], 9), -(x['ktc_value'] or 0), x['name']))

        # ── WRITE ALL ROSTERS TO DB ──────────────────────────────────────────
        # Save every manager's roster so chat has full league context
        now = datetime.now().isoformat()
        try:
            conn2 = sqlite3.connect(DB_PATH)
            c2 = conn2.cursor()
            # Clear old data for this league
            c2.execute("DELETE FROM league_rosters WHERE league=?", (league_key,))
            # Write all managers
            for roster in rosters:
                owner_id = roster.get('owner_id', '')
                mgr_name = user_map.get(owner_id, f"Team{roster['roster_id']}")
                # Mark dcatlet's team with arrow for easy identification
                if owner_id == my_user_id:
                    mgr_name = mgr_name  # keep clean for dcatlet
                for pid in (roster.get('players', []) or []):
                    p = players_db.get(pid, {})
                    name = f"{p.get('first_name','')} {p.get('last_name','')}".strip()
                    if not name: name = pid
                    pos  = p.get('position', '?')
                    team = p.get('team', 'FA') or 'FA'
                    c2.execute("""INSERT INTO league_rosters
                                  (league, manager, player_name, position, team, last_updated)
                                  VALUES (?,?,?,?,?,?)""",
                               (league_key, mgr_name, name, pos, team, now))
            conn2.commit()
            conn2.close()
        except Exception as e:
            pass  # DB write failure shouldn't break roster display

        # Get traded picks for this roster
        picks_r = requests.get(f"https://api.sleeper.app/v1/league/{league_id}/traded_picks", timeout=10)
        picks_raw = picks_r.json() if picks_r.status_code == 200 else []
        user_map = {u['user_id']: u.get('display_name', u['user_id']) for u in users}
        my_picks = []
        for pk in picks_raw:
            if pk.get('owner_id') == my_user_id:
                orig = user_map.get(pk.get('previous_owner_id', ''), 'Unknown')
                rnd  = pk.get('round', '?')
                yr   = pk.get('season', '')
                label = f"{yr} R{rnd}" + (f" (from {orig})" if orig != my_username else '')
                my_picks.append({'label': label, 'round': rnd, 'season': yr})

        return jsonify({
            "success":  True,
            "league":   league_key,
            "total":    len(players_out),
            "players":  players_out,
            "picks":    my_picks,
        })

    except Exception as e:
        return jsonify({"success": False, "error": str(e)})



@app.route('/api/sleeper/sync', methods=['POST'])
def sync_sleeper():
    try:
        url = "https://api.sleeper.app/v1/players/nfl"
        r = requests.get(url, timeout=30)
        players_db = r.json() if r.status_code == 200 else {}
    except:
        players_db = {}
    for league_name, league_id in SLEEPER_LEAGUE_IDS.items():
        try:
            rosters_r = requests.get(f"https://api.sleeper.app/v1/league/{league_id}/rosters", timeout=10)
            users_r = requests.get(f"https://api.sleeper.app/v1/league/{league_id}/users", timeout=10)
            if rosters_r.status_code == 200 and users_r.status_code == 200:
                rosters = rosters_r.json()
                users = users_r.json()
                user_map = {u['user_id']: u.get('display_name', u['user_id']) for u in users}
                league_rosters = []
                for roster in rosters:
                    owner_name = user_map.get(roster.get('owner_id'), 'Unknown')
                    player_names = []
                    for pid in roster.get('players', []):
                        if pid in players_db:
                            p = players_db[pid]
                            name = f"{p.get('first_name','')} {p.get('last_name','')}".strip()
                            player_names.append(f"{name} ({p.get('position','')}-{p.get('team','FA')})")
                    league_rosters.append({"owner": owner_name, "players": player_names})
                results[league_name] = league_rosters
        except Exception as e:
            results[league_name] = {"error": str(e)}
    return jsonify({"rosters": results, "success": True})

# ============================================================
# KTC SYNC — UNOFFICIAL API + MANUAL FALLBACK
# ============================================================

@app.route('/api/ktc/sync', methods=['POST'])
def sync_ktc():
    """Try KTC unofficial API, store values with timestamp"""
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (iPhone; CPU iPhone OS 16_0 like Mac OS X) AppleWebKit/605.1.15',
            'Accept': 'application/json',
            'Referer': 'https://keeptradecut.com/'
        }
        r = requests.get(
            'https://keeptradecut.com/api/players?format=2&limited=false',
            headers=headers, timeout=15
        )
        if r.status_code == 200:
            players = r.json()
            conn = sqlite3.connect(DB_PATH)
            c = conn.cursor()
            c.execute('''CREATE TABLE IF NOT EXISTS ktc_values
                        (id INTEGER PRIMARY KEY AUTOINCREMENT,
                         player_name TEXT NOT NULL, ktc_value INTEGER,
                         week_key TEXT NOT NULL, synced_at TEXT NOT NULL)''')
            week_key = get_week_key()
            count = 0
            for p in players:
                name = p.get('playerName', '') or f"{p.get('firstName','')} {p.get('lastName','')}".strip()
                value = p.get('superflex', p.get('value', 0))
                if name and value:
                    c.execute('''INSERT OR REPLACE INTO ktc_values (player_name, ktc_value, week_key, synced_at)
                                VALUES (?, ?, ?, ?)''',
                             (name, value, week_key, datetime.now().isoformat()))
                    count += 1
            conn.commit()
            conn.close()
            return jsonify({"success": True, "source": "ktc_api", "count": count,
                          "message": f"Synced {count} KTC values from live API"})
    except Exception as e:
        pass

    # Fallback — return baseline values from context
    return jsonify({"success": True, "source": "baseline",
                   "message": "KTC API unavailable. Using May 2026 baseline values. Paste updated values in chat to override."})

@app.route('/api/ktc/values', methods=['GET'])
def get_ktc_values():
    """Get stored KTC values with 4-week trend"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    try:
        c.execute('''CREATE TABLE IF NOT EXISTS ktc_values
                    (id INTEGER PRIMARY KEY AUTOINCREMENT,
                     player_name TEXT NOT NULL, ktc_value INTEGER,
                     week_key TEXT NOT NULL, synced_at TEXT NOT NULL)''')
        conn.commit()
        c.execute('''SELECT player_name, ktc_value, week_key FROM ktc_values
                    ORDER BY player_name, synced_at DESC''')
        rows = c.fetchall()
        conn.close()

        # Build player value history
        history = defaultdict(list)
        for name, val, week in rows:
            history[name].append({'value': val, 'week': week})

        result = {}
        for name, entries in history.items():
            latest = entries[0]['value']
            baseline = entries[-1]['value'] if len(entries) > 1 else latest
            change = latest - baseline
            result[name] = {
                'current': latest,
                'baseline': baseline,
                'change': change,
                'change_pct': round((change / baseline * 100) if baseline else 0, 1),
                'weeks_tracked': len(entries)
            }
        return jsonify({"values": result, "success": True})
    except Exception as e:
        conn.close()
        return jsonify({"values": {}, "success": True})

@app.route('/api/ktc/manual', methods=['POST'])
def manual_ktc_update():
    """Allow manual paste of KTC values"""
    updates = request.json.get('updates', [])
    week_key = get_week_key()
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS ktc_values
                (id INTEGER PRIMARY KEY AUTOINCREMENT,
                 player_name TEXT NOT NULL, ktc_value INTEGER,
                 week_key TEXT NOT NULL, synced_at TEXT NOT NULL)''')
    for item in updates:
        c.execute('''INSERT INTO ktc_values (player_name, ktc_value, week_key, synced_at)
                    VALUES (?, ?, ?, ?)''',
                 (item['name'], item['value'], week_key, datetime.now().isoformat()))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "updated": len(updates)})

# ============================================================
# SLEEPER ADP — FREE PUBLIC API
# ============================================================

@app.route('/api/sleeper/adp', methods=['GET'])
def get_sleeper_adp():
    """Get Sleeper dynasty ADP as ranking source"""
    try:
        # Sleeper trending players as ADP proxy
        r = requests.get('https://api.sleeper.app/v1/players/nfl/trending/add?lookback_hours=168&limit=200',
                        timeout=10)
        if r.status_code == 200:
            data = r.json()
            adp = {}
            for i, item in enumerate(data):
                pid = item.get('player_id', '')
                adp[pid] = {'rank': i + 1, 'adds': item.get('count', 0)}
            return jsonify({"adp": adp, "success": True, "count": len(adp)})
    except Exception as e:
        pass
    return jsonify({"adp": {}, "success": False})

# ============================================================
# COMPOSITE RANKINGS
# Weights: Personal ELO 40% | KTC 20% | Sleeper ADP 30% | Underdog 10%
# ============================================================

@app.route('/api/rankings/composite', methods=['GET'])
def get_composite_rankings():
    position_filter = request.args.get('position', 'ALL')

    # 1. Get personal ELO rankings
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''SELECT player_name, position, team, elo_score, comparisons
                FROM player_rankings ORDER BY elo_score DESC''')
    personal_rows = c.fetchall()
    conn.close()

    personal_ranks = {}
    for i, (name, pos, team, elo, comps) in enumerate(personal_rows):
        personal_ranks[name] = {
            'position': pos, 'team': team, 'elo': elo,
            'comparisons': comps, 'personal_rank': i + 1
        }

    # 2. Get KTC values from DB
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    ktc_ranks = {}
    try:
        c.execute('''CREATE TABLE IF NOT EXISTS ktc_values
                    (id INTEGER PRIMARY KEY AUTOINCREMENT,
                     player_name TEXT NOT NULL, ktc_value INTEGER,
                     week_key TEXT NOT NULL, synced_at TEXT NOT NULL)''')
        c.execute('''SELECT player_name, MAX(ktc_value) as val FROM ktc_values
                    GROUP BY player_name ORDER BY val DESC''')
        ktc_rows = c.fetchall()
        for i, (name, val) in enumerate(ktc_rows):
            ktc_ranks[name] = {'value': val, 'rank': i + 1}
    except:
        pass
    conn.close()

    # 3. Sleeper ADP placeholder (removed live fetch — unreliable from datacenter)
    sleeper_ranks = {}

    # 4. Build composite scores
    all_players = set(personal_ranks.keys())
    total = len(all_players)

    composite = []
    for name in all_players:
        p = personal_ranks[name]
        pos = p['position']

        if position_filter != 'ALL' and pos != position_filter:
            continue

        personal_rank = p['personal_rank']
        personal_score = (total - personal_rank + 1) / total

        # KTC score
        if name in ktc_ranks:
            ktc_rank = ktc_ranks[name]['rank']
            ktc_total = len(ktc_ranks)
            ktc_score = (ktc_total - ktc_rank + 1) / ktc_total
        else:
            ktc_score = personal_score * 0.7  # estimate from personal

        # Sleeper ADP score
        if name in sleeper_ranks:
            sl_rank = sleeper_ranks[name]
            sl_total = max(len(sleeper_ranks), 200)
            sleeper_score = (sl_total - sl_rank + 1) / sl_total
        else:
            sleeper_score = personal_score * 0.7

        # Composite: Personal 40%, KTC 20%, Sleeper 30%, buffer 10%
        composite_score = (
            personal_score * 0.40 +
            ktc_score * 0.20 +
            sleeper_score * 0.30 +
            personal_score * 0.10  # buffer uses personal as proxy
        )

        # Market rank (KTC-weighted)
        market_rank_score = ktc_score * 0.50 + sleeper_score * 0.50
        market_rank = 0  # will be assigned after sorting

        composite.append({
            'name': name,
            'position': pos,
            'team': p['team'],
            'elo': round(p['elo']),
            'comparisons': p['comparisons'],
            'personal_rank': personal_rank,
            'ktc_value': ktc_ranks.get(name, {}).get('value', 0),
            'ktc_rank': ktc_ranks.get(name, {}).get('rank', 0),
            'sleeper_rank': sleeper_ranks.get(name, 0),
            'composite_score': composite_score,
            'market_score': market_rank_score,
        })

    # Sort by composite score
    composite.sort(key=lambda x: x['composite_score'], reverse=True)

    # Assign market ranks
    market_sorted = sorted(composite, key=lambda x: x['market_score'], reverse=True)
    market_rank_map = {p['name']: i + 1 for i, p in enumerate(market_sorted)}

    for i, p in enumerate(composite):
        p['composite_rank'] = i + 1
        p['market_rank'] = market_rank_map[p['name']]
        delta = p['market_rank'] - p['personal_rank']
        p['delta'] = delta  # positive = you rank higher than market (buy signal)
        if abs(delta) >= 26:
            p['signal'] = 'MAJOR BUY' if delta > 0 else 'MAJOR SELL'
            p['signal_strength'] = 3
        elif abs(delta) >= 16:
            p['signal'] = 'BUY' if delta > 0 else 'SELL'
            p['signal_strength'] = 2
        elif abs(delta) >= 5:
            p['signal'] = 'SLIGHT BUY' if delta > 0 else 'SLIGHT SELL'
            p['signal_strength'] = 1
        else:
            p['signal'] = 'HOLD'
            p['signal_strength'] = 0

    return jsonify({
        "rankings": composite,
        "total": len(composite),
        "success": True,
        "weights": {"personal": "40%", "ktc": "20%", "sleeper": "30%", "buffer": "10%"}
    })

# ============================================================
# KNOWLEDGE BASE — YOUTUBE TRANSCRIPTS
# ============================================================

@app.route('/api/draft/assistant', methods=['POST'])
def draft_assistant():
    """Live startup draft assistant for Velvet Spade"""
    data = request.json
    picks_made = data.get('picks_made', [])  # list of {pick: "1.01", player: "Josh Allen", team: "pdwyer13"}
    current_pick = data.get('current_pick', '')
    my_roster = data.get('my_roster', [])
    available_input = data.get('available', '')

    # Build context for the agent
    picks_context = '\n'.join([f"Pick {p['pick']} ({p.get('team','')}): {p['player']}" for p in picks_made])
    roster_context = ', '.join(my_roster) if my_roster else 'None yet'

    prompt = f"""You are running the Velvet Spade startup draft assistant for MJBrutus (dcatlet).

VELVET SPADE SCORING: 6pt passing TDs | 1.5 TE premium | 1.0 PPR RB/WR | Bonuses: 40yd play +1, 400yd pass +2, 200yd rush/rec +2
LINEUP: 1QB/2RB/2WR/1TE/3FLEX/1SFLX | 23+5taxi+2IR | 28 rounds snake + 3rd round reversal
MY PICK SEQUENCE: 1.02, 2.11, 3.11, 4.02, 5.11, 6.02, 7.11, 8.02...

PICKS MADE SO FAR:
{picks_context if picks_context else 'None yet'}

MY CURRENT ROSTER: {roster_context}
CURRENT PICK ON CLOCK: {current_pick}
AVAILABLE PLAYERS: {available_input[:2000] if available_input else 'Not specified'}

Provide:
1. TOP RECOMMENDATION at this pick with clear reasoning
2. Next 2 alternatives if top pick is gone
3. Roster construction note (what positions still needed)
4. Value remaining alert (any elite players still available that are sliding)

Be decisive. One clear recommendation first. Format for mobile."""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=1000,
            system=SYSTEM_PROMPT,
            tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 2}],
            messages=[{"role": "user", "content": prompt}]
        )
        recommendation = "".join(b.text for b in response.content if hasattr(b, 'text'))
        return jsonify({"recommendation": recommendation, "success": True})
    except Exception as e:
        return jsonify({"error": str(e), "success": False})

@app.route('/api/draft/state', methods=['GET', 'POST'])
def draft_state():
    """Save and retrieve draft state"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS draft_state
                (id INTEGER PRIMARY KEY, state_data TEXT, updated_at TEXT)''')

    if request.method == 'POST':
        state = request.json.get('state', {})
        c.execute('''INSERT OR REPLACE INTO draft_state (id, state_data, updated_at)
                    VALUES (1, ?, ?)''', (json.dumps(state), datetime.now().isoformat()))
        conn.commit()
        conn.close()
        return jsonify({"success": True})
    else:
        c.execute("SELECT state_data FROM draft_state WHERE id=1")
        row = c.fetchone()
        conn.close()
        if row:
            return jsonify({"state": json.loads(row[0]), "success": True})
        return jsonify({"state": {}, "success": True})

# ============================================================
# VS RANKINGS — 6PT TD SPECIFIC HEAD-TO-HEAD
# ============================================================

def seed_vs_players():
    """Seed VS rankings pool with 6pt TD adjusted ELO — QBs boosted ~18%, copies ktc_tier"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM vs_rankings")
    if c.fetchone()[0] > 0:
        conn.close()
        return
    c.execute("SELECT player_name, position, team, elo_score, COALESCE(ktc_tier, 10) FROM player_rankings")
    rows = c.fetchall()
    for name, pos, team, elo, ktc_tier in rows:
        adjusted_elo = round(elo * 1.18) if pos == 'QB' else elo
        # QBs move up ~2 KTC tier buckets in 6pt TD
        adjusted_tier = max(1, ktc_tier - 2) if pos == 'QB' else ktc_tier
        c.execute('''INSERT OR IGNORE INTO vs_rankings
                    (player_name, position, team, elo_score, comparisons, last_updated, ktc_tier)
                    VALUES (?, ?, ?, ?, 0, ?, ?)''',
                 (name, pos, team, adjusted_elo, datetime.now().isoformat(), adjusted_tier))
    conn.commit()
    conn.close()

seed_vs_players()

@app.route('/api/vs/pair', methods=['GET'])
def get_vs_pair():
    """Proximity-based matchup for VS 6pt TD rankings — hybrid KTC tier matching"""
    position_filter = request.args.get('position', 'ALL')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    if position_filter != 'ALL':
        c.execute('''SELECT player_name, position, team, elo_score, comparisons, COALESCE(ktc_tier, 10)
                    FROM vs_rankings WHERE position=?
                    ORDER BY comparisons ASC, RANDOM() LIMIT 80''', (position_filter,))
    else:
        c.execute('''SELECT player_name, position, team, elo_score, comparisons, COALESCE(ktc_tier, 10)
                    FROM vs_rankings ORDER BY comparisons ASC, RANDOM() LIMIT 120''')
    rows = c.fetchall()
    conn.close()
    if len(rows) < 2:
        return jsonify({"players": [], "success": False})

    players = [{"name": r[0], "position": r[1], "team": r[2],
                "elo": r[3], "comparisons": r[4], "ktc_tier": r[5]} for r in rows]

    def get_comparison_tier(p):
        if p['comparisons'] < 10:
            kt = p['ktc_tier']
            if kt <= 1: return 1
            if kt <= 3: return 2
            if kt <= 6: return 3
            if kt <= 10: return 4
            if kt <= 14: return 5
            if kt <= 18: return 6
            return 7
        else:
            elo = p['elo']
            if elo >= 2800: return 1
            if elo >= 2400: return 2
            if elo >= 2000: return 3
            if elo >= 1600: return 4
            if elo >= 1200: return 5
            if elo >= 800: return 6
            return 7

    is_filtered = position_filter != 'ALL'
    pool = sorted(players, key=lambda x: x['comparisons'])
    p1 = pool[random.randint(0, min(9, len(pool)-1))]
    p1_tier = get_comparison_tier(p1)

    rand = random.random()
    if rand < 0.60:
        target_tiers = [p1_tier]
    elif rand < 0.90:
        target_tiers = [p1_tier - 1, p1_tier + 1]
    else:
        target_tiers = [p1_tier + i for i in range(-3, 4) if i != 0]
    target_tiers = [t for t in target_tiers if 1 <= t <= 7]

    p2_candidates = [p for p in players
                    if p['name'] != p1['name'] and get_comparison_tier(p) in target_tiers]

    if not p2_candidates and is_filtered:
        for spread in range(1, 4):
            expanded = [p1_tier + i for i in range(-spread, spread+1) if i != 0]
            expanded = [t for t in expanded if 1 <= t <= 7]
            p2_candidates = [p for p in players
                            if p['name'] != p1['name'] and get_comparison_tier(p) in expanded]
            if p2_candidates:
                break

    if not p2_candidates:
        p2_candidates = [p for p in players if p['name'] != p1['name']]

    p2_candidates.sort(key=lambda x: x['comparisons'])
    p2 = p2_candidates[random.randint(0, min(9, len(p2_candidates)-1))]
    return jsonify({
        "players": [p1, p2], "success": True,
        "tiers": [p1_tier, get_comparison_tier(p2)]
    })

@app.route('/api/vs/vote', methods=['POST'])
def vs_vote():
    """Record VS comparison vote and update ELO"""
    data = request.json
    winner = data.get('winner', '')
    loser = data.get('loser', '')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT elo_score, comparisons FROM vs_rankings WHERE player_name=?", (winner,))
    wr = c.fetchone()
    c.execute("SELECT elo_score, comparisons FROM vs_rankings WHERE player_name=?", (loser,))
    lr = c.fetchone()
    if wr and lr:
        K = 32
        we, le = wr[0], lr[0]
        exp_w = 1 / (1 + 10 ** ((le - we) / 400))
        new_we = we + K * (1 - exp_w)
        new_le = le + K * (0 - (1 - exp_w))
        c.execute("UPDATE vs_rankings SET elo_score=?, comparisons=comparisons+1, last_updated=? WHERE player_name=?",
                 (new_we, datetime.now().isoformat(), winner))
        c.execute("UPDATE vs_rankings SET elo_score=?, comparisons=comparisons+1, last_updated=? WHERE player_name=?",
                 (new_le, datetime.now().isoformat(), loser))
    conn.commit()
    conn.close()
    return jsonify({"success": True})

@app.route('/api/vs/list', methods=['GET'])
def get_vs_list():
    """Get VS personal rankings sorted by ELO, limited to top 150"""
    position_filter = request.args.get('position', 'ALL')
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    if position_filter != 'ALL':
        c.execute('''SELECT player_name, position, team, elo_score, comparisons
                    FROM vs_rankings WHERE position=?
                    ORDER BY elo_score DESC LIMIT 150''', (position_filter,))
    else:
        c.execute('''SELECT player_name, position, team, elo_score, comparisons
                    FROM vs_rankings ORDER BY elo_score DESC LIMIT 150''')
    rows = c.fetchall()
    conn.close()
    rankings = [{"rank": i+1, "name": r[0], "position": r[1], "team": r[2],
                 "elo": round(r[3]), "comparisons": r[4]} for i, r in enumerate(rows)]
    return jsonify({"rankings": rankings, "success": True})

@app.route('/api/vs/adjust', methods=['POST'])
def vs_adjust():
    """Manually adjust VS player ranking"""
    name = request.json.get('name', '')
    direction = request.json.get('direction', 'up')
    positions = request.json.get('positions', 5)
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT elo_score FROM vs_rankings WHERE player_name=?", (name,))
    row = c.fetchone()
    if not row:
        conn.close()
        return jsonify({"success": False, "error": "Player not found"})
    current_elo = row[0]
    c.execute("SELECT elo_score FROM vs_rankings ORDER BY elo_score DESC")
    all_elos = [r[0] for r in c.fetchall()]
    try:
        curr_pos = next(i for i, e in enumerate(all_elos) if abs(e - current_elo) < 0.01)
    except StopIteration:
        curr_pos = len(all_elos) // 2
    target_pos = max(0, curr_pos - positions) if direction == 'up' else min(len(all_elos)-1, curr_pos + positions)
    if target_pos != curr_pos:
        if direction == 'up' and target_pos > 0:
            new_elo = (all_elos[target_pos-1] + all_elos[target_pos]) / 2
        elif direction == 'down' and target_pos < len(all_elos)-1:
            new_elo = (all_elos[target_pos] + all_elos[target_pos+1]) / 2
        else:
            new_elo = all_elos[target_pos]
        c.execute("UPDATE vs_rankings SET elo_score=?, last_updated=? WHERE player_name=?",
                 (new_elo, datetime.now().isoformat(), name))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "new_elo": round(new_elo if target_pos != curr_pos else current_elo)})

@app.route('/api/vs/reset', methods=['POST'])
def vs_reset():
    """Reset VS rankings, re-seed from standard rankings with QB boost, then apply KTC tiers"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("DELETE FROM vs_rankings")
    conn.commit()
    conn.close()
    seed_vs_players()
    refresh_ktc_tiers()  # Critical: populate ktc_tier for boundary training
    return jsonify({"success": True, "message": "VS rankings reset with 6pt TD QB adjustments and KTC tiers applied"})

@app.route('/api/vs/refresh_tiers', methods=['POST'])
def vs_refresh_tiers():
    """Re-apply KTC tiers to VS rankings without wiping comparison history"""
    _do_refresh_ktc_tiers()
    return jsonify({"success": True, "message": "KTC tiers refreshed on VS rankings"})

# ============================================================
# MARKET DATA — PASTE KTC / DDL / SLEEPER ADP
# ============================================================

@app.route('/api/market/paste', methods=['POST'])
def paste_market_data():
    """
    Accept pasted rankings from KTC, Dynasty Data Lab, or Sleeper ADP.
    Parser handles messy copied text — just paste what you see on screen.
    """
    source = request.json.get('source', '')  # 'ktc', 'ddl', 'sleeper'
    raw_text = request.json.get('text', '')
    if not source or not raw_text:
        return jsonify({"success": False, "error": "source and text required"})

    players_parsed = []

    if source == 'ktc':
        # KTC format: "1 Josh Allen BUF QB1 9999"
        # Also handles the format from the paste: rank, name, team, pos, value
        lines = raw_text.strip().split('\n')
        rank = 0
        for line in lines:
            line = line.strip()
            if not line:
                continue
            # Extract KTC value (4-5 digit number at end)
            val_match = re.search(r'\b(\d{3,5})\s*$', line)
            if not val_match:
                continue
            val = int(val_match.group(1))
            if val < 100:
                continue
            # Extract position
            pos_match = re.search(r'\b(QB|RB|WR|TE|PICK)\b', line)
            if not pos_match:
                continue
            pos = pos_match.group(1)
            # Extract team (2-3 uppercase letters)
            team_match = re.search(r'\b([A-Z]{2,3})\b', line)
            team = team_match.group(1) if team_match else 'FA'
            # Player name: everything before team/pos markers
            name_part = re.sub(r'\b(QB|RB|WR|TE|PICK|QB\d+|RB\d+|WR\d+|TE\d+)\b.*', '', line)
            name_part = re.sub(r'^\d+\s*', '', name_part)
            name_part = re.sub(r'\b[A-Z]{2,3}\b', '', name_part)
            name = re.sub(r'\s+', ' ', name_part).strip()
            name = re.sub(r"[^\w\s'.]", '', name).strip()
            if len(name) < 3:
                continue
            rank += 1
            players_parsed.append({'name': name, 'rank': rank, 'value': val, 'position': pos, 'team': team})

    elif source == 'ddl':
        # DDL format: rank | name | team | pos | value
        # Handles both table and list formats
        lines = raw_text.strip().split('\n')
        rank = 0
        for line in lines:
            line = line.strip()
            if not line or len(line) < 5:
                continue
            # Skip headers
            if any(h in line.lower() for h in ['player', 'rank', 'name', 'value', '---']):
                continue
            pos_match = re.search(r'\b(QB|RB|WR|TE)\b', line)
            if not pos_match:
                continue
            pos = pos_match.group(1)
            val_match = re.search(r'\b(\d{3,5})\b', line)
            val = int(val_match.group(1)) if val_match else 5000
            # Name is usually before position marker
            name_part = line[:pos_match.start()]
            name_part = re.sub(r'^\d+[\.\):\s]+', '', name_part)
            name_part = re.sub(r'\b[A-Z]{2,3}\b\s*$', '', name_part)
            name = re.sub(r'\s+', ' ', name_part).strip()
            name = re.sub(r"[^\w\s'.]", '', name).strip()
            if len(name) < 3:
                continue
            rank += 1
            players_parsed.append({'name': name, 'rank': rank, 'value': val, 'position': pos, 'team': 'FA'})

    elif source == 'sleeper':
        # Sleeper ADP: usually "1. Player Name POS" or "Name (POS) - Team"
        lines = raw_text.strip().split('\n')
        rank = 0
        for line in lines:
            line = line.strip()
            if not line:
                continue
            pos_match = re.search(r'\b(QB|RB|WR|TE)\b', line)
            if not pos_match:
                continue
            pos = pos_match.group(1)
            team_match = re.search(r'\b([A-Z]{2,3})\b(?!.*\b(?:QB|RB|WR|TE)\b)', line)
            team = team_match.group(1) if team_match else 'FA'
            name_part = line[:pos_match.start()]
            name_part = re.sub(r'^\d+[\.\):\s]+', '', name_part)
            name = re.sub(r'\s+', ' ', name_part).strip()
            name = re.sub(r"[^\w\s'.]", '', name).strip()
            if len(name) < 3:
                continue
            rank += 1
            players_parsed.append({'name': name, 'rank': rank, 'value': 0, 'position': pos, 'team': team})

    if not players_parsed:
        return jsonify({"success": False, "error": "Could not parse any players. Try copying just the player list rows."})

    # Store in DB
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    for p in players_parsed:
        c.execute('''INSERT OR REPLACE INTO market_data
                    (source, player_name, rank, value, position, team, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)''',
                 (source, p['name'], p['rank'], p['value'], p['position'], p['team'],
                  datetime.now().isoformat()))
    conn.commit()
    conn.close()
    return jsonify({"success": True, "parsed": len(players_parsed),
                   "sample": [p['name'] for p in players_parsed[:5]]})

@app.route('/api/market/data', methods=['GET'])
def get_market_data():
    """Get all stored market data by source"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''SELECT source, player_name, rank, value, position, team, updated_at
                FROM market_data ORDER BY source, rank ASC''')
    rows = c.fetchall()
    conn.close()
    by_source = {}
    for source, name, rank, val, pos, team, updated in rows:
        if source not in by_source:
            by_source[source] = {'players': [], 'updated': updated}
        by_source[source]['players'].append({'name': name, 'rank': rank, 'value': val, 'position': pos, 'team': team})
    sources_available = list(by_source.keys())
    return jsonify({"by_source": by_source, "sources": sources_available, "success": True})

# ============================================================
# TIER BREAKS + TRADE UP/DOWN TARGETS
# ============================================================

@app.route('/api/tiers', methods=['GET'])
def get_tiers():
    """
    Positional tier breaks.
    - avg < 5 comps per player in position: KTC tier baseline
    - avg >= 5 comps per player: personal ELO tiers
    DDL ADP used for draft slot context with fuzzy name matching.
    """

    def normalize(name):
        import re as _re
        name = _re.sub(r"[^a-z0-9\s]", "", name.lower()).strip()
        name = _re.sub(r'\b(jr|sr|ii|iii|iv)\b', '', name).strip()
        return _re.sub(r'\s+', '', name)

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    c.execute("""SELECT player_name, position, elo_score, comparisons, COALESCE(ktc_tier, 10)
                FROM vs_rankings ORDER BY elo_score DESC LIMIT 300""")
    vs_rows = c.fetchall()
    vs_rank = {r[0]: {
        'rank': i+1, 'pos': r[1], 'elo': r[2],
        'comparisons': r[3], 'ktc_tier': r[4]
    } for i, r in enumerate(vs_rows)}

    c.execute("SELECT player_name, rank, value FROM market_data WHERE source='ktc' ORDER BY rank")
    ktc_rank_map = {normalize(r[0]): {'rank': r[1], 'value': r[2], 'name': r[0]} for r in c.fetchall()}

    c.execute("SELECT player_name, rank, team FROM market_data WHERE source='ddl' ORDER BY rank")
    ddl_raw = c.fetchall()
    conn.close()

    # Build DDL with fuzzy key
    ddl_rank_map = {}
    for name, rank, adp_str in ddl_raw:
        try:
            adp = float(adp_str)
        except:
            adp = float(rank)
        ddl_rank_map[normalize(name)] = {'rank': rank, 'adp': adp, 'name': name}

    has_ktc = len(ktc_rank_map) > 10
    has_ddl = len(ddl_rank_map) > 10

    def adp_to_pick(adp):
        if not adp:
            return '—'
        r = int(adp // 12) + 1
        s = int(adp % 12) + 1
        return f"{r}.{str(s).zfill(2)}"

    def get_ddl(player_name):
        key = normalize(player_name)
        return ddl_rank_map.get(key, {})

    def enrich(name, rank, elo, ktc_tier):
        ddl = get_ddl(name)
        adp = ddl.get('adp')
        return {
            'name': name,
            'vs_rank': rank,
            'elo': round(elo),
            'ktc_tier': ktc_tier,
            'adp': adp,
            'adp_pick': adp_to_pick(adp),
        }

    def build_ktc_tiers(pos_players):
        """
        One tier per KTC tier number. Each KTC tier = one logical tier break.
        This gives 5-12 tiers per position with 1-9 players each.
        pos_players: list of (name, rank, elo, ktc_tier)
        """
        from collections import OrderedDict

        KTC_TIER_LABELS = {
            1: 'Generational', 2: 'Elite', 3: 'Premium',
            4: 'Strong', 5: 'Strong', 6: 'Solid', 7: 'Solid',
            8: 'Value', 9: 'Value', 10: 'Deep Value', 11: 'Deep Value',
            12: 'Late Round', 13: 'Late Round', 14: 'Speculative',
            15: 'Speculative', 16: 'Deep Spec', 17: 'Very Deep',
        }

        # Group by exact KTC tier number
        tier_groups = OrderedDict()
        for p in pos_players:
            kt = p[3]
            if kt not in tier_groups:
                tier_groups[kt] = []
            tier_groups[kt].append(p)

        tiers = []
        for kt, group in sorted(tier_groups.items()):
            group_sorted = sorted(group, key=lambda x: -x[2])  # ELO desc
            last = group_sorted[-1]
            last_ddl = get_ddl(last[0])
            last_adp = last_ddl.get('adp')
            label = f"KTC T{kt} — {KTC_TIER_LABELS.get(kt, 'Deep')}"
            tiers.append({
                'tier_num': len(tiers) + 1,
                'label': label,
                'players': [enrich(p[0], p[1], p[2], p[3]) for p in group_sorted],
                'count': len(group_sorted),
                'break_after': last[0],
                'last_player': last[0],
                'last_adp_pick': adp_to_pick(last_adp),
                'gap': 0,
                'is_major': kt <= 3,
                'is_moderate': 4 <= kt <= 6,
                'data_source': 'ktc',
            })
        return tiers


    def build_elo_tiers(pos_players):
        """Build tiers from ELO gaps. pos_players: (name, rank, elo, ktc_tier)"""
        pos_elo_max = pos_players[0][2]
        pos_elo_min = pos_players[-1][2]
        pos_elo_range = max(pos_elo_max - pos_elo_min, 1)
        # Use 6% of range as threshold — more sensitive than before
        gap_threshold = max(60, pos_elo_range * 0.06)

        raw_tiers = []
        current = []
        for i, (name, rank, elo, ktc_tier) in enumerate(pos_players):
            current.append(enrich(name, rank, elo, ktc_tier))
            if i < len(pos_players) - 1:
                gap = elo - pos_players[i+1][2]
                if gap >= gap_threshold:
                    raw_tiers.append({'players': current.copy(), 'gap': round(gap), 'break_after': name})
                    current = []
        if current:
            raw_tiers.append({'players': current.copy(), 'gap': 0, 'break_after': None})

        # Merge isolated single-player tiers (not at very top)
        merged = []
        i = 0
        while i < len(raw_tiers):
            t = raw_tiers[i]
            if len(t['players']) == 1 and i > 0 and i + 1 < len(raw_tiers) and t['gap'] < gap_threshold * 1.5:
                # Merge into previous tier
                merged[-1]['players'].extend(t['players'])
                merged[-1]['gap'] = t['gap']
                merged[-1]['break_after'] = t['break_after']
            else:
                merged.append(t)
            i += 1

        tiers = []
        for ti, t in enumerate(merged):
            last = t['players'][-1]
            is_major = t['gap'] > gap_threshold * 2.5
            is_moderate = gap_threshold * 1.5 <= t['gap'] <= gap_threshold * 2.5
            tiers.append({
                'tier_num': ti + 1,
                'label': None,
                'players': t['players'],
                'count': len(t['players']),
                'break_after': t['break_after'],
                'last_player': last['name'],
                'last_adp_pick': last['adp_pick'],
                'gap': t['gap'],
                'is_major': is_major,
                'is_moderate': is_moderate,
                'data_source': 'personal',
            })
        return tiers

    positions = ['QB', 'RB', 'WR', 'TE']
    tier_data = {}
    comp_counts = {}

    for pos in positions:
        pos_players = [
            (name, d['rank'], d['elo'], d['ktc_tier'])
            for name, d in vs_rank.items() if d['pos'] == pos
        ]
        if len(pos_players) < 2:
            continue

        total_comps = sum(vs_rank[p[0]]['comparisons'] for p in pos_players)
        n_players = len(pos_players)
        avg_comps = total_comps / n_players if n_players > 0 else 0
        comp_counts[pos] = round(avg_comps, 1)

        pos_players.sort(key=lambda x: x[1])  # sort by vs_rank

        # Switch to personal at avg 5 comps per player
        if avg_comps >= 5:
            tier_data[pos] = build_elo_tiers(pos_players)
        else:
            tier_data[pos] = build_ktc_tiers(pos_players)

    # Trade targets
    trade_targets = []
    for name, vd in vs_rank.items():
        if vd['rank'] > 150:
            continue
        ddl = get_ddl(name)
        ktc_norm = normalize(name)
        my_rank = vd['rank']

        if ddl:
            market_rank = ddl['rank']
            adp_float = ddl.get('adp')
            market_source = 'DDL'
        elif ktc_norm in ktc_rank_map:
            market_rank = ktc_rank_map[ktc_norm]['rank']
            adp_float = None
            market_source = 'KTC'
        else:
            continue

        delta = market_rank - my_rank
        if abs(delta) < 10:
            continue

        if delta >= 30: signal = 'STRONG BUY'
        elif delta >= 15: signal = 'BUY'
        elif delta >= 10: signal = 'SLIGHT BUY'
        elif delta <= -30: signal = 'STRONG SELL'
        elif delta <= -15: signal = 'SELL'
        else: signal = 'SLIGHT SELL'

        est_pick = adp_to_pick(adp_float) if adp_float else adp_to_pick(market_rank)
        trade_targets.append({
            'name': name, 'position': vd['pos'],
            'my_rank': my_rank, 'market_rank': market_rank,
            'market_source': market_source, 'adp': adp_float,
            'delta': delta, 'signal': signal, 'est_draft_pick': est_pick,
        })

    trade_targets.sort(key=lambda x: abs(x['delta']), reverse=True)

    return jsonify({
        "tiers": tier_data,
        "trade_targets": trade_targets[:40],
        "comp_counts": comp_counts,
        "sources_available": (['ktc'] if has_ktc else []) + (['ddl'] if has_ddl else []),
        "ktc_count": len(ktc_rank_map),
        "ddl_count": len(ddl_rank_map),
        "success": True
    })

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # VS rankings with comparison counts and ktc_tier
    c.execute("""SELECT player_name, position, elo_score, comparisons, COALESCE(ktc_tier, 10)
                FROM vs_rankings ORDER BY elo_score DESC LIMIT 300""")
    vs_rows = c.fetchall()
    vs_rank = {r[0]: {
        'rank': i+1, 'pos': r[1], 'elo': r[2],
        'comparisons': r[3], 'ktc_tier': r[4]
    } for i, r in enumerate(vs_rows)}

    # KTC data
    c.execute("SELECT player_name, rank, value FROM market_data WHERE source='ktc' ORDER BY rank ASC")
    ktc_rank_map = {r[0]: {'rank': r[1], 'value': r[2]} for r in c.fetchall()}

    # DDL ADP
    c.execute("SELECT player_name, rank, team FROM market_data WHERE source='ddl' ORDER BY rank ASC")
    ddl_rank_map = {}
    for name, rank, adp_str in c.fetchall():
        try:
            ddl_rank_map[name] = {'rank': rank, 'adp': float(adp_str)}
        except:
            ddl_rank_map[name] = {'rank': rank, 'adp': float(rank)}

    conn.close()

    has_ktc = len(ktc_rank_map) > 10
    has_ddl = len(ddl_rank_map) > 10

    def adp_to_pick(adp):
        if not adp:
            return '—'
        r = int(adp // 12) + 1
        s = int(adp % 12) + 1
        return f"{r}.{str(s).zfill(2)}"

    def enrich_player(name, rank, elo, ktc_tier):
        ddl = ddl_rank_map.get(name, {})
        adp = ddl.get('adp')
        return {
            'name': name,
            'vs_rank': rank,
            'elo': round(elo),
            'ktc_tier': ktc_tier,
            'adp': adp,
            'adp_pick': adp_to_pick(adp),
        }

    def build_elo_tiers(pos_players_raw, pos):
        """Build tiers from personal ELO gaps. pos_players_raw: (name, rank, elo, ktc_tier)"""
        pos_players = [(p[0], p[1], p[2]) for p in pos_players_raw]

        pos_elo_max = pos_players[0][2]
        pos_elo_min = pos_players[-1][2]
        pos_elo_range = max(pos_elo_max - pos_elo_min, 1)
        gap_threshold = max(80, pos_elo_range * 0.08)

        raw_tiers = []
        current_tier = []
        for i, (name, rank, elo) in enumerate(pos_players):
            p_data = vs_rank.get(name, {})
            current_tier.append(enrich_player(name, rank, elo, p_data.get('ktc_tier', 10)))
            if i < len(pos_players) - 1:
                gap = elo - pos_players[i+1][2]
                if gap >= gap_threshold:
                    raw_tiers.append({'players': current_tier.copy(), 'gap': round(gap),
                                     'break_after': name, 'gap_threshold': gap_threshold})
                    current_tier = []
        if current_tier:
            raw_tiers.append({'players': current_tier.copy(), 'gap': 0,
                             'break_after': None, 'gap_threshold': gap_threshold})

        # Merge single-player tiers
        merged = []
        i = 0
        while i < len(raw_tiers):
            t = raw_tiers[i]
            if len(t['players']) == 1 and i + 1 < len(raw_tiers) and t['gap'] < gap_threshold * 2:
                combined = t['players'] + raw_tiers[i+1]['players']
                merged.append({'players': combined, 'gap': raw_tiers[i+1]['gap'],
                               'break_after': raw_tiers[i+1]['break_after'],
                               'gap_threshold': gap_threshold})
                i += 2
            else:
                merged.append(t)
                i += 1

        tiers = []
        for ti, t in enumerate(merged):
            last = t['players'][-1]
            is_major = t['gap'] > gap_threshold * 2
            is_moderate = gap_threshold <= t['gap'] <= gap_threshold * 2
            tiers.append({
                'tier_num': ti + 1,
                'label': None,
                'players': t['players'],
                'count': len(t['players']),
                'break_after': t['break_after'],
                'last_player': last['name'],
                'last_adp_pick': last['adp_pick'],
                'gap': t['gap'],
                'is_major': is_major,
                'is_moderate': is_moderate,
                'data_source': 'personal',
            })
        return tiers

    # Build tiers per position
    positions = ['QB', 'RB', 'WR', 'TE']
    tier_data = {}
    comp_counts = {}

    for pos in positions:
        # (name, vs_rank, elo, ktc_tier)
        pos_players = [
            (name, d['rank'], d['elo'], d['ktc_tier'])
            for name, d in vs_rank.items() if d['pos'] == pos
        ]
        if len(pos_players) < 2:
            continue

        # Count total comparisons for this position
        total_comps = sum(vs_rank[p[0]]['comparisons'] for p in pos_players)
        comp_counts[pos] = total_comps

        # Threshold: 10 comparisons per position player on average (so ~10*n total)
        # Simpler: just use total > 20 as minimum meaningful signal
        has_enough_data = total_comps >= 20

        pos_players.sort(key=lambda x: x[1])  # sort by vs_rank

        if has_enough_data:
            tier_data[pos] = build_elo_tiers(pos_players, pos)
        else:
            tier_data[pos] = build_ktc_tiers(pos_players)

    # Trade targets (unchanged)
    trade_targets = []
    for name, vd in vs_rank.items():
        if vd['rank'] > 150:
            continue
        my_rank = vd['rank']
        market_rank = None
        market_source = None
        adp_float = None

        if name in ddl_rank_map:
            market_rank = ddl_rank_map[name]['rank']
            adp_float = ddl_rank_map[name]['adp']
            market_source = 'DDL'
        elif name in ktc_rank_map:
            market_rank = ktc_rank_map[name]['rank']
            market_source = 'KTC'

        if not market_rank:
            continue

        delta = market_rank - my_rank
        if abs(delta) < 10:
            continue

        if delta >= 30: signal = 'STRONG BUY'
        elif delta >= 15: signal = 'BUY'
        elif delta >= 10: signal = 'SLIGHT BUY'
        elif delta <= -30: signal = 'STRONG SELL'
        elif delta <= -15: signal = 'SELL'
        else: signal = 'SLIGHT SELL'

        est_draft_pick = adp_to_pick(adp_float) if adp_float else adp_to_pick(market_rank)

        trade_targets.append({
            'name': name, 'position': vd['pos'],
            'my_rank': my_rank, 'market_rank': market_rank,
            'market_source': market_source, 'adp': adp_float,
            'delta': delta, 'signal': signal, 'est_draft_pick': est_draft_pick,
            'note': f"Market ~{est_draft_pick}" if delta > 0 else f"Slides to ~{est_draft_pick}",
        })

    trade_targets.sort(key=lambda x: abs(x['delta']), reverse=True)

    return jsonify({
        "tiers": tier_data,
        "trade_targets": trade_targets[:40],
        "comp_counts": comp_counts,
        "sources_available": (['ktc'] if has_ktc else []) + (['ddl'] if has_ddl else []),
        "ktc_count": len(ktc_rank_map),
        "ddl_count": len(ddl_rank_map),
        "success": True
    })

def seed_ddl_market_data():
    """Store DDL ADP data — seeded from May 12 2026 paste"""
    DDL_PLAYERS = [
    ('Josh Allen', 'QB', 1, 1.95),
    ('Bijan Robinson', 'RB', 2, 2.7),
    ('Drake Maye', 'QB', 3, 3.76),
    ('Ja\'Marr Chase', 'WR', 4, 4.68),
    ('Jahmyr Gibbs', 'RB', 5, 4.97),
    ('Jaxon Smith-Njigba', 'WR', 6, 6.45),
    ('Puka Nacua', 'WR', 7, 7.43),
    ('Jayden Daniels', 'QB', 8, 10.51),
    ('Amon-Ra St. Brown', 'WR', 9, 11.39),
    ('Caleb Williams', 'QB', 10, 12.58),
    ('Brock Bowers', 'TE', 11, 12.72),
    ('Lamar Jackson', 'QB', 12, 12.8),
    ('Malik Nabers', 'WR', 13, 13.72),
    ('Joe Burrow', 'QB', 14, 15.12),
    ('Ashton Jeanty', 'RB', 15, 15.34),
    ('Trey McBride', 'TE', 16, 15.67),
    ('Justin Jefferson', 'WR', 17, 16.22),
    ('Jeremiyah Love', 'RB', 18, 16.84),
    ('CeeDee Lamb', 'WR', 19, 19.34),
    ('Devon Achane', 'RB', 20, 21.27),
    ('Omarion Hampton', 'RB', 21, 21.88),
    ('Jaxson Dart', 'QB', 22, 22.08),
    ('Justin Herbert', 'QB', 23, 24.06),
    ('Drake London', 'WR', 24, 24.22),
    ('Jalen Hurts', 'QB', 25, 26.37),
    ('Jonathan Taylor', 'RB', 26, 27.47),
    ('Patrick Mahomes', 'QB', 27, 28.33),
    ('Tetairoa McMillan', 'WR', 28, 28.49),
    ('James Cook', 'RB', 29, 28.97),
    ('Colston Loveland', 'TE', 30, 29.22),
    ('Trevor Lawrence', 'QB', 31, 31.04),
    ('George Pickens', 'WR', 32, 33.91),
    ('Bo Nix', 'QB', 33, 34.43),
    ('Emeka Egbuka', 'WR', 34, 35.5),
    ('Tyler Warren', 'TE', 35, 36.88),
    ('Nico Collins', 'WR', 36, 37.69),
    ('Brock Purdy', 'QB', 37, 39.08),
    ('Christian McCaffrey', 'RB', 38, 40.83),
    ('Kenneth Walker III', 'RB', 39, 41.15),
    ('Chris Olave', 'WR', 40, 41.44),
    ('TreVeyon Henderson', 'RB', 41, 42.67),
    ('Carnell Tate', 'WR', 42, 43.29),
    ('Chase Brown', 'RB', 43, 43.7),
    ('Garrett Wilson', 'WR', 44, 43.97),
    ('Rashee Rice', 'WR', 45, 44.12),
    ('Dak Prescott', 'QB', 46, 47.14),
    ('Ladd McConkey', 'WR', 47, 49.0),
    ('Breece Hall', 'RB', 48, 49.48),
    ('Quinshon Judkins', 'RB', 49, 51.06),
    ('Fernando Mendoza', 'QB', 50, 51.76),
    ('Harold Fannin', 'TE', 51, 51.83),
    ('Jordan Love', 'QB', 52, 53.27),
    ('Luther Burden', 'WR', 53, 53.29),
    ('Bucky Irving', 'RB', 54, 55.06),
    ('Saquon Barkley', 'RB', 55, 55.08),
    ('Jordyn Tyson', 'WR', 56, 56.08),
    ('Rome Odunze', 'WR', 57, 57.49),
    ('A.J. Brown', 'WR', 58, 58.17),
    ('Tucker Kraft', 'TE', 59, 58.95),
    ('Kyren Williams', 'RB', 60, 59.64),
    ('Marvin Harrison Jr.', 'WR', 61, 62.51),
    ('Makai Lemon', 'WR', 62, 63.6),
    ('Zay Flowers', 'WR', 63, 64.15),
    ('Cam Ward', 'QB', 64, 65.25),
    ('Tee Higgins', 'WR', 65, 66.42),
    ('Cam Skattebo', 'RB', 66, 67.09),
    ('Brian Thomas Jr.', 'WR', 67, 67.69),
    ('Javonte Williams', 'RB', 68, 68.3),
    ('Jared Goff', 'QB', 69, 69.32),
    ('Jadarian Price', 'RB', 70, 69.33),
    ('DeVonta Smith', 'WR', 71, 69.52),
    ('Josh Jacobs', 'RB', 72, 69.58),
    ('Tyler Shough', 'QB', 73, 70.32),
    ('Baker Mayfield', 'QB', 74, 70.51),
    ('Sam LaPorta', 'TE', 75, 70.92),
    ('Travis Etienne', 'RB', 76, 71.31),
    ('Jameson Williams', 'WR', 77, 74.97),
    ('Kyle Pitts', 'TE', 78, 75.46),
    ('Jaylen Waddle', 'WR', 79, 75.48),
    ('C.J. Stroud', 'QB', 80, 77.6),
    ('Derrick Henry', 'RB', 81, 78.23),
    ('Sam Darnold', 'QB', 82, 83.25),
    ('Kenyon Sadiq', 'TE', 83, 83.34),
    ('KC Concepcion', 'WR', 84, 84.5),
    ('Bhayshul Tuten', 'RB', 85, 84.67),
    ('Oronde Gadsden', 'TE', 86, 85.56),
    ('Kyler Murray', 'QB', 87, 87.12),
    ('D.J. Moore', 'WR', 88, 88.1),
    ('Alec Pierce', 'WR', 89, 88.86),
    ('Matthew Stafford', 'QB', 90, 88.97),
    ('RJ Harvey', 'RB', 91, 90.54),
    ('Bryce Young', 'QB', 92, 91.51),
    ('David Montgomery', 'RB', 93, 95.69),
    ('Malik Willis', 'QB', 94, 96.94),
    ('D\'Andre Swift', 'RB', 95, 97.55),
    ('Christian Watson', 'WR', 96, 97.89),
    ('Jordan Addison', 'WR', 97, 98.55),
    ('Michael Wilson', 'WR', 98, 99.89),
    ('Kyle Monangai', 'RB', 99, 100.17),
    ('Omar Cooper', 'WR', 100, 101.0),
    ('Terry McLaurin', 'WR', 101, 102.16),
    ('Jake Ferguson', 'TE', 102, 103.61),
    ('Daniel Jones', 'QB', 103, 104.28),
    ('Chuba Hubbard', 'RB', 104, 104.39),
    ('Wan\'Dale Robinson', 'WR', 105, 106.93),
    ('Parker Washington', 'WR', 106, 107.69),
    ('Davante Adams', 'WR', 107, 108.69),
    ('D.K. Metcalf', 'WR', 108, 109.17),
    ('Dalton Kincaid', 'TE', 109, 109.29),
    ('Ricky Pearsall', 'WR', 110, 109.48),
    ('Eli Stowers', 'TE', 111, 110.9),
    ('Ty Simpson', 'QB', 112, 110.93),
    ('Isaiah Likely', 'TE', 113, 112.76),
    ('Mike Evans', 'WR', 114, 113.32),
    ('Matthew Golden', 'WR', 115, 115.41),
    ('Jaylen Warren', 'RB', 116, 116.33),
    ('George Kittle', 'TE', 117, 116.89),
    ('Jayden Higgins', 'WR', 118, 117.97),
    ('Denzel Boston', 'WR', 119, 118.53),
    ('Brenton Strange', 'TE', 120, 120.51),
    ('Jonah Coleman', 'RB', 121, 121.38),
    ('Blake Corum', 'RB', 122, 121.73),
    ('Rico Dowdle', 'RB', 123, 123.61),
    ('Quentin Johnston', 'WR', 124, 125.26),
    ('Michael Pittman Jr.', 'WR', 125, 127.04),
    ('Xavier Worthy', 'WR', 126, 128.03),
    ('Josh Downs', 'WR', 127, 128.39),
    ('Zach Charbonnet', 'RB', 128, 128.85),
    ('Travis Hunter', 'WR', 129, 129.24),
    ('Jayden Reed', 'WR', 130, 130.32),
    ('Rhamondre Stevenson', 'RB', 131, 130.81),
    ('Courtland Sutton', 'WR', 132, 131.56),
    ('Jacory Croskey-Merritt', 'RB', 133, 133.04),
    ('Romeo Doubs', 'WR', 134, 134.34),
    ('Chris Bell', 'WR', 135, 134.7),
    ('Jonathon Brooks', 'RB', 136, 135.9),
    ('Nicholas Singleton', 'RB', 137, 135.92),
    ('Jakobi Meyers', 'WR', 138, 137.76),
    ('J.K. Dobbins', 'RB', 139, 138.88),
    ('AJ Barner', 'TE', 140, 139.65),
    ('Jalen Coker', 'WR', 141, 141.4),
    ('Tony Pollard', 'RB', 142, 143.0),
    ('Chris Godwin', 'WR', 143, 144.14),
    ('Germie Bernard', 'WR', 144, 144.33),
    ('Kenneth Gainwell', 'RB', 145, 146.65),
    ('Mark Andrews', 'TE', 146, 146.81),
    ('Woody Marks', 'RB', 147, 150.57),
    ('Tyler Allgeier', 'RB', 148, 151.89),
    ('Jordan Mason', 'RB', 149, 152.01),
    ('Khalil Shakir', 'WR', 150, 152.4),
    ('Travis Kelce', 'TE', 151, 154.32),
    ('Jacoby Brissett', 'QB', 152, 154.33),
    ('Michael Penix Jr.', 'QB', 153, 155.53),
    ('Juwan Johnson', 'TE', 154, 157.15),
    ('Tua Tagovailoa', 'QB', 155, 157.17),
    ('Elijah Sarratt', 'WR', 156, 157.21),
    ('Jalen McMillan', 'WR', 157, 157.41),
    ('Rachaad White', 'RB', 158, 157.97),
    ('Antonio Williams', 'WR', 159, 159.57),
    ('T.J. Hockenson', 'TE', 160, 159.66),
    ('Chigoziem Okonkwo', 'TE', 161, 161.33),
    ('Shedeur Sanders', 'QB', 162, 161.33),
    ('Emmett Johnson', 'RB', 163, 161.53),
    ('Dallas Goedert', 'TE', 164, 161.63),
    ('Gunnar Helm', 'TE', 165, 165.15),
    ('Tyrone Tracy', 'RB', 166, 165.65),
    ('Geno Smith', 'QB', 167, 166.76),
    ('Rashid Shaheed', 'WR', 168, 168.97),
    ('Chris Brazzell', 'WR', 169, 172.25),
    ('Zachariah Branch', 'WR', 170, 172.5),
    ('Braelon Allen', 'RB', 171, 172.75),
    ('Chris Rodriguez Jr.', 'RB', 172, 173.44),
    ('J.J. McCarthy', 'QB', 173, 173.99),
    ('Tre Harris', 'WR', 174, 175.81),
    ('Kaytron Allen', 'RB', 176, 176.15),
    ('Brandon Aiyuk', 'WR', 177, 177.79),
    ('Dylan Sampson', 'RB', 178, 178.02),
    ]
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM market_data WHERE source='ddl'")
    if c.fetchone()[0] > 50:
        conn.close()
        return  # Already loaded
    for name, pos, rank, adp in DDL_PLAYERS:
        # Store ADP as value (multiplied by 100 to keep as int, adp_float stored in team field)
        c.execute('''INSERT OR REPLACE INTO market_data
                    (source, player_name, rank, value, position, team, updated_at)
                    VALUES (?, ?, ?, ?, ?, ?, ?)''',
                 ('ddl', name, rank, int(adp * 10), pos, str(adp), datetime.now().isoformat()))
    conn.commit()
    conn.close()

seed_ddl_market_data()

def seed_player_values():
    """
    Seed player_values table from current hardcoded values in system prompt.
    Called once on startup. If table already has data, skip.
    """
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM player_values")
    if c.fetchone()[0] > 0:
        conn.close()
        return  # already seeded

    # Parse the values block embedded in SYSTEM_PROMPT
    import re as _re
    pattern = _re.compile(
        r'^(.+?)\s+\((\w+)\)\s+\|\s+MY:(\d+)\s+KTC:(\d+)\s+Δ:([+-]?\d+)\s+\|\s+(.+?)(?:\s+★)?$',
        _re.MULTILINE
    )
    now = datetime.now().isoformat()
    count = 0
    for m in pattern.finditer(SYSTEM_PROMPT):
        name, pos, my_val, ktc_val, delta, tier = (
            m.group(1).strip(), m.group(2), int(m.group(3)),
            int(m.group(4)), int(m.group(5)), m.group(6).strip()
        )
        try:
            c.execute("""INSERT OR REPLACE INTO player_values
                        (player_name, position, my_value, ktc_value, delta, tier, last_updated)
                        VALUES (?,?,?,?,?,?,?)""",
                     (name, pos, my_val, ktc_val, delta, tier, now))
            count += 1
        except: pass
    conn.commit()
    conn.close()
    if count > 0:
        print(f"  Seeded {count} player values from system prompt")


def fetch_sleeper_rosters_live(league_key):
    """
    Fetch ALL manager rosters directly from Sleeper API at request time.
    No DB dependency — always current. Called during chat request.
    Falls back to empty string if API unavailable.
    """
    league_id = SLEEPER_LEAGUE_IDS.get(league_key)
    if not league_id:
        return ""

    try:
        players_r = requests.get("https://api.sleeper.app/v1/players/nfl", timeout=20)
        if players_r.status_code != 200:
            return ""
        players_db = players_r.json()

        users_r = requests.get(
            f"https://api.sleeper.app/v1/league/{league_id}/users", timeout=10)
        rosters_r = requests.get(
            f"https://api.sleeper.app/v1/league/{league_id}/rosters", timeout=10)
        if users_r.status_code != 200 or rosters_r.status_code != 200:
            return ""

        users   = users_r.json()
        rosters = rosters_r.json()
        user_map = {u['user_id']: u.get('display_name', u.get('username','?'))
                    for u in users}

        # Find dcatlet's user_id
        my_id = next(
            (u['user_id'] for u in users
             if u.get('username','').lower() == 'dcatlet'
             or u.get('display_name','').lower() == 'dcatlet'),
            None)

        pos_order = {'QB':0,'RB':1,'WR':2,'TE':3,'K':4,'DEF':5}
        lines = [f"\n{league_key.upper().replace('_',' ')} ALL ROSTERS (live from Sleeper):"]

        for roster in sorted(rosters, key=lambda r: r['roster_id']):
            owner_id = roster.get('owner_id','')
            mgr = user_map.get(owner_id, f"Team{roster['roster_id']}")
            is_me = (owner_id == my_id)

            by_pos = defaultdict(list)
            for pid in (roster.get('players',[]) or []):
                p = players_db.get(pid, {})
                name = f"{p.get('first_name','')} {p.get('last_name','')}".strip()
                pos  = p.get('position','?')
                if pos in ('QB','RB','WR','TE') and name:
                    by_pos[pos].append(name)

            parts = []
            for pos in ['QB','RB','WR','TE']:
                if by_pos[pos]:
                    parts.append(f"{pos}:{','.join(by_pos[pos])}")

            prefix = "★ MY TEAM — " if is_me else ""
            lines.append(f"{prefix}{mgr}: {' | '.join(parts)}")

        return '\n'.join(lines)

    except Exception:
        return ""


def get_league_context_block(league_key):
    """Wrapper — try live API first, fall back to DB."""
    live = fetch_sleeper_rosters_live(league_key)
    if live:
        return live
    # Fallback: read from DB if API unavailable
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""SELECT manager, position, player_name FROM league_rosters
                 WHERE league=?
                 ORDER BY manager, CASE position
                   WHEN 'QB' THEN 1 WHEN 'RB' THEN 2
                   WHEN 'WR' THEN 3 WHEN 'TE' THEN 4 ELSE 5 END""", (league_key,))
    rows = c.fetchall()
    conn.close()
    if not rows:
        return ""
    mgr_pos = defaultdict(lambda: defaultdict(list))
    for mgr, pos, name in rows:
        if pos in ('QB','RB','WR','TE'):
            mgr_pos[mgr][pos].append(name)
    lines = [f"\n{league_key.upper().replace('_',' ')} ALL ROSTERS (from DB):"]
    for mgr in sorted(mgr_pos.keys()):
        parts = []
        for pos in ['QB','RB','WR','TE']:
            if mgr_pos[mgr][pos]:
                parts.append(f"{pos}:{','.join(mgr_pos[mgr][pos])}")
        lines.append(f"{mgr}: {' | '.join(parts)}")
    return '\n'.join(lines)


    """Pull dcatlet's roster from league_rosters DB. Flexible manager name matching."""
    MY_NAMES = {'dcatlet', 'mjbrutus', 'capital gains', 'twenty run savages'}

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Get all distinct managers in this league
    c.execute("SELECT DISTINCT manager FROM league_rosters WHERE league=?", (league_key,))
    all_managers = [r[0] for r in c.fetchall()]

    if not all_managers:
        conn.close()
        return fallback if fallback else f"[{league_key} roster not synced — tap \u23f3 Sync Sleeper]"

    # Strip arrow markers and whitespace before matching
    # DB stores names like "Capital Gains  ←" — need to clean before comparing
    def clean(s):
        return s.replace('\u2190','').replace('\u2192','').replace('←','').replace('→','').strip().lower().replace(' ','')

    my_manager = None
    for mgr in all_managers:
        if clean(mgr) in {n.replace(' ','') for n in MY_NAMES}:
            my_manager = mgr
            break
    if not my_manager:
        # Fallback: partial match
        for mgr in all_managers:
            cm = clean(mgr)
            if any(n.replace(' ','') in cm or cm in n.replace(' ','') for n in MY_NAMES):
                my_manager = mgr
                break
    if not my_manager:
        # Last resort: manager with most players
        c.execute("""SELECT manager, COUNT(*) as cnt FROM league_rosters
                     WHERE league=? GROUP BY manager ORDER BY cnt DESC LIMIT 1""", (league_key,))
        row = c.fetchone()
        if row: my_manager = row[0]

    if not my_manager:
        conn.close()
        return f"[{league_key}: could not identify dcatlet. Managers: {', '.join(all_managers[:8])}]"

    c.execute("""SELECT position, player_name, team FROM league_rosters
                 WHERE league=? AND manager=?
                 ORDER BY CASE position
                   WHEN 'QB' THEN 1 WHEN 'RB' THEN 2
                   WHEN 'WR' THEN 3 WHEN 'TE' THEN 4
                   WHEN 'K'  THEN 5 ELSE 6 END, player_name""",
              (league_key, my_manager))
    rows = c.fetchall()
    conn.close()

    if not rows:
        return fallback if fallback else f"[{league_key} roster empty for manager: {my_manager}]"

    by_pos = defaultdict(list)
    for pos, name, team in rows:
        by_pos[pos].append(f"{name}({team})" if team else name)

    total = sum(len(v) for v in by_pos.values())
    lines = [f"MY {league_key.upper().replace('_',' ')} ROSTER — {total} players (manager: {my_manager}):"]
    for pos in ['QB','RB','WR','TE','K']:
        if by_pos[pos]:
            lines.append(f"{pos}: {' | '.join(by_pos[pos])}")
    return '\n'.join(lines)


@app.route('/api/values/debug_upload', methods=['POST'])
def debug_upload():
    """Debug endpoint: shows exactly what the parser sees in an uploaded Excel file."""
    data = request.json
    file_obj = data.get('file_data', {})
    raw = file_obj.get('data', '')
    if not raw:
        return jsonify({"error": "No file data"})
    try:
        import io, base64 as b64lib
        from openpyxl import load_workbook as _lwb
        b64 = raw.split(',')[1] if ',' in raw else raw
        wb_xl = _lwb(io.BytesIO(b64lib.b64decode(b64)), data_only=True)

        result = {"sheets": []}
        for sname in wb_xl.sheetnames:
            ws = wb_xl[sname]
            # Show first 3 rows of each sheet
            rows_preview = []
            for r in range(1, 5):
                row_vals = [str(ws.cell(r, c).value or '') for c in range(1, 26)]
                rows_preview.append(row_vals)
            result["sheets"].append({
                "name": sname,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "first_4_rows": rows_preview
            })
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)})


@app.route('/api/values/update', methods=['POST'])
def update_player_values():
    data         = request.json
    file_data    = data.get('file_data')
    ktc_paste    = data.get('ktc_paste', '')
    manual       = data.get('manual_updates', [])
    now          = datetime.now().isoformat()

    conn = sqlite3.connect(DB_PATH)
    c    = conn.cursor()
    updated = 0
    errors  = []

    if file_data:
        try:
            import io, base64 as b64lib
            from openpyxl import load_workbook as _lwb
            raw = file_data.get('data','')
            b64 = raw.split(',')[1] if ',' in raw else raw
            wb_xl = _lwb(io.BytesIO(b64lib.b64decode(b64)), data_only=True)

            for sname in wb_xl.sheetnames:
                ws_xl = wb_xl[sname]
                if ws_xl.max_row < 5: continue

                # Read header row — try rows 1-3
                headers = []
                header_row = None
                for try_row in range(1, 4):
                    h = [str(ws_xl.cell(try_row, i).value or '').strip()
                         for i in range(1, min(ws_xl.max_column+1, 30))]
                    # Valid header row: has Position, Team, and at least one value column
                    has_pos   = any('position' in x.lower() for x in h)
                    has_value = any('value' in x.lower() or 'my value' in x.lower() for x in h)
                    if has_pos and has_value:
                        headers = h
                        header_row = try_row
                        break

                if not headers:
                    # Special case: Community Trade Value format
                    # Row 1 col A = timestamp, col C = Position, col D = Team, col E = KTC SF value
                    h1 = [str(ws_xl.cell(1, i).value or '').strip()
                          for i in range(1, min(ws_xl.max_column+1, 30))]
                    # Detect by checking if col C looks like a position header
                    if len(h1) >= 5 and 'position' in h1[2].lower():
                        headers = h1
                        header_row = 1

                if not headers or header_row is None: continue

                hl = [h.lower() for h in headers]

                def find_col(keywords):
                    for kw in keywords:
                        for i, h in enumerate(hl):
                            if kw in h: return i
                    return None

                def find_col_exact(keywords):
                    """Prefer exact match, fall back to contains."""
                    for kw in keywords:
                        for i, h in enumerate(hl):
                            if h == kw.lower(): return i
                    return find_col(keywords)

                # Column mapping — use exact match where possible to avoid
                # "Position Rank" matching before "Position"
                name_col = 0  # Col A always player name
                pos_col  = find_col_exact(['position']) or 2
                team_col = find_col_exact(['team']) or 3

                # KTC SF TE+ value: col E "Value" (not "KTC 1QB Value", not "Redraft")
                ktc_col = None
                for i, h in enumerate(hl):
                    if h == 'value':  # exact match for col E
                        ktc_col = i; break
                if ktc_col is None:
                    for i, h in enumerate(hl):
                        if 'value' in h and 'ktc' in h and '1qb' not in h and 'redraft' not in h:
                            ktc_col = i; break
                if ktc_col is None: ktc_col = 4

                # My Value: exact label "my value"
                my_col = find_col_exact(['my value'])

                # FantasyCalc SF value — must have actual values, not just a header label
                fc_col = None
                for i, h in enumerate(hl):
                    if 'fantasycalc' in h and 'value' in h and '1qb' not in h \
                       and 'redraft' not in h and '->' not in h:
                        fc_col = i; break

                age_col  = find_col(['age'])
                tier_col = find_col(['tier'])

                rows_parsed = 0
                for row in ws_xl.iter_rows(min_row=header_row+1, values_only=True):
                    try:
                        if all(v is None for v in row): continue
                        name = str(row[name_col] or '').strip()
                        if not name or name in ('None', 'Player', 'Name') or len(name) < 3: continue
                        # Skip rows that look like headers or section dividers
                        if name.lower() in ('quarterback','running back','wide receiver','tight end',
                                            'qb','rb','wr','te','pick','player','name'): continue

                        pos  = str(row[pos_col] or '').strip()  if pos_col < len(row) else ''
                        team = str(row[team_col] or '').strip() if team_col < len(row) else ''
                        tier = str(row[tier_col] or '').strip() if tier_col is not None and tier_col < len(row) else ''

                        def safe_int(val):
                            if val is None: return 0
                            try: return int(float(str(val).replace(',','')))
                            except: return 0

                        ktc_val = safe_int(row[ktc_col]) if ktc_col < len(row) else 0
                        my_val  = safe_int(row[my_col])  if my_col is not None and my_col < len(row) else 0
                        fc_val  = safe_int(row[fc_col])  if fc_col is not None and fc_col < len(row) else 0

                        if ktc_val <= 0 and my_val <= 0 and fc_val <= 0: continue

                        # If no personal value set, use KTC as both
                        use_my  = my_val if my_val > 0 else ktc_val
                        use_ktc = ktc_val if ktc_val > 0 else fc_val
                        delta   = use_my - use_ktc if use_ktc > 0 else 0

                        c.execute("""INSERT OR REPLACE INTO player_values
                                    (player_name, position, my_value, ktc_value, delta, tier, last_updated)
                                    VALUES (?,?,?,?,?,?,?)""",
                                 (name, pos, use_my, use_ktc, delta, tier, now))
                        c.execute("""INSERT OR REPLACE INTO market_data
                                    (source, player_name, rank, value, position, team, updated_at)
                                    VALUES ('ktc',?,?,?,?,?,?)""",
                                 (name, updated+1, use_ktc, pos, team, now))
                        updated += 1
                        rows_parsed += 1
                    except: pass

                if rows_parsed > 10:
                    break  # found a valid data sheet

        except Exception as ex:
            errors.append(f"Excel parse error: {ex}")

    if ktc_paste:
        # Parse KTC paste inline — same logic as /api/market/paste
        try:
            lines = ktc_paste.strip().split('\n')
            ktc_rank = 0
            ktc_parsed = 0
            for line in lines:
                line = line.strip()
                if not line: continue
                val_match = re.search(r'\b(\d{3,5})\s*$', line)
                if not val_match: continue
                val = int(val_match.group(1))
                if val < 100: continue
                pos_match = re.search(r'\b(QB|RB|WR|TE|PICK)\b', line)
                if not pos_match: continue
                pos = pos_match.group(1)
                team_match = re.search(r'\b([A-Z]{2,3})\b', line)
                team = team_match.group(1) if team_match else 'FA'
                name_part = re.sub(r'\b(QB|RB|WR|TE|PICK|QB\d+|RB\d+|WR\d+|TE\d+)\b.*', '', line)
                name_part = re.sub(r'^\d+\s*', '', name_part)
                name_part = re.sub(r'\b[A-Z]{2,3}\b', '', name_part)
                name = re.sub(r'\s+', ' ', name_part).strip()
                name = re.sub(r"[^\w\s'.]", '', name).strip()
                if len(name) < 3: continue
                ktc_rank += 1
                # Store in market_data
                c.execute("""INSERT OR REPLACE INTO market_data
                            (source, player_name, rank, value, position, team, updated_at)
                            VALUES ('ktc',?,?,?,?,?,?)""",
                         (name, ktc_rank, val, pos, team, now))
                # Update player_values if player exists there
                c.execute("""UPDATE player_values SET ktc_value=?, delta=my_value-?,
                            last_updated=? WHERE LOWER(player_name)=LOWER(?)""",
                         (val, val, now, name))
                ktc_parsed += 1
            updated += ktc_parsed
        except Exception as ex:
            errors.append(f"KTC paste error: {ex}")

    for item in manual:
        try:
            name    = item.get('name','').strip()
            my_val  = int(item.get('my_value', 0))
            ktc_val = int(item.get('ktc_value', 0))
            pos     = item.get('position','')
            tier    = item.get('tier','')
            delta   = my_val - ktc_val
            c.execute("""INSERT OR REPLACE INTO player_values
                        (player_name,position,my_value,ktc_value,delta,tier,last_updated)
                        VALUES (?,?,?,?,?,?,?)""",
                     (name,pos,my_val,ktc_val,delta,tier,now))
            updated += 1
        except: pass

    conn.commit()
    conn.close()

    return jsonify({
        "success": True,
        "updated": updated,
        "errors": errors,
        "message": f"Updated {updated} player values. Changes take effect immediately in all chat responses."
    })


@app.route('/api/values/summary', methods=['GET'])
def values_summary():
    """Return current player values — merges player_values (composite) and market_data (KTC paste)."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Primary: player_values (your composite)
    c.execute("""SELECT player_name, position, my_value, ktc_value, delta, tier, last_updated
                 FROM player_values ORDER BY my_value DESC""")
    pv_rows = c.fetchall()

    # Secondary: market_data KTC — players not in player_values
    c.execute("""SELECT m.player_name, m.rank, m.value
                 FROM market_data m
                 WHERE m.source='ktc'
                 AND LOWER(m.player_name) NOT IN
                     (SELECT LOWER(player_name) FROM player_values)
                 ORDER BY m.rank ASC LIMIT 300""")
    md_rows = c.fetchall()

    c.execute("SELECT MAX(last_updated) FROM player_values")
    last_updated = c.fetchone()[0] or 'Never'
    conn.close()

    values = [{"name":r[0],"pos":r[1],"my_value":r[2],"ktc_value":r[3],
               "delta":r[4],"tier":r[5]} for r in pv_rows]

    # Add market_data players with ktc_value only (no my_value)
    for name, rank, ktc_val in md_rows:
        values.append({"name":name,"pos":"","my_value":0,
                       "ktc_value":ktc_val,"delta":0,"tier":"KTC only"})

    return jsonify({
        "success": True,
        "count": len(values),
        "last_updated": last_updated,
        "values": values
    })



_do_refresh_ktc_tiers()
seed_player_values()

# VS pick map — who holds each pick (updated as trades happen)
VS_PICK_MAP = {
    'pdwyer13': {'picks': ['1.01','2.12','3.12','4.01','5.12','6.01','7.12','8.01','9.12','10.01','11.12','12.01','13.12','14.01','15.12','16.01','17.12','18.01','19.12','20.01','21.12','22.01','23.12','24.01','25.12','26.01','27.12','28.01']},
    'dcatlet':  {'picks': ['1.02','2.11','3.11','4.02','5.11','6.02','7.11','8.02','9.11','10.02','11.11','12.02','13.11','14.02','15.11','16.02','17.11','18.02','19.11','20.02','21.11','22.02','23.11','24.02','25.11','26.02','27.11','28.02']},
    'yerkdog':  {'picks': ['1.03','2.10','3.10','4.03','5.10','6.03','7.10','8.03','9.10','10.03','11.10','12.03','13.10','14.03','15.10','16.03','17.10','18.03','19.10','20.03','21.10','22.03','23.10','24.03','25.10','26.03','27.10','28.03']},
    'jefisk24': {'picks': ['1.04','2.09','3.09','4.04','5.09','6.04','7.09','8.04','9.09','10.04','11.09','12.04','13.09','14.04','15.09','16.04','17.09','18.04','19.09','20.04','21.09','22.04','23.09','24.04','25.09','26.04','27.09','28.04']},
    'Smohr609': {'picks': ['1.05','2.08','3.08','4.05','5.08','6.05','7.08','8.05','9.08','10.05','11.08','12.05','13.08','14.05','15.08','16.05','17.08','18.05','19.08','20.05','21.08','22.05','23.08','24.05','25.08','26.05','27.08','28.05']},
    'ColeTrain8300': {'picks': ['1.06','2.07','3.07','4.06','5.07','6.06','7.07','8.06','9.07','10.06','11.07','12.06','13.07','14.06','15.07','16.06','17.07','18.06','19.07','20.06','21.07','22.06','23.07','24.06','25.07','26.06','27.07','28.06']},
    'DrTrollPhD': {'picks': ['1.07','2.06','3.06','4.07','5.06','6.07','7.06','8.07','9.06','10.07','11.06','12.07','13.06','14.07','15.06','16.07','17.06','18.07','19.06','20.07','21.06','22.07','23.06','24.07','25.06','26.07','27.06','28.07']},
    'colinmonie': {'picks': ['1.08','2.05','3.05','4.08','5.05','6.08','7.05','8.08','9.05','10.08','11.05','12.08','13.05','14.08','15.05','16.08','17.05','18.08','19.05','20.08','21.05','22.08','23.05','24.08','25.05','26.08','27.05','28.08']},
    'EazyDakar': {'picks': ['1.09','2.04','3.04','4.09','5.04','6.09','7.04','8.09','9.04','10.09','11.04','12.09','13.04','14.09','15.04','16.09','17.04','18.09','19.04','20.09','21.04','22.09','23.04','24.09','25.04','26.09','27.04','28.09']},
    'jakemills69': {'picks': ['1.10','2.03','3.03','4.10','5.03','6.10','7.03','8.10','9.03','10.10','11.03','12.10','13.03','14.10','15.03','16.10','17.03','18.10','19.03','20.10','21.03','22.10','23.03','24.10','25.03','26.10','27.03','28.10']},
    'NateSneller': {'picks': ['1.11','2.02','3.02','4.11','5.02','6.11','7.02','8.11','9.02','10.11','11.02','12.11','13.02','14.11','15.02','16.11','17.02','18.11','19.02','20.11','21.02','22.11','23.02','24.11','25.02','26.11','27.02','28.11']},
    'sneller':  {'picks': ['1.12','2.01','3.01','4.12','5.01','6.12','7.01','8.12','9.01','10.12','11.01','12.12','13.01','14.12','15.01','16.12','17.01','18.12','19.01','20.12','21.01','22.12','23.01','24.12','25.01','26.12','27.01','28.12']},
}

# KTC value adjustment table for multi-asset trades
# Based on KTC's published adjustment factors
TRADE_ADJUSTMENT = {
    1: 1.00,  # 1 asset: no adjustment
    2: 0.90,  # 2 assets: each worth 90% (10% discount for complexity)
    3: 0.80,  # 3 assets: each worth 80%
    4: 0.72,  # 4 assets: each worth 72%
    5: 0.65,  # 5+ assets: diminishing returns
}

def adjust_trade_value(assets_values):
    """Apply KTC multi-asset discount"""
    n = len(assets_values)
    factor = TRADE_ADJUSTMENT.get(min(n, 5), 0.65)
    return sum(v * factor for v in assets_values)

def get_pick_owner(pick_slot):
    """Get current owner of a pick slot from VS_PICK_MAP"""
    for manager, data in VS_PICK_MAP.items():
        if pick_slot in data['picks']:
            return manager
    return None

def get_startup_pick_value(pick_slot):
    """Get KTC startup value for a pick slot"""
    STARTUP_VALUES = {
        '1.01':9999,'1.02':9987,'1.03':9918,'1.04':9565,'1.05':9412,'1.06':8765,
        '1.07':8603,'1.08':8337,'1.09':7935,'1.10':7861,'1.11':7838,'1.12':7778,
        '2.01':7713,'2.02':7697,'2.03':7679,'2.04':7566,'2.05':7499,'2.06':7287,
        '2.07':6888,'2.08':6887,'2.09':6818,'2.10':6814,'2.11':6702,'2.12':6692,
        '3.01':6673,'3.02':6556,'3.03':6328,'3.04':6220,'3.05':6197,'3.06':6181,
        '3.07':6139,'3.08':6056,'3.09':6044,'3.10':5957,'3.11':5841,'3.12':5830,
        '4.01':5719,'4.02':5672,'4.03':5668,'4.04':5638,'4.05':5571,'4.06':5474,
        '4.07':5461,'4.08':5459,'4.09':5402,'4.10':5373,'4.11':5357,'4.12':5356,
        '5.01':5251,'5.02':5243,'5.03':5214,'5.04':5197,'5.05':5076,'5.06':5042,
        '5.07':4993,'5.08':4992,'5.09':4929,'5.10':4925,'5.11':4920,'5.12':4904,
        '6.01':4899,'6.02':4898,'6.03':4889,'6.04':4885,'6.05':4866,'6.06':4853,
        '6.07':4850,'6.08':4810,'6.09':4747,'6.10':4741,'6.11':4710,'6.12':4696,
        '7.01':4695,'7.02':4620,'7.03':4619,'7.04':4604,'7.05':4560,'7.06':4512,
        '7.07':4494,'7.08':4464,'7.09':4454,'7.10':4428,'7.11':4417,'7.12':4338,
        '8.01':4110,'8.02':4103,'8.03':4059,'8.04':3969,'8.05':3967,'8.06':3960,
        '8.07':3945,'8.08':3924,'8.09':3915,'8.10':3880,'8.11':3866,'8.12':3825,
    }
    return STARTUP_VALUES.get(pick_slot, 3000)

# ============================================================
# DRAFT DECISION ENGINE
# ============================================================

@app.route('/api/draft/decisions', methods=['GET'])
def get_draft_decisions():
    """
    Pre-draft and live decision signals for each of your 28 picks.
    For each pick: Stick | Reach | Trade Up | Trade Back
    With specific trade packages and manager targets.
    """
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Get your VS rankings
    c.execute("SELECT player_name, position, elo_score FROM vs_rankings ORDER BY elo_score DESC LIMIT 200")
    vs_rows = c.fetchall()
    your_rank = {r[0]: {'rank': i+1, 'pos': r[1], 'elo': r[2]} for i, r in enumerate(vs_rows)}

    # Get DDL ADP
    c.execute("SELECT player_name, rank, team FROM market_data WHERE source='ddl' ORDER BY rank")
    ddl_rows = c.fetchall()
    ddl_rank = {}
    for name, rank, adp_str in ddl_rows:
        try: ddl_rank[name] = {'rank': rank, 'adp': float(adp_str)}
        except: ddl_rank[name] = {'rank': rank, 'adp': rank}

    # Get already drafted players
    c.execute("SELECT pick_slot, player_name, manager FROM draft_activity WHERE draft_id='vs_2026' ORDER BY pick_slot")
    drafted = {r[1]: {'slot': r[0], 'manager': r[2]} for r in c.fetchall()}
    drafted_players = set(drafted.keys())

    conn.close()

    my_picks = ['1.02','2.11','3.11','4.02','5.11','6.02','7.11','8.02',
                '9.11','10.02','11.11','12.02','13.11','14.02','15.11','16.02',
                '17.11','18.02','19.11','20.02','21.11','22.02','23.11','24.02',
                '25.11','26.02','27.11','28.02']

    # Remove already used picks
    remaining_picks = [p for p in my_picks if not any(d['slot'] == p for d in drafted.values() if d['manager'] == 'dcatlet')]

    decisions = []

    for i, my_pick in enumerate(remaining_picks[:14]):  # Focus on meaningful picks
        my_pick_val = get_startup_pick_value(my_pick)
        pick_num = my_picks.index(my_pick) + 1  # Overall pick number in draft order

        # Find best available players at this pick window (±3 picks of my slot)
        # by DDL ADP
        pick_slot_num = int(my_pick.split('.')[0]) * 12 + int(my_pick.split('.')[1])

        # Players likely available at this pick (DDL ADP within window)
        available_window = []
        for name, dd in ddl_rank.items():
            if name in drafted_players:
                continue
            if abs(dd['adp'] - (pick_slot_num - 1)) <= 6:
                yr = your_rank.get(name, {})
                available_window.append({
                    'name': name,
                    'pos': yr.get('pos', 'WR'),
                    'your_rank': yr.get('rank', 200),
                    'ddl_adp': dd['adp'],
                    'ddl_rank': dd['rank'],
                    'delta': yr.get('rank', 200) - dd['rank'],
                })

        available_window.sort(key=lambda x: x['your_rank'])
        top_target = available_window[0] if available_window else None

        # Determine decision
        decision = {'pick': my_pick, 'pick_value': my_pick_val, 'action': 'STICK',
                   'target': None, 'reasoning': '', 'trade_package': None}

        if top_target:
            delta = top_target['delta']
            adp = top_target['ddl_adp']
            name = top_target['name']

            # Find next pick after this one
            next_pick = remaining_picks[i+1] if i+1 < len(remaining_picks) else None
            next_pick_slot = (int(next_pick.split('.')[0]) * 12 + int(next_pick.split('.')[1])) if next_pick else 999
            picks_until_next = next_pick_slot - pick_slot_num if next_pick else 999

            if delta <= -20:
                # Market values much higher than you — trade back
                decision['action'] = 'TRADE BACK'
                decision['target'] = name
                decision['reasoning'] = f"You rank {name} #{top_target['your_rank']}, DDL has him at {adp:.1f}. Market overvalues him here. Trade back for future capital."
                # Find who picks next in sequence
                trade_back_target = _find_trade_back_target(my_pick, my_pick_val)
                decision['trade_package'] = trade_back_target

            elif delta >= 15 and picks_until_next > 8:
                # You value player much more than market AND they won't last to your next pick
                decision['action'] = 'TRADE UP'
                decision['target'] = name
                # Find who picks before this player goes
                goes_at_slot = int(adp)
                goes_at_pick = _slot_to_pick(goes_at_slot - 2)
                trade_up = _build_trade_up(my_pick, goes_at_pick, my_pick_val, remaining_picks, i)
                decision['trade_package'] = trade_up
                decision['reasoning'] = f"{name} is your #{top_target['your_rank']} but DDL ADP is {adp:.1f} — goes {picks_until_next} picks before your next selection. Trade up now."

            elif delta >= 10 and picks_until_next <= 5:
                # You value higher, but next pick is close — just reach slightly
                decision['action'] = 'REACH'
                decision['target'] = name
                decision['reasoning'] = f"{name}: your #{top_target['your_rank']} vs DDL {adp:.1f}. Next pick only {picks_until_next} slots away — slight reach is acceptable, don't risk losing him."

            else:
                decision['action'] = 'STICK'
                decision['target'] = name
                decision['reasoning'] = f"{name} projects available at {my_pick}. DDL ADP {adp:.1f} aligns with your slot. Draft as planned."

        decisions.append(decision)

    return jsonify({"decisions": decisions, "success": True,
                   "picks_remaining": remaining_picks, "drafted_count": len(drafted)})

def _slot_to_pick(slot_num):
    """Convert sequential slot number to pick notation"""
    if slot_num < 1: slot_num = 1
    rnd = (slot_num - 1) // 12 + 1
    pick = (slot_num - 1) % 12 + 1
    return f"{rnd}.{str(pick).zfill(2)}"

def _find_trade_back_target(my_pick, my_pick_val):
    """Find who to trade back with and what to get"""
    my_round = int(my_pick.split('.')[0])
    my_slot = int(my_pick.split('.')[1])

    # Find managers picking in the next 3 slots after mine
    next_slots = []
    for offset in range(1, 4):
        target_slot = my_slot + offset
        if target_slot > 12: break
        target_pick = f"{my_round}.{str(target_slot).zfill(2)}"
        owner = get_pick_owner(target_pick)
        if owner and owner != 'dcatlet':
            later_pick_val = get_startup_pick_value(f"{my_round + 2}.{str(my_slot).zfill(2)}")
            next_slots.append({
                'manager': owner,
                'their_pick': target_pick,
                'their_pick_val': get_startup_pick_value(target_pick),
                'offer': f"Trade {my_pick} ({my_pick_val:,}) to {owner} for {target_pick} ({get_startup_pick_value(target_pick):,}) + future capital",
                'value_gain': get_startup_pick_value(target_pick) - my_pick_val + later_pick_val,
            })
    return next_slots[0] if next_slots else None

def _build_trade_up(my_pick, target_pick, my_val, remaining_picks, current_idx):
    """Build a specific trade-up package with KTC value adjustment"""
    target_val = get_startup_pick_value(target_pick)
    target_owner = get_pick_owner(target_pick)
    deficit = target_val - my_val

    if deficit <= 0:
        return {'owner': target_owner, 'offer': f"{my_pick} straight up for {target_pick}",
                'value_check': 'Even value', 'deficit': 0}

    # Find sweetener picks from remaining picks
    sweetener_picks = []
    sweetener_total = 0
    for p in remaining_picks[current_idx+2:]:
        if sweetener_total >= deficit * 0.85:
            break
        pv = get_startup_pick_value(p)
        if pv < my_val * 0.4:  # Don't give up picks worth less than 40% of main pick
            sweetener_picks.append(p)
            sweetener_total += pv

    all_giving = [my_pick] + sweetener_picks
    giving_vals = [get_startup_pick_value(p) for p in all_giving]
    adjusted_total = adjust_trade_value(giving_vals)

    surplus_pct = round((adjusted_total - target_val) / target_val * 100, 1)
    sweetener_str = ' + '.join(sweetener_picks) if sweetener_picks else 'straight up'

    return {
        'owner': target_owner or 'Unknown',
        'target_pick': target_pick,
        'offer': f"Offer {my_pick} + {sweetener_str} to {target_owner} for {target_pick}",
        'your_total_raw': sum(giving_vals),
        'your_total_adjusted': round(adjusted_total),
        'their_value': target_val,
        'surplus_pct': surplus_pct,
        'verdict': 'Fair' if abs(surplus_pct) <= 10 else ('Overpay' if surplus_pct < 0 else 'Good value'),
        'sweetener_picks': sweetener_picks,
    }

# ============================================================
# SCREENSHOT PARSING — DRAFT BOARD + TRADE ACTIVITY
# ============================================================

@app.route('/api/draft/parse_screenshot', methods=['POST'])
def parse_draft_screenshot():
    """
    Parse a Sleeper draft board screenshot using Claude vision.
    Returns extracted picks for user confirmation before logging.
    """
    data = request.json
    image_data = data.get('image', '')
    draft_context = data.get('context', 'grid')  # 'grid' or 'trade'

    if not image_data:
        return jsonify({"success": False, "error": "No image provided"})

    image_content = image_data.split(',')[1] if ',' in image_data else image_data
    media_type = 'image/png' if 'png' in image_data[:30] else 'image/jpeg'

    if draft_context == 'trade':
        prompt = """This is a screenshot of a fantasy football draft trade or draft activity. 
Extract any trade information: which picks or players were exchanged, between which managers.
Return JSON only:
{"type": "trade", "trades": [{"from": "manager1", "to": "manager2", "assets_sent": ["pick or player"], "assets_received": ["pick or player"]}]}"""
    else:
        prompt = """This is a screenshot of a Sleeper fantasy football startup draft board.
The board shows a grid of picks already made (player name, team, position) and the current pick on clock.
Extract all visible drafted picks. For each: pick slot (e.g. 1.01, 1.02), player name, manager/team name if visible.
Return JSON only:
{"picks": [{"slot": "1.01", "player": "Josh Allen", "manager": "pdwyer13", "position": "QB"}], "current_pick": "2.03", "notes": "any relevant observations"}
If you cannot clearly read a pick, skip it rather than guess."""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=1500,
            messages=[{"role": "user", "content": [
                {"type": "image", "source": {"type": "base64", "media_type": media_type, "data": image_content}},
                {"type": "text", "text": prompt}
            ]}]
        )
        text = response.content[0].text.strip()
        start = text.find('{')
        end = text.rfind('}') + 1
        if start >= 0 and end > start:
            parsed = json.loads(text[start:end])
            return jsonify({"parsed": parsed, "success": True, "confirmed": False})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})

    return jsonify({"success": False, "error": "Could not parse image"})

@app.route('/api/draft/confirm_picks', methods=['POST'])
def confirm_draft_picks():
    """Commit confirmed parsed picks to draft_activity table"""
    data = request.json
    picks = data.get('picks', [])
    if not picks:
        return jsonify({"success": False, "error": "No picks to commit"})

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    committed = 0
    for pick in picks:
        slot = pick.get('slot', '')
        player = pick.get('player', '')
        manager = pick.get('manager', '')
        pos = pick.get('position', '')
        if slot and player:
            c.execute('''INSERT OR REPLACE INTO draft_activity
                        (draft_id, pick_slot, player_name, manager, position, logged_at, source)
                        VALUES (?, ?, ?, ?, ?, ?, ?)''',
                     ('vs_2026', slot, player, manager, pos,
                      datetime.now().isoformat(), 'screenshot'))
            committed += 1
    conn.commit()
    conn.close()
    return jsonify({"success": True, "committed": committed})

@app.route('/api/draft/activity', methods=['GET'])
def get_draft_activity():
    """Get all logged draft activity"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''SELECT pick_slot, player_name, manager, position, logged_at, source
                FROM draft_activity WHERE draft_id='vs_2026'
                ORDER BY pick_slot ASC''')
    rows = c.fetchall()
    conn.close()
    picks = [{'slot': r[0], 'player': r[1], 'manager': r[2],
              'position': r[3], 'logged_at': r[4], 'source': r[5]} for r in rows]

    # Build manager pick counts for tendency analysis
    manager_pos = defaultdict(lambda: defaultdict(int))
    for p in picks:
        if p['manager'] and p['position']:
            manager_pos[p['manager']][p['position']] += 1

    tendencies = {}
    for mgr, pos_counts in manager_pos.items():
        total = sum(pos_counts.values())
        tendencies[mgr] = {pos: round(count/total*100) for pos, count in pos_counts.items() if total > 0}

    return jsonify({"picks": picks, "total": len(picks), "tendencies": tendencies, "success": True})

@app.route('/api/draft/live_recommendation', methods=['POST'])
def live_draft_recommendation():
    """
    Real-time draft recommendation for your current pick.
    Incorporates: your VS rankings, DDL ADP, picks already made,
    manager tendencies, and positional scarcity.
    """
    data = request.json
    current_pick = data.get('current_pick', '')
    picks_made = data.get('picks_made', [])
    my_roster = data.get('my_roster', [])
    available_input = data.get('available', '')

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # Get manager tendencies from logged draft activity
    c.execute('''SELECT manager, position, COUNT(*) as cnt FROM draft_activity
                WHERE draft_id='vs_2026' GROUP BY manager, position''')
    tendency_rows = c.fetchall()

    # Get VS rankings top 50
    c.execute("SELECT player_name, position, elo_score FROM vs_rankings ORDER BY elo_score DESC LIMIT 60")
    vs_top = c.fetchall()

    # Get DDL ADP top 100
    c.execute("SELECT player_name, rank, team FROM market_data WHERE source='ddl' ORDER BY rank LIMIT 100")
    ddl_top = {r[0]: {'rank': r[1], 'adp': float(r[2]) if r[2] else r[1]} for r in c.fetchall()}

    conn.close()

    # Build manager tendency summary
    mgr_pos = defaultdict(lambda: defaultdict(int))
    for mgr, pos, cnt in tendency_rows:
        mgr_pos[mgr][pos] = cnt
    mgr_summary = {}
    for mgr, pos_counts in mgr_pos.items():
        total = sum(pos_counts.values())
        if total > 0:
            top_pos = sorted(pos_counts.items(), key=lambda x: -x[1])[:2]
            mgr_summary[mgr] = f"favors {', '.join([f'{p}({c})' for p,c in top_pos])}"

    drafted_names = [p.get('player', '') for p in picks_made]
    vs_available = [(r[0], r[1], r[2]) for r in vs_top if r[0] not in drafted_names]

    pick_num = current_pick
    my_next_picks = []
    all_my = ['1.02','2.11','3.11','4.02','5.11','6.02','7.11','8.02',
              '9.11','10.02','11.11','12.02','13.11','14.02','15.11','16.02',
              '17.11','18.02','19.11','20.02','21.11','22.02','23.11','24.02',
              '25.11','26.02','27.11','28.02']

    if current_pick in all_my:
        idx = all_my.index(current_pick)
        my_next_picks = all_my[idx+1:idx+4]

    prompt = f"""You are the draft assistant for MJBrutus (dcatlet) in the Velvet Spade startup draft.
FORMAT: SuperFlex | TE Premium 1.5x | 6pt TDs | 12 teams | 28 rounds | Snake + 3rd round reversal

CURRENT PICK: {current_pick}
MY NEXT PICKS: {', '.join(my_next_picks)}
MY CURRENT ROSTER: {', '.join(my_roster) if my_roster else 'None yet'}

TOP AVAILABLE BY MY VS RANKINGS (6pt TD adjusted):
{chr(10).join([f"#{i+1} {r[0]} ({r[1]})" for i, r in enumerate(vs_available[:15])])}

DDL ADP CONTEXT: {', '.join([f"{n} (ADP {d['adp']:.1f})" for n, d in list(ddl_top.items())[:8] if n not in drafted_names])}

PICKS MADE SO FAR ({len(picks_made)} total):
{chr(10).join([f"{p.get('pick','?')} {p.get('player','?')} ({p.get('team','?')})" for p in picks_made[-12:]]) if picks_made else 'Draft just started'}

MANAGER TENDENCIES FROM DRAFT SO FAR:
{chr(10).join([f"{mgr}: {tendency}" for mgr, tendency in mgr_summary.items()]) if mgr_summary else 'Not enough data yet'}

AVAILABLE INPUT: {available_input[:500] if available_input else 'Not provided'}

Provide:
1. PRIMARY RECOMMENDATION — specific player to draft with 2-sentence reason
2. DECISION TYPE — one of: STICK (draft as planned) | REACH (take now before next pick) | TRADE UP (package to move up) | TRADE BACK (pick is weak, get future capital)
3. If TRADE UP: who to approach and what to offer (use manager names from the pick map)
4. If TRADE BACK: who picks next and what to request
5. ALTERNATIVES — 2 backup options if primary goes
6. POSITIONAL SCARCITY ALERT — flag any position tier breaks happening now

Be decisive. One clear pick first. Mobile-friendly format."""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=1000,
            system=SYSTEM_PROMPT,
            tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 1}],
            messages=[{"role": "user", "content": prompt}]
        )
        rec = "".join(b.text for b in response.content if hasattr(b, 'text'))
        return jsonify({"recommendation": rec, "success": True,
                       "my_next_picks": my_next_picks})
    except Exception as e:
        return jsonify({"error": str(e), "success": False})

@app.route('/api/trade/analyze_screenshots', methods=['POST'])
def analyze_trade_screenshots():
    data        = request.json
    images      = data.get('images', [])[:6]
    prompt_text = data.get('prompt', '')
    league      = data.get('league', 'dynasty league')
    context     = data.get('context', '')
    prompt_type = data.get('prompt_type', 'screenshot_eval')

    # Pull player values from DB
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""SELECT player_name, my_value, ktc_value, delta, tier
                 FROM player_values ORDER BY my_value DESC LIMIT 150""")
    value_rows = c.fetchall()
    conn.close()

    # Build compact value reference
    value_ref = "MY PLAYER VALUES (MY=personal composite, KTC=market, Δ=difference, ★=significant divergence):\n"
    for name, my_val, ktc_val, delta, tier in value_rows:
        d = f"+{delta}" if delta >= 0 else str(delta)
        flag = " ★" if abs(delta) > 500 else ""
        value_ref += f"{name}: MY {my_val:,} KTC {ktc_val:,} Δ{d} | {tier}{flag}\n"

    league_settings = {
        "Velvet Spade":        "12-team SuperFlex | 1.5x TE Premium | 6pt Pass TD | DRAFT COMPLETE — trade season",
        "Capital Gains":       "FFPC #430 | SuperFlex | Dynasty | Rebuild — heavy pick capital",
        "Twenty Run Savages":  "FFPC #210 | SuperFlex | Dynasty | Competing — Drake Maye core",
        "Gentleman's Dynasty": "Sleeper | Dynasty | Rebuild phase | Mahomes+Bowers untouchable",
    }
    league_ctx = league_settings.get(league, f"{league} — dynasty league")

    _vs = get_full_league_context('velvet_spade')
    _gl = get_full_league_context('gentlemans_dynasty')
    _sp = (SYSTEM_PROMPT
        .replace('{VS_ROSTER_BLOCK}',  _vs['my_roster'])
        .replace('{GL_ROSTER_BLOCK}',  _gl['my_roster'])
        .replace('{CG_ROSTER_BLOCK}',  get_roster_block('Capital Gains'))
        .replace('{TRS_ROSTER_BLOCK}', get_roster_block('TRS')))
    if _vs['all_rosters']: _sp += _vs['all_rosters']
    if _gl['all_rosters']: _sp += _gl['all_rosters']
    system = f"""{_sp}

{value_ref}

LEAGUE: {league} | {league_ctx}

━━━━ TRADE ANALYSIS PROTOCOL ━━━━

STEP 1 — ASSET EXTRACTION (do this first, silently):
Read the screenshot carefully. List every asset being traded by each side.
ONLY use player names clearly visible in the screenshot.
For any player not in MY PLAYER VALUES above, flag with ⚠ VALUE NOT IN DB.
Do NOT infer, combine, or generate player names from memory.

STEP 2 — NFL SITUATION CHECK (mandatory for every player mentioned):
Use web_search to verify each player's current NFL situation:
- Current team and QB (training data is outdated — always verify)
- Any injuries, suspensions, or role changes in the last 30 days
- If web search unavailable, flag claim as ⚠ UNVERIFIED — DO NOT state as fact

STEP 3 — VALUE CALCULATION:
Use MY values (not KTC) as primary. Show KTC as secondary reference.
PACKAGE VALUATION — KTC-STYLE DYNAMIC MODEL (always apply, always show the math):
This mirrors how KTC actually values packages. Flat discounts are wrong.
The core principle: secondary assets contribute diminishing value the further they fall below the lead asset.

RULES:
1. Identify the LEAD asset (highest value on that side)
2. For each additional asset, calculate its CONTRIBUTION based on its value relative to the lead:
   - Additional asset worth ≥ 70% of lead → contributes 65% of its face value
   - Additional asset worth 50-69% of lead → contributes 50% of its face value
   - Additional asset worth 30-49% of lead → contributes 35% of its face value
   - Additional asset worth < 30% of lead → contributes 20% of its face value
3. Package value = lead asset (100%) + sum of all additional asset contributions
4. Apply to EACH side independently

EXAMPLES:
  Chase (9,999) + Tate (5,985):
  Tate is 60% of Chase → contributes 50% → 2,993
  Package value = 9,999 + 2,993 = 12,992 (NOT 15,984)

  CeeDee (7,494) + 2027 1st (5,200) + Tate (5,985):
  Lead = CeeDee 7,494
  2027 1st is 69% of lead → contributes 50% → 2,600
  Tate is 80% of lead → contributes 65% → 3,890
  Package value = 7,494 + 2,600 + 3,890 = 13,984 (NOT 18,679)

  Three depth assets (2,000 + 1,800 + 1,500):
  Lead = 2,000
  1,800 is 90% of lead → contributes 65% → 1,170
  1,500 is 75% of lead → contributes 65% → 975
  Package value = 2,000 + 1,170 + 975 = 4,145 (NOT 5,300)
  Note: three depth players still can't match one elite player

Always show full calculation so the math is transparent.
NEVER just add up face values on a multi-asset side.

STEP 4 — VERDICT (ACCEPT / DECLINE / COUNTER):
State verdict first, one word, in caps.
Support with: adjusted value gap | roster fit for {league} strategy | counterparty motivation.

STEP 5 — COUNTER (only if COUNTER verdict):
ONE primary counter only.
ONLY use assets from MY confirmed rosters (visible in screenshot or in system prompt roster sections).
Counter must pass OTHER SIDE TEST: would a rational GM accept this?
If you cannot construct a realistic counter from confirmed assets, say so explicitly rather than inventing one.
One Sleeper message: under 20 words, direct, addresses what THEY want.

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━"""

    content_parts = []
    for img_data in images:
        b64 = img_data.split(',')[1] if ',' in img_data else img_data
        media_type = 'image/png' if 'png' in img_data[:30].lower() else 'image/jpeg'
        content_parts.append({
            "type": "image",
            "source": {"type": "base64", "media_type": media_type, "data": b64}
        })

    content_parts.append({"type": "text", "text": prompt_text or
        f"Analyze this trade for {league}. Follow the 5-step protocol."})

    try:
        response = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=1800,
            system=system,
            tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 4}],
            messages=[{"role": "user", "content": content_parts}]
        )
        analysis = ""
        for block in response.content:
            if hasattr(block, 'text'):
                analysis += block.text
        return jsonify({"success": True, "analysis": analysis})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route('/api/trade/analyze_block', methods=['POST'])
def analyze_trade_block():
    """Analyze a manager's trade block — why they're selling, fit, proposed offers."""
    data    = request.json
    images  = data.get('images', [])[:6]
    prompt  = data.get('prompt', '')
    league  = data.get('league', '')
    manager = data.get('manager', '')

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("""SELECT player_name, my_value, ktc_value, delta, tier, position
                 FROM player_values ORDER BY my_value DESC LIMIT 150""")
    value_rows = c.fetchall()
    conn.close()

    value_ref = "MY PLAYER VALUES (for offer building):\n"
    for name, my_val, ktc_val, delta, tier, pos in value_rows:
        d = f"+{delta}" if delta >= 0 else str(delta)
        value_ref += f"{name} ({pos}): MY {my_val:,} KTC {ktc_val:,} Δ{d}\n"

    league_contexts = {
        "Velvet Spade":        "12-team SuperFlex | 1.5x TE | 6pt TD | 2027 window | 4x 2027 1sts",
        "Capital Gains":       "FFPC #430 | SuperFlex | Rebuild | 4x 2027 1sts",
        "Twenty Run Savages":  "FFPC #210 | SuperFlex | Competing | Drake Maye core",
        "Gentleman's Dynasty": "Sleeper | Rebuild | Mahomes+Bowers untouchable",
    }
    league_ctx = league_contexts.get(league, league)

    # Get full league roster context so model knows who owns what
    league_roster_ctx = get_league_roster_context(league)
    if not league_roster_ctx:
        league_roster_ctx = ("\n⚠ LEAGUE ROSTER DATA NOT SYNCED — Go to Rosters tab → Sync Sleeper "
                             "to load full VS/GL manager rosters for accurate trade block analysis.")

    _vs = get_full_league_context('velvet_spade')
    _gl = get_full_league_context('gentlemans_dynasty')
    _sp = (SYSTEM_PROMPT
        .replace('{VS_ROSTER_BLOCK}',  _vs['my_roster'])
        .replace('{GL_ROSTER_BLOCK}',  _gl['my_roster'])
        .replace('{CG_ROSTER_BLOCK}',  get_roster_block('Capital Gains'))
        .replace('{TRS_ROSTER_BLOCK}', get_roster_block('TRS')))
    if _vs['all_rosters']: _sp += _vs['all_rosters']
    if _gl['all_rosters']: _sp += _gl['all_rosters']
    system = f"""{_sp}

{value_ref}

LEAGUE: {league} | {league_ctx}
{f'MANAGER BEING ANALYZED: {manager}' if manager else ''}
{league_roster_ctx}

TRADE BLOCK RULES:
1. FIRST: Resolve every player name in the screenshot against the league roster data above.
   Use the roster table to identify which manager owns each player — this tells you WHO is selling.
   Common abbreviations: "D Jones" = Daniel Jones, "K Williams RB LAR" = Kyren Williams, etc.
2. Web search each player's current NFL situation before any claim. Flag ⚠ UNVERIFIED if not confirmed.
3. WHY SELLING: read the manager's full roster from the league data above — rebuilding, selling aging vets, window mismatch, needs picks? Give a specific read based on THEIR actual roster.
4. FIT: direct assessment — STRONG FIT / NEUTRAL / POOR FIT with one-line reason tied to my {league} strategy. One sentence for POOR FIT players, do not over-analyze.
5. OFFERS: KTC-style dynamic package valuation. Lead asset contributes 100%. Each additional asset contributes based on its % of lead value (≥70%→65%, 50-69%→50%, 30-49%→35%, <30%→20%). Only use my confirmed {league} roster assets. Must pass other side test — would they actually accept?
6. ONE priority recommendation at the end.
7. NEVER reference players, managers, or picks from other leagues ({', '.join([l for l in ['Capital Gains','Twenty Run Savages',"Gentleman's Dynasty",'Velvet Spade'] if l != league])})."""

    content_parts = []
    for img_data in images:
        b64 = img_data.split(',')[1] if ',' in img_data else img_data
        media_type = 'image/png' if 'png' in img_data[:30].lower() else 'image/jpeg'
        content_parts.append({"type": "image",
            "source": {"type": "base64", "media_type": media_type, "data": b64}})
    content_parts.append({"type": "text", "text": prompt})

    try:
        response = client.messages.create(
            model="claude-sonnet-4-5", max_tokens=2000, system=system,
            tools=[{"type": "web_search_20250305", "name": "web_search", "max_uses": 5}],
            messages=[{"role": "user", "content": content_parts}]
        )
        analysis = "".join(b.text for b in response.content if hasattr(b, 'text'))
        return jsonify({"success": True, "analysis": analysis})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route('/api/trade/evaluate', methods=['POST'])
def trade_evaluate():
    """Evaluate a trade using Claude with full league context and KTC values"""
    data = request.json
    league = data.get('league', '')
    giving = data.get('giving', [])
    receiving = data.get('receiving', [])

    if not giving and not receiving:
        return jsonify({"success": False, "error": "No assets provided"})

    # Look up KTC values for mentioned players
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    ktc_context = {}
    for asset in giving + receiving:
        norm = re.sub(r"[^a-z0-9]", "", asset.lower())
        c.execute("SELECT player_name, value FROM market_data WHERE source='ktc' ORDER BY rank LIMIT 200")
        all_ktc = c.fetchall()
        for name, val in all_ktc:
            if re.sub(r"[^a-z0-9]", "", name.lower()) == norm:
                ktc_context[asset] = val
                break
    conn.close()

    ktc_str = "\n".join([f"  {k}: {v:,}" for k, v in ktc_context.items()]) if ktc_context else "  (No KTC data found — use your judgment)"

    prompt = f"""Evaluate this dynasty fantasy football trade for {league}.

GIVING: {', '.join(giving)}
RECEIVING: {', '.join(receiving)}

KTC VALUES (SuperFlex+TE):
{ktc_str}

Provide:
1. VERDICT: ACCEPT / DECLINE / COUNTER
2. value_giving: total KTC value of assets given
3. value_receiving: total KTC value of assets received  
4. surplus_pct: percentage surplus (positive = receiving more)
5. analysis: 2-3 sentences on the deal
6. my_perspective: how this fits MJBrutus's roster/strategy
7. their_perspective: why the other side takes this
8. trade_message: under 20 words to send the offer
9. counter: if declining, what counter to propose

Respond as JSON only. No markdown, no preamble."""

    try:
        response = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=800,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": prompt}]
        )
        text = response.content[0].text.strip()
        start = text.find('{')
        end = text.rfind('}') + 1
        if start >= 0 and end > start:
            result = json.loads(text[start:end])
            return jsonify({"success": True, "result": result})
        return jsonify({"success": False, "error": "Could not parse response"})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})


@app.route('/api/trade/analyze_negotiation', methods=['POST'])
def analyze_negotiation():
    """
    Analyze trade negotiation screenshots.
    Accepts up to 5 images, reads the conversation,
    and returns strategic analysis of partner positioning.
    """
    data = request.json
    images = data.get('images', [])[:5]
    context = data.get('context', '')
    league = data.get('league', '')

    if not images:
        return jsonify({"success": False, "error": "No images provided"})

    # Build multi-image message content
    content_parts = []
    for i, img_data in enumerate(images):
        img_b64 = img_data.split(',')[1] if ',' in img_data else img_data
        media_type = 'image/png' if 'png' in img_data[:30] else 'image/jpeg'
        content_parts.append({
            "type": "image",
            "source": {"type": "base64", "media_type": media_type, "data": img_b64}
        })

    prompt = f"""You are analyzing trade negotiation screenshots from a dynasty fantasy football chat.
League: {league or 'Dynasty league'}
Additional context: {context or 'None provided'}

These {len(images)} screenshot(s) show trade negotiation messages between MJBrutus and a trade partner.

Analyze:
1. PARTNER POSITIONING — What does the partner want? What are their pain points? Are they motivated sellers or just fishing?
2. OFFER ASSESSMENT — What was offered, counter-offered? Who has leverage?
3. NEGOTIATION SIGNALS — Any tells? Urgency? Reluctance? Specific language that indicates willingness to deal?
4. RECOMMENDED STRATEGY — Exactly how should MJBrutus respond? What counter-offer, what framing, what to emphasize?
5. DRAFT A RESPONSE — Write a ready-to-send reply (under 40 words) that advances MJBrutus's position.

Be direct and specific. Reference actual players/picks mentioned in the screenshots."""

    content_parts.append({"type": "text", "text": prompt})

    try:
        response = client.messages.create(
            model="claude-sonnet-4-5",
            max_tokens=1200,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": content_parts}]
        )
        analysis = response.content[0].text
        return jsonify({"success": True, "analysis": analysis})
    except Exception as e:
        return jsonify({"success": False, "error": str(e)})



if __name__ == '__main__':
    port = int(os.environ.get('PORT', 8080))
    app.run(host='0.0.0.0', port=port, debug=False)
@app.route('/api/board/debug', methods=['GET'])
def board_debug():
    """Quick debug: shows what the board sees without full roster fetch."""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM player_values WHERE ktc_value > 0")
    val_count = c.fetchone()[0]
    c.execute("SELECT COUNT(*) FROM market_data WHERE source='ktc'")
    ktc_count = c.fetchone()[0]
    c.execute("SELECT player_name, ktc_value FROM player_values ORDER BY ktc_value DESC LIMIT 5")
    top5 = c.fetchall()
    conn.close()
    return jsonify({
        "player_values_with_ktc": val_count,
        "market_data_ktc": ktc_count,
        "top5_players": [{"name": r[0], "ktc": r[1]} for r in top5],
        "has_values": val_count > 0 or ktc_count > 0
    })



@app.route('/api/board/<league_key>', methods=['GET'])
def get_board(league_key):
    try:
        return _get_board_inner(league_key)
    except Exception as e:
        import traceback
        return jsonify({"success": False, "error": str(e), "trace": traceback.format_exc()[-500:]})


def _get_board_inner(league_key):
    # Resolve display name → sleeper key
    name_map = {
        'velvet_spade': 'velvet_spade', 'gentlemans_dynasty': 'gentlemans_dynasty',
        'vs': 'velvet_spade', 'gl': 'gentlemans_dynasty',
    }
    sleeper_key = name_map.get(league_key.lower())

    # Load my values for valuation
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT player_name, my_value, ktc_value, position FROM player_values")
    pv = c.fetchall()
    c.execute("SELECT player_name, value FROM market_data WHERE source='ktc'")
    md = c.fetchall()
    conn.close()

    def norm(s):
        import re as _re
        # Remove punctuation, lowercase
        s = _re.sub(r"[^a-z0-9\s]", "", s.lower()).strip()
        # Normalize suffixes: jr, sr, ii, iii, iv
        s = _re.sub(r'\s+(jr|sr|ii|iii|iv)$', '', s)
        return _re.sub(r'\s+', '', s)  # remove all spaces

    val_exact = {r[0].lower(): r[1] for r in pv}      # my_value
    val_norm  = {norm(r[0]): r[1] for r in pv}
    ktc_exact = {r[0].lower(): r[1] for r in pv}
    ktc_norm  = {norm(r[0]): r[1] for r in pv}
    md_exact  = {r[0].lower(): r[1] for r in md}
    md_norm   = {norm(r[0]): r[1] for r in md}

    def value_of(name):
        nl, nn = name.lower(), norm(name)
        if nl in val_exact: return val_exact[nl]
        if nn in val_norm:  return val_norm[nn]
        if nl in md_exact:  return md_exact[nl]
        if nn in md_norm:   return md_norm[nn]
        return 0

    # Get rosters — Sleeper live or DB for FFPC
    managers = []  # list of {manager, is_me, players:[{name,pos,team,age,value}]}

    if sleeper_key:
        league_id = SLEEPER_LEAGUE_IDS.get(sleeper_key)
        try:
            players_db = requests.get("https://api.sleeper.app/v1/players/nfl", timeout=20).json()
            users      = requests.get(f"https://api.sleeper.app/v1/league/{league_id}/users", timeout=10).json()
            rosters    = requests.get(f"https://api.sleeper.app/v1/league/{league_id}/rosters", timeout=10).json()
            user_map   = {u['user_id']: u.get('display_name', u.get('username','?')) for u in users}
            my_id = next((u['user_id'] for u in users
                          if u.get('username','').lower()=='dcatlet'
                          or u.get('display_name','').lower()=='dcatlet'), None)

            for r in rosters:
                oid = r.get('owner_id','')
                mgr = user_map.get(oid, f"Team{r['roster_id']}")
                plist = []
                for pid in (r.get('players',[]) or []):
                    p = players_db.get(pid, {})
                    name = f"{p.get('first_name','')} {p.get('last_name','')}".strip()
                    if not name: continue
                    plist.append({
                        'name': name, 'pos': p.get('position','?'),
                        'team': p.get('team','FA') or 'FA',
                        'age':  p.get('age'),
                        'value': value_of(name),
                    })
                managers.append({'manager': mgr, 'is_me': oid==my_id, 'players': plist})
        except Exception as e:
            return jsonify({"success": False, "error": f"Sleeper fetch failed: {e}"})
    else:
        # FFPC — from DB
        db_league = 'Capital Gains' if 'capital' in league_key.lower() or league_key.lower()=='cg' else 'TRS'
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT manager, player_name, position, team FROM league_rosters WHERE league=?", (db_league,))
        rows = c.fetchall()
        conn.close()
        from collections import defaultdict
        by_mgr = defaultdict(list)
        for mgr, name, pos, team in rows:
            by_mgr[mgr].append({'name':name,'pos':pos,'team':team,'age':None,'value':value_of(name)})
        my_clean = {'capitalgains','twentyrunsavages','dcatlet','mjbrutus'}
        for mgr, plist in by_mgr.items():
            is_me = mgr.replace('←','').replace('→','').strip().lower().replace(' ','') in my_clean
            managers.append({'manager': mgr, 'is_me': is_me, 'players': plist})

    if not managers:
        return jsonify({"success": False, "error": "No roster data. Sync Sleeper or upload spreadsheet."})

    # Add draft pick capital — reuse already-fetched rosters/users, just add traded_picks call
    if sleeper_key:
        try:
            league_id = SLEEPER_LEAGUE_IDS.get(sleeper_key)
            picks_r = requests.get(
                f"https://api.sleeper.app/v1/league/{league_id}/traded_picks", timeout=10)

            if picks_r.status_code == 200 and managers:
                traded_picks_raw = picks_r.json()

                # Build roster_id → manager name from already-fetched data
                # managers list has 'manager' names; we need roster_id → mgr mapping
                # Re-use the rosters variable from above (still in scope)
                rid_to_mgr = {}
                for r in rosters:
                    oid = r.get('owner_id', '')
                    mgr_name = user_map.get(oid, f"Team{r['roster_id']}")
                    rid_to_mgr[str(r['roster_id'])] = mgr_name

                # Simple approach: traded_picks tells us the CURRENT state
                # owner_id = who owns it now (roster_id)
                # previous_owner_id = who had it before (roster_id)
                # roster_id = original team it belongs to (roster_id)
                # season = draft year, round = pick round

                # Start with each manager owning all their own picks
                # then override with what traded_picks says
                mgr_picks = defaultdict(set)

                # Get future seasons from traded picks
                future_seasons = sorted(set(
                    tp.get('season','') for tp in traded_picks_raw
                    if tp.get('season','')
                ))

                for season in future_seasons:
                    # Every manager starts with their own R1-R7
                    for r in rosters:
                        mgr = rid_to_mgr.get(str(r['roster_id']), f"Team{r['roster_id']}")
                        for rnd in range(1, 8):
                            mgr_picks[mgr].add(f"{season}|R{rnd}|{mgr}|own")

                    # Apply trades: remove from previous owner, add to current owner
                    for tp in traded_picks_raw:
                        if tp.get('season') != season: continue
                        rnd  = tp.get('round')
                        orig_rid  = str(tp.get('roster_id', ''))
                        owner_rid = str(tp.get('owner_id', ''))
                        prev_rid  = str(tp.get('previous_owner_id', ''))

                        orig_mgr  = rid_to_mgr.get(orig_rid, orig_rid)
                        owner_mgr = rid_to_mgr.get(owner_rid, owner_rid)
                        prev_mgr  = rid_to_mgr.get(prev_rid, prev_rid)

                        # Remove this pick from previous owner (regardless of label)
                        mgr_picks[prev_mgr] = {
                            p for p in mgr_picks[prev_mgr]
                            if not (p.startswith(f"{season}|R{rnd}|{orig_mgr}|"))
                        }
                        # Add to current owner
                        label = 'own' if owner_mgr == orig_mgr else f'from {orig_mgr}'
                        mgr_picks[owner_mgr].add(f"{season}|R{rnd}|{orig_mgr}|{label}")

                # Convert to sorted display strings and attach to managers
                for m in managers:
                    raw_picks = sorted(mgr_picks.get(m['manager'], []))
                    display = []
                    for p in raw_picks:
                        parts = p.split('|')
                        if len(parts) == 4:
                            season, rnd, orig, label = parts
                            display.append(
                                f"{season} {rnd}" + ('' if label == 'own' else f" ({label})")
                            )
                    m['picks'] = display

        except Exception as ex:
            for m in managers:
                if 'picks' not in m:
                    m['picks'] = []
    else:
        for m in managers:
            m['picks'] = []

    # Compute window + needs for each manager
    # Thresholds: top starter quality in dynasty SF TE+ context
    STARTER_THRESHOLDS = {'QB': 3500, 'RB': 2500, 'WR': 2500, 'TE': 2500}
    ELITE_THRESHOLD = 5500  # true elite asset
    for m in managers:
        pos_players = defaultdict(list)
        for p in m['players']:
            if p['pos'] in ('QB','RB','WR','TE'):
                pos_players[p['pos']].append(p['value'])
        for pos in pos_players:
            pos_players[pos].sort(reverse=True)

        total_value = sum(p['value'] for p in m['players'])
        top_tier    = sum(1 for p in m['players'] if p['value'] >= ELITE_THRESHOLD)
        young_studs = sum(1 for p in m['players']
                          if p['value'] >= 4000 and (p['age'] is None or p['age'] <= 25))
        aging_vets  = sum(1 for p in m['players']
                          if p['value'] >= 3000 and p['age'] is not None and p['age'] >= 28)

        # Window: use relative ranking if most values are 0
        valued_players = [p for p in m['players'] if p['value'] > 0]
        if len(valued_players) < 5:
            # Fall back to raw count-based inference
            window = 'Unknown'
        elif young_studs >= 3 and top_tier >= 2:
            window = 'Contending'
        elif aging_vets >= 3 and young_studs <= 1:
            window = 'Win-Now'
        elif top_tier <= 1 and young_studs <= 1:
            window = 'Rebuilding'
        else:
            window = 'Transitioning'

        # Needs/surplus — only compute if we have enough value data
        needs, surplus = [], []
        if len(valued_players) >= 5:
            for pos, thresh in STARTER_THRESHOLDS.items():
                vals        = pos_players.get(pos, [])
                starters    = [v for v in vals if v >= thresh]
                need_count  = 2 if pos in ('RB','WR') else 1
                if len(starters) < need_count:
                    needs.append(pos)
                elif len(starters) >= need_count + 2:
                    surplus.append(pos)

        m['window']      = window
        m['needs']       = needs
        m['surplus']     = surplus
        m['total_value'] = total_value
        m['players'].sort(key=lambda p: -p['value'])

    # Sort: me first, then by total value
    managers.sort(key=lambda m: (not m['is_me'], -m['total_value']))

    return jsonify({"success": True, "league": league_key, "managers": managers})



    """
    Single Sleeper API call that returns BOTH:
    - my_roster: dcatlet's position-grouped roster string
    - all_rosters: all 12 manager rosters for full league context
    Uses 3 API calls total (players, users, rosters) — reused for both outputs.
    Falls back gracefully if API unavailable.
    """
    empty = {'my_roster': f'[{league_key} roster not available]', 'all_rosters': ''}

    league_id = SLEEPER_LEAGUE_IDS.get(league_key)
    if not league_id:
        return empty

    try:
        players_r = requests.get("https://api.sleeper.app/v1/players/nfl", timeout=20)
        users_r   = requests.get(f"https://api.sleeper.app/v1/league/{league_id}/users", timeout=10)
        rosters_r = requests.get(f"https://api.sleeper.app/v1/league/{league_id}/rosters", timeout=10)

        if any(r.status_code != 200 for r in [players_r, users_r, rosters_r]):
            return empty

        players_db = players_r.json()
        users      = users_r.json()
        rosters    = rosters_r.json()
        user_map   = {u['user_id']: u.get('display_name', u.get('username','?')) for u in users}

        my_id = next(
            (u['user_id'] for u in users
             if u.get('username','').lower() == 'dcatlet'
             or u.get('display_name','').lower() == 'dcatlet'),
            None)

        pos_order = {'QB':0,'RB':1,'WR':2,'TE':3,'K':4,'DEF':5}
        my_roster_lines  = []
        all_roster_lines = [f"\n{league_key.upper().replace('_',' ')} FULL LEAGUE ROSTERS (live):"]

        for roster in sorted(rosters, key=lambda r: r['roster_id']):
            owner_id = roster.get('owner_id','')
            mgr      = user_map.get(owner_id, f"Team{roster['roster_id']}")
            is_me    = (owner_id == my_id)

            by_pos = defaultdict(list)
            for pid in (roster.get('players',[]) or []):
                p    = players_db.get(pid, {})
                name = f"{p.get('first_name','')} {p.get('last_name','')}".strip()
                pos  = p.get('position','?')
                team = p.get('team','FA') or 'FA'
                ktc  = p.get('search_rank', 999)  # use search_rank as proxy for value sorting
                if name and pos in ('QB','RB','WR','TE','K'):
                    by_pos[pos].append((name, team, ktc))

            # Sort each position by value (search_rank ascending = higher value first)
            for pos in by_pos:
                by_pos[pos].sort(key=lambda x: x[2])

            # Build compact all-roster line
            parts = []
            for pos in ['QB','RB','WR','TE']:
                if by_pos[pos]:
                    parts.append(f"{pos}:{','.join(n for n,t,_ in by_pos[pos])}")
            prefix = "★MY TEAM " if is_me else ""
            all_roster_lines.append(f"{prefix}{mgr}: {' | '.join(parts)}")

            # Build my full roster block
            if is_me:
                total = sum(len(v) for v in by_pos.values())
                my_roster_lines.append(
                    f"MY {league_key.upper().replace('_',' ')} ROSTER — {total} players (live):")
                for pos in ['QB','RB','WR','TE','K']:
                    if by_pos[pos]:
                        players_str = ' | '.join(f"{n}({t})" for n,t,_ in by_pos[pos])
                        my_roster_lines.append(f"{pos}: {players_str}")

        # Pull traded picks — who owns what
        try:
            picks_r = requests.get(
                f"https://api.sleeper.app/v1/league/{league_id}/traded_picks", timeout=10)
            if picks_r.status_code == 200:
                picks_raw = picks_r.json()
                # Group picks by current owner
                my_picks = []
                all_pick_lines = [f"\n{league_key.upper().replace('_',' ')} TRADED PICKS (current ownership):"]
                owner_picks = defaultdict(list)
                for pk in picks_raw:
                    owner_id = pk.get('owner_id','')
                    prev_id  = pk.get('previous_owner_id','')
                    rnd  = pk.get('round','?')
                    yr   = pk.get('season','')
                    owner_name = user_map.get(owner_id, owner_id)
                    orig_name  = user_map.get(prev_id, prev_id)
                    label = f"{yr} R{rnd}" + (f" (orig: {orig_name})" if prev_id != owner_id else '')
                    owner_picks[owner_name].append(label)
                    if owner_id == my_id:
                        my_picks.append(label)
                # Add my picks to my_roster block
                if my_picks:
                    my_roster_lines.append(f"PICKS OWNED: {' | '.join(sorted(my_picks))}")
                # Add all pick ownership to all_rosters context
                for owner, picks in sorted(owner_picks.items()):
                    all_pick_lines.append(f"{owner}: {', '.join(sorted(picks))}")
                all_roster_lines.extend(all_pick_lines)
        except Exception:
            pass  # picks are supplementary — don't fail if unavailable

        return {
            'my_roster':   '\n'.join(my_roster_lines) if my_roster_lines else empty['my_roster'],
            'all_rosters': '\n'.join(all_roster_lines),
            'my_picks':    ' | '.join(sorted(my_picks)) if my_picks else '[no traded picks — own standard R1-R7]',
        }

    except Exception as e:
        return {'my_roster': f'[{league_key}: {str(e)[:60]}]', 'all_rosters': ''}


def get_roster_block(league_key, fallback=''):
    """
    Get dcatlet's roster for a league.
    Sleeper leagues: fetch live from API (always current).
    FFPC leagues: read from DB (populated via KB import).
    """
    SLEEPER_LEAGUES = {'velvet_spade', 'gentlemans_dynasty'}

    if league_key in SLEEPER_LEAGUES:
        # Fetch live from Sleeper — same approach as Rosters tab
        league_id = SLEEPER_LEAGUE_IDS.get(league_key)
        if not league_id:
            return fallback or f"[{league_key}: no league ID configured]"
        try:
            players_r = requests.get("https://api.sleeper.app/v1/players/nfl", timeout=20)
            users_r   = requests.get(f"https://api.sleeper.app/v1/league/{league_id}/users", timeout=10)
            rosters_r = requests.get(f"https://api.sleeper.app/v1/league/{league_id}/rosters", timeout=10)

            if any(r.status_code != 200 for r in [players_r, users_r, rosters_r]):
                return fallback or f"[{league_key}: Sleeper API unavailable]"

            players_db = players_r.json()
            users      = users_r.json()
            rosters    = rosters_r.json()

            my_id = next(
                (u['user_id'] for u in users
                 if u.get('username','').lower() == 'dcatlet'
                 or u.get('display_name','').lower() == 'dcatlet'),
                None)

            my_roster = next(
                (r for r in rosters if r.get('owner_id') == my_id), None)

            if not my_roster:
                return fallback or f"[{league_key}: could not find dcatlet roster]"

            by_pos = defaultdict(list)
            for pid in (my_roster.get('players',[]) or []):
                p = players_db.get(pid, {})
                name = f"{p.get('first_name','')} {p.get('last_name','')}".strip()
                pos  = p.get('position','?')
                team = p.get('team','FA') or 'FA'
                if pos in ('QB','RB','WR','TE','K') and name:
                    by_pos[pos].append(f"{name}({team})")

            total = sum(len(v) for v in by_pos.values())
            lines = [f"MY {league_key.upper().replace('_',' ')} ROSTER — {total} players (live):"]
            for pos in ['QB','RB','WR','TE','K']:
                if by_pos[pos]:
                    lines.append(f"{pos}: {' | '.join(by_pos[pos])}")
            return '\n'.join(lines)

        except Exception as e:
            return fallback or f"[{league_key}: API error — {str(e)[:50]}]"

    else:
        # FFPC leagues — read from DB (populated via KB spreadsheet import)
        MY_NAMES = {'dcatlet', 'mjbrutus', 'capital gains', 'twenty run savages'}

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute("SELECT DISTINCT manager FROM league_rosters WHERE league=?", (league_key,))
        all_managers = [r[0] for r in c.fetchall()]

        def clean(s):
            return s.replace('\u2190','').replace('\u2192','').replace('←','').replace('→','').strip().lower().replace(' ','')

        my_manager = next(
            (m for m in all_managers if clean(m) in {n.replace(' ','') for n in MY_NAMES}),
            None)

        if not my_manager:
            conn.close()
            return fallback or f"[{league_key}: upload spreadsheet to load FFPC roster]"

        c.execute("""SELECT position, player_name, team FROM league_rosters
                     WHERE league=? AND manager=?
                     ORDER BY CASE position
                       WHEN 'QB' THEN 1 WHEN 'RB' THEN 2
                       WHEN 'WR' THEN 3 WHEN 'TE' THEN 4 ELSE 5 END, player_name""",
                  (league_key, my_manager))
        rows = c.fetchall()
        conn.close()

        if not rows:
            return fallback or f"[{league_key}: no roster data — upload spreadsheet]"

        by_pos = defaultdict(list)
        for pos, name, team in rows:
            by_pos[pos].append(f"{name}({team})" if team else name)

        total = sum(len(v) for v in by_pos.values())
        lines = [f"MY {league_key.upper().replace('_',' ')} ROSTER — {total} players:"]
        for pos in ['QB','RB','WR','TE','K']:
            if by_pos[pos]:
                lines.append(f"{pos}: {' | '.join(by_pos[pos])}")
        return '\n'.join(lines)


def get_week_key():
    now = datetime.now()
    week_start = now - timedelta(days=now.weekday())
    return week_start.strftime("%Y-W%U")
